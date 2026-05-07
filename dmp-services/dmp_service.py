import asyncio
import logging
import math
import os
import re
import shutil
import socket
import tempfile
import threading
import time
from contextlib import asynccontextmanager, contextmanager
from datetime import date
from io import BytesIO
from pathlib import Path

import pyodbc
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from typing import Optional

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


DMP_STATION_NAME: str = os.environ.get("DMP_STATION_NAME", "").strip()
VONIKO_SERVER_URL: str = os.environ.get("VONIKO_SERVER_URL", "").rstrip("/")
DMP_STATION_PORT: int = int(os.environ.get("DMP_STATION_PORT", "8766"))
DMP_DATA_DIR: str = os.environ.get("DMP_DATA_DIR", r"C:\DMP\Data")
DMP_TEMPLATES_DIR: str = os.environ.get("DMP_TEMPLATES_DIR", "./dmp_templates")
DM2000_DATA_DIR: str = os.environ.get("DM2000_DATA_DIR", r"D:\DM2000\dmdatabase")
DM2000_TEMPLATES_DIR: str = os.environ.get("DM2000_TEMPLATES_DIR", "./dm2000_templates")
DM2000_PERF_TEMPLATES_DIR: str = os.environ.get("DM2000_PERF_TEMPLATES_DIR", "./dm2000_perf_templates")
DMP_PERF_TEMPLATES_DIR: str = os.environ.get("DMP_PERF_TEMPLATES_DIR", "./dmp_perf_templates")
# Local cache directory: persistent copies of dmdata_ls.mdb are kept here so that
# every request reads from the cached copy instead of making a new shadow copy.
DM2000_CACHE_DIR: str = os.environ.get("DM2000_CACHE_DIR", "../backend/data/dm2000_cache")
# Local cache directory: persistent copy of DMPDATA.mdb is kept here so that
# every batch/channel request reads from the cached copy instead of making a
# new shadow copy each time.
DMPDATA_CACHE_DIR: str = os.environ.get("DMPDATA_CACHE_DIR", "../backend/data/dmpdata_cache")
# Configurable company name shown in reports (e.g. "Asia Matsushita Electric Pte Ltd").
DM2000_COMPANY_NAME: str = os.environ.get("DM2000_COMPANY_NAME", "")
WATCH_INTERVAL_SECONDS: int = 5
# Maximum valid DM2000 battery channel number (channels are numbered 1..MAX_BATTERY_NUMBER)
MAX_BATTERY_NUMBER: int = 99
_EXCEL_MAX_SHEET_NAME: int = 31  # Excel sheet names are capped at 31 characters

_WATCH_LOCK = threading.Lock()
_ACCESS_QUERY_LOCK = threading.Semaphore(3)
_ACCESS_QUERY_TIMEOUT: float = 20.0  # seconds to wait for a DB slot before returning 503

# ── Persistent local cache for dmdata_ls.mdb ─────────────────────────────────
# Instead of creating a temporary shadow copy for every request (which caused
# resource exhaustion and 503 errors after ~90 s of load), we maintain ONE
# persistent local copy of the file.  A background watcher checks the source
# mtime every WATCH_INTERVAL_SECONDS and atomically replaces the cached file
# when the source has been updated.  All read operations use the cached path
# directly; no per-request file copy or concurrency lock is needed.
_DM2000_LS_CACHE_PATH: str = ""           # path to the current cached copy
_DM2000_LS_SOURCE_MTIME: float = -1.0    # mtime of the last successfully cached source
_DM2000_LS_CACHE_WRITE_LOCK = threading.Lock()   # one writer at a time
_DM2000_LS_CACHE_READY = threading.Event()        # set once the first copy is done
# Seconds to wait for the initial cache build before returning 503 to callers.
_DM2000_LS_CACHE_READY_TIMEOUT: float = 30.0

# ── Persistent local cache for DMPDATA.mdb ───────────────────────────────────
# DMPDATA.mdb accumulates data since 2019 and grows very large.  Maintaining a
# single persistent local copy eliminates the per-request shadow-copy overhead
# and lets the year-filter fast-path work entirely against the local file.
# The background watcher refreshes the copy whenever the source mtime changes.
_DMPDATA_CACHE_PATH: str = ""
_DMPDATA_SOURCE_MTIME: float = -1.0
_DMPDATA_CACHE_WRITE_LOCK = threading.Lock()
_DMPDATA_CACHE_READY = threading.Event()
_DMPDATA_CACHE_READY_TIMEOUT: float = 30.0
_TELEMETRY_CACHE: dict[tuple, tuple[list, float]] = {}
_TELEMETRY_CACHE_LOCK = threading.Lock()
_TELEMETRY_CACHE_TTL: float = 60.0  # seconds
_DM2000_CURVE_CACHE: dict[tuple, tuple[list, float]] = {}
_DM2000_CURVE_CACHE_LOCK = threading.Lock()
_DM2000_CURVE_CACHE_TTL: float = 300.0  # seconds
_DM2000_ARCHIVES_CACHE: dict[str, tuple[list, float]] = {}
_DM2000_ARCHIVES_CACHE_LOCK = threading.Lock()
_DM2000_ARCHIVES_CACHE_TTL: float = 60.0  # seconds
_DM2000_BATTERIES_CACHE: dict[str, tuple[list, float]] = {}
_DM2000_BATTERIES_CACHE_LOCK = threading.Lock()
_DM2000_BATTERIES_CACHE_TTL: float = 60.0  # seconds
_DM2000_STATS_CACHE: dict[tuple, tuple[dict, float]] = {}
_DM2000_STATS_CACHE_LOCK = threading.Lock()
_DM2000_STATS_CACHE_TTL: float = 300.0  # seconds
_DM2000_CACHE_MAX_ENTRIES: int = 100
_WATCHED_MDB_MTIME: dict[str, float] = {}
_WATCHED_CHANGES: dict[str, float] = {}
_SCHEMA_TABLE_WHITELIST = {
    "para_singl",
    "para_pub",
    "para_pur",
    "vidata",
    "ls_jb_cs",
    "ls_pam2",
    "ls_vtime",
    "ls_evolt",
    "ls_timev",
}
def _get_local_ip() -> str:
    try:
        host = VONIKO_SERVER_URL.split("://")[-1].split(":")[0].split("/")[0]
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect((host, 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def _registration_loop() -> None:
    """Background thread: register/heartbeat to Voniko server every 30s."""
    import requests as _req
    while True:
        try:
            ip = _get_local_ip()
            url = f"http://{ip}:{DMP_STATION_PORT}"
            _req.post(
                f"{VONIKO_SERVER_URL}/api/dmp/register",
                json={"name": DMP_STATION_NAME, "url": url},
                timeout=5,
            )
        except Exception:
            pass
        time.sleep(30)


def _scan_dynamic_mdb_files() -> dict[str, float]:
    data_dir = Path(DMP_DATA_DIR).resolve()
    if not data_dir.exists():
        return {}

    result: dict[str, float] = {}
    for entry in data_dir.glob("*.mdb"):
        if not entry.is_file():
            continue
        if entry.name.lower() == "dmpdata.mdb":
            continue
        try:
            result[entry.stem] = entry.stat().st_mtime
        except OSError:
            continue
    return result


def _watch_dmp_changes_loop() -> None:
    with _WATCH_LOCK:
        _WATCHED_MDB_MTIME.update(_scan_dynamic_mdb_files())

    while True:
        try:
            current = _scan_dynamic_mdb_files()
            now = time.time()
            with _WATCH_LOCK:
                deleted_stems = set(_WATCHED_MDB_MTIME.keys()) - set(current.keys())
                for stem in deleted_stems:
                    _WATCHED_CHANGES.pop(stem, None)
                    _WATCHED_MDB_MTIME.pop(stem, None)
                for stem, mtime in current.items():
                    previous = _WATCHED_MDB_MTIME.get(stem)
                    if previous is None or mtime > previous:
                        _WATCHED_CHANGES[stem] = now
                    _WATCHED_MDB_MTIME[stem] = mtime
            # Keep dm2000 and dmpdata local caches in sync with source files
            _dm2000_refresh_ls_cache()
            _dmpdata_refresh_cache()
        except (OSError, ValueError, pyodbc.Error, HTTPException) as exc:
            # Never let the watcher thread die — log and continue.
            logger.warning("watch_loop: unexpected error (will retry): %s", exc)
        time.sleep(WATCH_INTERVAL_SECONDS)


def _dm2000_refresh_ls_cache(force: bool = False) -> None:
    """Copy dmdata_ls.mdb to the local cache directory if the source has changed.

    Uses an atomic rename (os.replace) so readers always see a complete file.
    Invalidates all DM2000 in-memory caches after a successful refresh.
    If the source cannot be read but a previous cache exists, keeps using it.
    """
    global _DM2000_LS_CACHE_PATH, _DM2000_LS_SOURCE_MTIME  # noqa: PLW0603

    ls_path = Path(get_dm2000_ls_path())
    if not ls_path.exists():
        # Source not present; signal ready using live path as fallback if cache exists
        if not _DM2000_LS_CACHE_READY.is_set():
            _DM2000_LS_CACHE_READY.set()
        return

    try:
        current_mtime = ls_path.stat().st_mtime
    except OSError:
        if not _DM2000_LS_CACHE_READY.is_set():
            _DM2000_LS_CACHE_READY.set()
        return

    with _DM2000_LS_CACHE_WRITE_LOCK:
        if not force and current_mtime <= _DM2000_LS_SOURCE_MTIME and _DM2000_LS_CACHE_READY.is_set():
            return  # source unchanged

        cache_dir = Path(DM2000_CACHE_DIR).resolve()
        try:
            cache_dir.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            logger.warning("dm2000_cache: cannot create cache dir %s: %s", cache_dir, exc)
            if not _DM2000_LS_CACHE_READY.is_set():
                _DM2000_LS_CACHE_PATH = str(ls_path)
                _DM2000_LS_CACHE_READY.set()
            return

        cache_final = str(cache_dir / "dmdata_ls.mdb")
        cache_tmp = str(cache_dir / "dmdata_ls_new.mdb")
        try:
            shutil.copy2(str(ls_path), cache_tmp)
        except (OSError, PermissionError) as exc:
            logger.warning("dm2000_cache: copy failed (%s), falling back to live source: %s", ls_path, exc)
            try:
                os.unlink(cache_tmp)
            except OSError:
                pass
            # Fall back to the live source path so ODBC reads attempt to
            # pick up the latest data.  If the source is exclusively locked,
            # ODBC will also fail, but this is no worse than using a stale copy.
            preferred = str(ls_path)
            if not _DM2000_LS_CACHE_READY.is_set():
                _DM2000_LS_CACHE_PATH = preferred
                _DM2000_LS_CACHE_READY.set()
            else:
                # Always update the path and clear caches so the next ODBC
                # read picks up whatever data is available in preferred.
                _DM2000_LS_CACHE_PATH = preferred
                with _DM2000_ARCHIVES_CACHE_LOCK:
                    _DM2000_ARCHIVES_CACHE.clear()
                with _DM2000_BATTERIES_CACHE_LOCK:
                    _DM2000_BATTERIES_CACHE.clear()
                with _DM2000_STATS_CACHE_LOCK:
                    _DM2000_STATS_CACHE.clear()
                with _DM2000_CURVE_CACHE_LOCK:
                    _DM2000_CURVE_CACHE.clear()
            return

        # Validate magic bytes before replacing.
        # The first 4 bytes of a valid Jet/Access .mdb file are 0x00 0x01 0x00 0x00
        # (Jet3/Jet4 format identifier).  This prevents caching a partially-written
        # or corrupt file.
        try:
            size = os.path.getsize(cache_tmp)
            if size >= 32 * 1024:
                with open(cache_tmp, "rb") as fh:
                    magic = fh.read(4)
                if magic != b"\x00\x01\x00\x00":
                    raise ValueError(f"Not a valid Access DB (magic={magic!r})")
        except (OSError, ValueError) as exc:
            logger.warning("dm2000_cache: validation failed for %s: %s", cache_tmp, exc)
            try:
                os.unlink(cache_tmp)
            except OSError:
                pass
            # Fall back to the live source path so ODBC reads are never
            # blocked by a partially-written or corrupt temporary file.
            preferred = str(ls_path)
            if not _DM2000_LS_CACHE_READY.is_set():
                _DM2000_LS_CACHE_PATH = preferred
                _DM2000_LS_CACHE_READY.set()
            else:
                _DM2000_LS_CACHE_PATH = preferred
                with _DM2000_ARCHIVES_CACHE_LOCK:
                    _DM2000_ARCHIVES_CACHE.clear()
                with _DM2000_BATTERIES_CACHE_LOCK:
                    _DM2000_BATTERIES_CACHE.clear()
                with _DM2000_STATS_CACHE_LOCK:
                    _DM2000_STATS_CACHE.clear()
                with _DM2000_CURVE_CACHE_LOCK:
                    _DM2000_CURVE_CACHE.clear()
            return

        # Atomically replace the cached copy.  On Windows a concurrent pyodbc
        # reader may hold cache_final open, causing os.replace to fail with
        # PermissionError.  Retry a few times with a short sleep to let
        # short-lived ODBC connections finish before giving up.
        replace_exc: "OSError | None" = None
        for attempt in range(6):
            try:
                os.replace(cache_tmp, cache_final)
                replace_exc = None
                break
            except OSError as exc:
                replace_exc = exc
                if attempt < 5:
                    time.sleep(0.4)
        if replace_exc is not None:
            # cache_final is held open by a process (e.g. Access).  Decide the
            # best read target without overwriting user edits.
            #
            # shutil.copy2 preserves the source file's mtime, so cache_tmp.mtime
            # equals the source's current mtime (current_mtime).  If cache_final
            # has a *newer* mtime it means the user edited it directly in Access
            # (e.g. adding remarks), and we must NOT replace it with the older
            # source copy.  In that case keep cache_final as the read target.
            #
            # Otherwise fall back to the live source path as before.  We do NOT
            # keep cache_tmp in this branch: the background watcher also writes
            # to that same path, so using it as a live cache could cause
            # concurrent-write races; the source file is a safer choice.
            cache_final_path = Path(cache_final)
            cache_final_mtime = 0.0
            if cache_final_path.exists():
                try:
                    cache_final_mtime = cache_final_path.stat().st_mtime
                except OSError:
                    pass

            try:
                os.unlink(cache_tmp)
            except OSError:
                pass

            if cache_final_mtime > current_mtime:
                # User edited the cache file; keep it as the read target.
                logger.warning(
                    "dm2000_cache: rename failed (file in use?), keeping user-edited"
                    " cache_final (mtime %.0f > source %.0f): %s",
                    cache_final_mtime,
                    current_mtime,
                    replace_exc,
                )
                preferred = cache_final
            else:
                # Source is newer or cache_final absent; fall back to source.
                logger.warning(
                    "dm2000_cache: rename failed (file in use?), falling back to"
                    " source: %s",
                    replace_exc,
                )
                preferred = str(ls_path)

            if not _DM2000_LS_CACHE_READY.is_set():
                _DM2000_LS_CACHE_PATH = preferred
                _DM2000_LS_CACHE_READY.set()
            else:
                # Always update the path and clear caches so the next ODBC read
                # sees fresh data, even if preferred equals the previous path.
                _DM2000_LS_CACHE_PATH = preferred
                with _DM2000_ARCHIVES_CACHE_LOCK:
                    _DM2000_ARCHIVES_CACHE.clear()
                with _DM2000_BATTERIES_CACHE_LOCK:
                    _DM2000_BATTERIES_CACHE.clear()
                with _DM2000_STATS_CACHE_LOCK:
                    _DM2000_STATS_CACHE.clear()
                with _DM2000_CURVE_CACHE_LOCK:
                    _DM2000_CURVE_CACHE.clear()
            return
        _DM2000_LS_CACHE_PATH = cache_final
        _DM2000_LS_SOURCE_MTIME = current_mtime
        logger.info("dm2000_cache: refreshed local copy from %s (mtime=%s)", ls_path, current_mtime)

        # Invalidate all DM2000 in-memory caches
        with _DM2000_CURVE_CACHE_LOCK:
            _DM2000_CURVE_CACHE.clear()
        with _DM2000_ARCHIVES_CACHE_LOCK:
            _DM2000_ARCHIVES_CACHE.clear()
        with _DM2000_BATTERIES_CACHE_LOCK:
            _DM2000_BATTERIES_CACHE.clear()
        with _DM2000_STATS_CACHE_LOCK:
            _DM2000_STATS_CACHE.clear()

        _DM2000_LS_CACHE_READY.set()


def _dmpdata_refresh_cache(force: bool = False) -> None:
    """Copy DMPDATA.mdb to the local cache directory if the source has changed.

    Uses an atomic rename (os.replace) so readers always see a complete file.
    Invalidates the DMPDATA in-memory caches after a successful refresh.
    If the source cannot be read but a previous cache exists, keeps using it.
    """
    global _DMPDATA_CACHE_PATH, _DMPDATA_SOURCE_MTIME  # noqa: PLW0603

    dmpdata_path = Path(get_dmpdata_path())
    if not dmpdata_path.exists():
        if not _DMPDATA_CACHE_READY.is_set():
            _DMPDATA_CACHE_READY.set()
        return

    try:
        current_mtime = dmpdata_path.stat().st_mtime
    except OSError:
        if not _DMPDATA_CACHE_READY.is_set():
            _DMPDATA_CACHE_READY.set()
        return

    with _DMPDATA_CACHE_WRITE_LOCK:
        if not force and current_mtime <= _DMPDATA_SOURCE_MTIME and _DMPDATA_CACHE_READY.is_set():
            return  # source unchanged

        cache_dir = Path(DMPDATA_CACHE_DIR).resolve()
        try:
            cache_dir.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            logger.warning("dmpdata_cache: cannot create cache dir %s: %s", cache_dir, exc)
            if not _DMPDATA_CACHE_READY.is_set():
                _DMPDATA_CACHE_PATH = str(dmpdata_path)
                _DMPDATA_CACHE_READY.set()
            return

        cache_final = str(cache_dir / "DMPDATA.mdb")
        cache_tmp = str(cache_dir / "DMPDATA_new.mdb")
        try:
            shutil.copy2(str(dmpdata_path), cache_tmp)
        except (OSError, PermissionError) as exc:
            logger.warning("dmpdata_cache: copy failed (%s), continuing with existing cache: %s", dmpdata_path, exc)
            try:
                os.unlink(cache_tmp)
            except OSError:
                pass
            source_str = str(dmpdata_path)
            if not _DMPDATA_CACHE_READY.is_set():
                _DMPDATA_CACHE_PATH = source_str
                _DMPDATA_CACHE_READY.set()
            elif _DMPDATA_CACHE_PATH != source_str:
                _DMPDATA_CACHE_PATH = source_str
            return

        # Validate magic bytes before replacing.
        try:
            size = os.path.getsize(cache_tmp)
            if size >= 32 * 1024:
                with open(cache_tmp, "rb") as fh:
                    magic = fh.read(4)
                if magic != b"\x00\x01\x00\x00":
                    raise ValueError(f"Not a valid Access DB (magic={magic!r})")
        except (OSError, ValueError) as exc:
            logger.warning("dmpdata_cache: validation failed for %s: %s", cache_tmp, exc)
            try:
                os.unlink(cache_tmp)
            except OSError:
                pass
            source_str = str(dmpdata_path)
            if not _DMPDATA_CACHE_READY.is_set():
                _DMPDATA_CACHE_PATH = source_str
                _DMPDATA_CACHE_READY.set()
            elif _DMPDATA_CACHE_PATH != source_str:
                _DMPDATA_CACHE_PATH = source_str
            return

        # Atomically replace the cached copy.  Retry on transient ODBC locks.
        replace_exc: "OSError | None" = None
        for attempt in range(6):
            try:
                os.replace(cache_tmp, cache_final)
                replace_exc = None
                break
            except OSError as exc:
                replace_exc = exc
                if attempt < 5:
                    time.sleep(0.4)
        if replace_exc is not None:
            # All retries failed — fall back to reading directly from source.
            logger.warning(
                "dmpdata_cache: rename failed after retries (file in use?), falling back to source: %s",
                replace_exc,
            )
            try:
                os.unlink(cache_tmp)
            except OSError:
                pass
            source_str = str(dmpdata_path)
            if not _DMPDATA_CACHE_READY.is_set():
                _DMPDATA_CACHE_PATH = source_str
                _DMPDATA_CACHE_READY.set()
            elif _DMPDATA_CACHE_PATH != source_str:
                _DMPDATA_CACHE_PATH = source_str
            return

        _DMPDATA_CACHE_PATH = cache_final
        _DMPDATA_SOURCE_MTIME = current_mtime
        logger.info("dmpdata_cache: refreshed local copy from %s (mtime=%s)", dmpdata_path, current_mtime)

        _DMPDATA_CACHE_READY.set()


@asynccontextmanager
async def _lifespan(application):
    watcher_thread = threading.Thread(target=_watch_dmp_changes_loop, daemon=True)
    watcher_thread.start()
    # Build the initial local caches in thread executors so the async event loop
    # is not blocked during startup.  Simple endpoints (templates, config) can
    # respond immediately; DB-backed endpoints wait for the *_CACHE_READY events.
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, lambda: _dm2000_refresh_ls_cache(force=True))
    await loop.run_in_executor(None, lambda: _dmpdata_refresh_cache(force=True))
    # Register with the Voniko server only AFTER the local caches are ready so
    # that requests proxied to this station are not met with a premature 503.
    if VONIKO_SERVER_URL and DMP_STATION_NAME:
        t = threading.Thread(target=_registration_loop, daemon=True)
        t.start()
    yield


app = FastAPI(title="DMP Battery Data Bridge", version="1.0.0", lifespan=_lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


@contextmanager
def shadow_copy_any(mdb_path: str, allowed_dir: str):
    """Generic shadow copy with path restriction under allowed_dir."""
    source = Path(mdb_path).resolve()
    base_dir = Path(allowed_dir).resolve()
    if source.suffix.lower() != ".mdb":
        raise HTTPException(status_code=400, detail="Invalid file type")
    if not source.exists() or not source.is_file():
        raise HTTPException(status_code=404, detail="MDB file not found")
    try:
        source.relative_to(base_dir)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid path") from exc

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".mdb")
    os.close(tmp_fd)
    try:
        try:
            shutil.copy2(str(source), tmp_path)
        except PermissionError as exc:
            logger.warning("shadow_copy: PermissionError copying %s — file locked by DMP app: %s", source, exc)
            raise HTTPException(
                status_code=503,
                detail=f"MDB file is locked by DMP application, retry later: {exc}",
            ) from exc
        except OSError as exc:
            logger.warning("shadow_copy: OSError copying %s: %s", source, exc)
            raise HTTPException(
                status_code=503,
                detail=f"Cannot read MDB file: {exc}",
            ) from exc

        # Validate: must be at least 32 KB and start with Jet/ACE magic bytes
        copied_size = os.path.getsize(tmp_path)
        if copied_size < 32 * 1024:
            raise HTTPException(
                status_code=422,
                detail=f"File too small ({copied_size} bytes) to be a valid Access database (possible Windows shortcut)",
            )
        with open(tmp_path, "rb") as fh:
            magic = fh.read(4)
        if magic != b"\x00\x01\x00\x00":
            raise HTTPException(
                status_code=422,
                detail="Not a valid Access database header (possible Windows shortcut)",
            )

        yield tmp_path
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


@contextmanager
def shadow_copy(mdb_path: str):
    """Copy live DMP .mdb to temp location before querying to avoid file-lock errors."""
    with shadow_copy_any(mdb_path, DMP_DATA_DIR) as tmp:
        yield tmp


@contextmanager
def shadow_copy_dm2000(mdb_path: str):
    """Copy DM2000 .mdb to a temp file before querying to avoid lock conflicts."""
    with shadow_copy_any(mdb_path, DM2000_DATA_DIR) as tmp:
        yield tmp


def _inline_params(sql: str, params: tuple) -> str:
    """Inline params into SQL for Access ODBC drivers that don't support SQLBindParameter."""
    parts = sql.split("?")
    if len(parts) != len(params) + 1:
        raise ValueError("Parameter count mismatch")
    result = parts[0]
    for i, param in enumerate(params):
        if param is None:
            result += "NULL"
        elif isinstance(param, (int, float)):
            result += str(param)
        else:
            value = str(param)
            # Permit only a conservative character set used by batch/cdmc identifiers
            # and date-like strings when fallback inlining is required by Access ODBC.
            # Includes () , + / # @ which appear in cdmc values and model names.
            if not re.fullmatch(r"[A-Za-z0-9 _:.(),%+/#@-]+", value):
                raise ValueError("Unsafe string parameter for inline SQL fallback")
            escaped = value.replace("'", "''")
            result += f"'{escaped}'"
        result += parts[i + 1]
    return result


@contextmanager
def _acquire_query_lock():
    """Acquire the Access DB semaphore with a timeout.

    Raises HTTPException(503) instead of blocking indefinitely when the
    database is saturated with concurrent requests.
    """
    acquired = _ACCESS_QUERY_LOCK.acquire(timeout=_ACCESS_QUERY_TIMEOUT)
    if not acquired:
        raise HTTPException(
            status_code=503,
            detail="Database is busy with too many concurrent requests, please retry shortly",
        )
    try:
        yield
    finally:
        _ACCESS_QUERY_LOCK.release()


def query_mdb(mdb_path: str, sql: str, params: tuple = (), retries: int = 2) -> list[dict]:
    """Run an Access query with bounded concurrency and short HY000 retries.

    Retries are only applied to transient HY000 driver errors using a linear
    backoff of 0.3s * attempt and preserve existing HYC00 inline fallback.
    """
    conn_str = (
        r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
        f"DBQ={mdb_path};"
    )
    for attempt_num in range(1, retries + 2):
        try:
            with _acquire_query_lock():
                with pyodbc.connect(conn_str, timeout=10) as conn:
                    cursor = conn.cursor()
                    try:
                        cursor.execute(sql, params) if params else cursor.execute(sql)
                    except pyodbc.Error as exc:
                        err_str = str(exc)
                        if params and ("HYC00" in err_str or "SQLBindParameter" in err_str or "07002" in err_str):
                            try:
                                cursor = conn.cursor()
                                cursor.execute(_inline_params(sql, params))
                            except (pyodbc.Error, ValueError) as inline_exc:
                                logger.warning(
                                    "Inline parameter fallback failed (%s); re-raising original execute error (%s)",
                                    inline_exc,
                                    exc,
                                )
                                raise exc from inline_exc
                        else:
                            raise
                    columns = [col[0] for col in cursor.description] if cursor.description else []
                    return [dict(zip(columns, row)) for row in cursor.fetchall()]
        except pyodbc.Error as exc:
            err_str = str(exc)
            if "HY000" in err_str and attempt_num <= retries:
                wait = 0.3 * attempt_num
                logger.debug("HY000 on attempt %d, retrying in %.1fs: %s", attempt_num, wait, exc)
                time.sleep(wait)
                continue
            raise


def resolve_data_file(filename: str) -> str:
    """Resolve a filename to absolute path within DMP_DATA_DIR. Raises HTTPException on traversal."""
    parsed = Path(filename)
    if parsed.parent != Path("."):
        raise HTTPException(status_code=400, detail="Invalid filename")
    stem = parsed.stem
    if not re.match(r'^[A-Za-z0-9_-]+$', stem):
        raise HTTPException(status_code=400, detail="Invalid filename")
    result = Path(DMP_DATA_DIR).resolve() / (stem + ".mdb")
    if not str(result).startswith(str(Path(DMP_DATA_DIR).resolve())):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not result.exists():
        raise HTTPException(status_code=404, detail="Data file not found")
    return str(result)


def get_dmpdata_path() -> str:
    return str(Path(DMP_DATA_DIR).resolve() / "DMPDATA.mdb")


def get_dm2000_ls_path() -> str:
    return str(Path(DM2000_DATA_DIR).resolve() / "dmdata_ls.mdb")


def get_dm2000_main_path() -> str:
    return str(Path(DM2000_DATA_DIR).resolve() / "DM2000.mdb")


def compute_stats(rows: list[dict]) -> dict:
    def safe_float(value):
        if value == "--" or value == "" or value is None:
            return None
        try:
            f = float(value)
            return None if math.isnan(f) else f
        except (TypeError, ValueError):
            return None

    def get_value(row: dict, *keys):
        for key in keys:
            if key in row and row.get(key) is not None:
                v = row.get(key)
                if v == "--" or v == "":
                    continue
                return v
        return None

    volt_vals = [safe_float(get_value(r, "VOLT", "volt", "Volt")) for r in rows]
    volt_vals = [v for v in volt_vals if v is not None]
    im_vals_all = [safe_float(get_value(r, "Im", "IM", "im")) for r in rows]
    im_vals_all = [i for i in im_vals_all if i is not None]
    im_vals_active = [i for i in im_vals_all if i > 0]

    def agg(vals):
        if not vals:
            return {"max": None, "min": None, "avg": None}
        return {
            "max": round(max(vals), 4),
            "min": round(min(vals), 4),
            "avg": round(sum(vals) / len(vals), 4),
        }

    v = agg(volt_vals)
    i_all = agg(im_vals_all)
    i_active = agg(im_vals_active)
    return {
        "VOLT_MAX": v["max"],
        "VOLT_MIN": v["min"],
        "VOLT_AVG": v["avg"],
        "IM_MAX": i_all["max"],
        "IM_MIN": i_all["min"],
        "IM_AVG": i_active["avg"],
    }


class ReportRequest(BaseModel):
    batch_id: str
    cdmc: str
    channel: int
    template_name: str


class DMPSimpleReportRequest(BaseModel):
    batch_id: str
    # Legacy single-channel fields (used when ``batys`` is not provided so older
    # callers keep working).  When ``batys`` is non-empty the new DM2000-style
    # multi-battery preview format is generated and these fields are ignored.
    cdmc: Optional[str] = None
    channel: Optional[int] = None
    # New multi-battery fields mirroring DM2000SimpleReportRequest
    batys: list[int] = Field(default=[])
    override_battery_type: Optional[str] = None
    override_manufacturer: Optional[str] = None
    endpoint_cutoff: Optional[float] = None


class DM2000ReportRequest(BaseModel):
    archname: str
    baty: int = Field(ge=0)
    template_name: str = Field(min_length=6)
    override_archname: Optional[str] = None
    override_start_date: Optional[str] = None
    override_battery_type: Optional[str] = None
    override_batch_name: Optional[str] = None
    override_discharge_condition: Optional[str] = None
    override_manufacturer: Optional[str] = None
    override_made_date: Optional[str] = None
    override_serial_no: Optional[str] = None
    override_remarks: Optional[str] = None


class DM2000SimpleReportRequest(BaseModel):
    archname: str
    batys: list[int] = Field(default=[])
    override_battery_type: Optional[str] = None
    override_manufacturer: Optional[str] = None
    endpoint_cutoff: Optional[float] = None


class PerfReportEntry(BaseModel):
    archname: str
    battery_type: str  # e.g. "HP", "UD", "UD+"
    batys: list[int] = Field(default=[])  # empty = use all detected batteries
    sheet_name: str = ""  # auto-derived from dcxh+serialno when empty


class PerfReportRequest(BaseModel):
    entries: list[PerfReportEntry]
    template_name: Optional[str] = None


class DmpPerfGroup(BaseModel):
    loai: str  # e.g. "UD", "HP", "UD+"
    chuyen: str  # production-line number, e.g. "501"
    trays: list[int] = Field(default=[])  # channel numbers (1-9); empty = auto-assigned


class DmpPerfEntry(BaseModel):
    batch_id: str  # para_pub.id
    model: str  # e.g. "LR6", "LR03", "LR61", "9V"
    groups: list[DmpPerfGroup]
    special_type: str = "normal"  # "normal" | "6020" | "3thang" | "6thang"
    report_date: Optional[str] = None  # YYYY-MM-DD from SQLite; used as row-label fallback
    raw_remark: Optional[str] = None  # free-text remark; used as fallback to search para_pub.bz
    dm2000_archname: Optional[str] = None  # DM2000 archive name; when set, data is read from DM2000 instead of DMP


class DmpPerfReportRequest(BaseModel):
    entries: list[DmpPerfEntry]
    template_name: Optional[str] = None


def render_excel_template(template_path: str, context: dict) -> bytes:
    wb = load_workbook(template_path)
    for ws in wb.worksheets:
        _process_worksheet(ws, context)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _interpolate_cell(value, ctx: dict):
    if not isinstance(value, str):
        return value
    m = re.fullmatch(r'\{\{(\w+)\}\}', value.strip())
    if m:
        key = m.group(1)
        return ctx.get(key, value)

    def replacer(match):
        key = match.group(1)
        v = ctx.get(key, match.group(0))
        return str(v) if v is not None else ''

    return re.sub(r'\{\{(\w+)\}\}', replacer, value)


def _process_worksheet(ws, ctx: dict):
    history_data = ctx.get("HISTORY_DATA", [])

    rows_to_expand = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '{{#HISTORY_DATA}}' in cell.value:
                rows_to_expand.append(cell.row)
                break

    for template_row_idx in sorted(rows_to_expand, reverse=True):
        template_row = list(ws.iter_rows(min_row=template_row_idx, max_row=template_row_idx))[0]

        if history_data:
            if len(history_data) > 1:
                ws.insert_rows(template_row_idx + 1, len(history_data) - 1)

            for i, item in enumerate(history_data):
                target_row = template_row_idx + i
                for j, tmpl_cell in enumerate(template_row):
                    target_cell = ws.cell(row=target_row, column=tmpl_cell.column)
                    if i > 0:
                        target_cell._style = tmpl_cell._style
                    raw = tmpl_cell.value
                    if isinstance(raw, str):
                        raw = raw.replace('{{#HISTORY_DATA}}', '').strip()
                    target_cell.value = _interpolate_cell(raw, item) if raw else raw
        else:
            for cell in template_row:
                cell.value = None

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '{{' in cell.value and '{{#HISTORY_DATA}}' not in cell.value:
                cell.value = _interpolate_cell(cell.value, ctx)


def _resolve_template_path(template_name: str) -> str:
    if not _is_valid_template_name(template_name):
        raise HTTPException(status_code=400, detail="Invalid template")

    base = Path(DMP_TEMPLATES_DIR).resolve()
    allowed = {
        f.name for f in base.iterdir()
        if f.is_file() and _is_valid_template_name(f.name)
    } if base.exists() else set()
    if template_name not in allowed:
        raise HTTPException(status_code=404, detail="Template not found")
    result = (base / template_name).resolve()
    try:
        result.relative_to(base)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Template path traversal detected") from exc
    if not result.exists():
        raise HTTPException(status_code=404, detail="Template not found")
    return str(result)


def _read_dmpdata(sql: str, params: tuple = ()) -> list[dict]:
    """Query the persistent local cache of DMPDATA.mdb.

    Waits up to _DMPDATA_CACHE_READY_TIMEOUT seconds for the initial cache to be
    built on startup, then reads directly from the cached copy without creating a
    new shadow copy per request.  Falls back to a fresh shadow copy of the source
    file when the cache has not been built (e.g. DMPDATA_CACHE_DIR unavailable).
    """
    if _DMPDATA_CACHE_READY.wait(timeout=_DMPDATA_CACHE_READY_TIMEOUT):
        if _DMPDATA_CACHE_PATH:
            return query_mdb(_DMPDATA_CACHE_PATH, sql, params)
    # Fallback: original shadow-copy path (cache not available)
    dmpdata = Path(get_dmpdata_path())
    if not dmpdata.exists():
        raise HTTPException(status_code=404, detail="DMPDATA.mdb not found")
    with shadow_copy(str(dmpdata)) as copied:
        return query_mdb(copied, sql, params)


def _read_dm2000_ls(sql: str, params: tuple = ()) -> list[dict]:
    """Query the persistent local cache of dmdata_ls.mdb.

    Waits up to _DM2000_LS_CACHE_READY_TIMEOUT seconds for the initial cache to be built (startup), then reads
    directly from the cached copy without creating a new shadow copy per request.
    The global ACCESS_QUERY_LOCK inside query_mdb still caps concurrent ODBC
    connections, but no per-request file duplication overhead exists.
    """
    if not _DM2000_LS_CACHE_READY.wait(timeout=_DM2000_LS_CACHE_READY_TIMEOUT):
        raise HTTPException(
            status_code=503,
            detail="DM2000 database not ready, please retry shortly",
        )
    if not _DM2000_LS_CACHE_PATH:
        raise HTTPException(status_code=404, detail="dmdata_ls.mdb not found")
    return query_mdb(_DM2000_LS_CACHE_PATH, sql, params)


def _read_dm2000_ls_multi(queries: list[tuple[str, tuple]]) -> list[dict]:
    """Execute a list of (sql, params) queries against the persistent local cache
    of dmdata_ls.mdb, returning the first successful result.

    Raises HTTPException(503) when the cache is not yet ready.
    Re-raises the last :class:`pyodbc.Error` when every query in *queries* fails.
    """
    if not _DM2000_LS_CACHE_READY.wait(timeout=_DM2000_LS_CACHE_READY_TIMEOUT):
        raise HTTPException(
            status_code=503,
            detail="DM2000 database not ready, please retry shortly",
        )
    if not _DM2000_LS_CACHE_PATH:
        raise HTTPException(status_code=404, detail="dmdata_ls.mdb not found")
    last_exc: "pyodbc.Error | None" = None
    for sql, params in queries:
        try:
            return query_mdb(_DM2000_LS_CACHE_PATH, sql, params)
        except pyodbc.Error as exc:
            last_exc = exc
    if last_exc is not None:
        raise last_exc
    return []  # only reached when queries is empty


def _resolve_dm2000_template_path(template_name: str) -> str:
    if not _is_valid_template_name(template_name):
        raise HTTPException(status_code=400, detail="Invalid template")

    base = Path(DM2000_TEMPLATES_DIR).resolve()
    allowed = {
        f.name for f in base.iterdir()
        if f.is_file() and _is_valid_template_name(f.name)
    } if base.exists() else set()
    if template_name not in allowed:
        raise HTTPException(status_code=404, detail="Template not found")
    result = (base / template_name).resolve()
    try:
        result.relative_to(base)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Template path traversal detected") from exc
    if not result.exists():
        raise HTTPException(status_code=404, detail="Template not found")
    return str(result)


def _resolve_perf_template_path(template_name: str) -> str:
    if not _is_valid_template_name(template_name):
        raise HTTPException(status_code=400, detail="Invalid template")

    base = Path(DM2000_PERF_TEMPLATES_DIR).resolve()
    allowed = {
        f.name: f for f in base.iterdir()
        if f.is_file() and _is_valid_template_name(f.name)
    } if base.exists() else {}
    if template_name not in allowed:
        raise HTTPException(status_code=404, detail="Perf template not found")
    # Use the filesystem-derived Path object (not raw user input) to avoid taint
    result = allowed[template_name].resolve()
    try:
        result.relative_to(base)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Template path traversal detected") from exc
    if not result.is_file():
        raise HTTPException(status_code=404, detail="Perf template not found")
    return str(result)


def _perf_cell_value_with_merge(ws, row: int, col: int):
    """Return the effective display value of a cell, accounting for merged ranges."""
    for merge_range in ws.merged_cells.ranges:
        if (
            merge_range.min_row <= row <= merge_range.max_row
            and merge_range.min_col <= col <= merge_range.max_col
        ):
            return ws.cell(row=merge_range.min_row, column=merge_range.min_col).value
    return ws.cell(row=row, column=col).value


def _perf_col_header(ws, col: int, below_row: int) -> str:
    """Scan upward from *below_row* and return the first non-empty cell value
    in *col*, respecting merged cells.  Returns an empty string if not found."""
    for row_idx in range(below_row - 1, 0, -1):
        val = _perf_cell_value_with_merge(ws, row_idx, col)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def _perf_find_fdfs_for_col(
    ws, col: int, below_row: int, all_fdfs: list[str]
) -> tuple[str, str]:
    """Scan *all* rows above *below_row* in *col* and return the first
    ``(fdfs_label, header_text)`` pair where the header matches a known fdfs.
    Returns ``("", "")`` when no match is found.

    This correctly handles templates where multiple header rows exist above the
    data-tag row (e.g. a merged fdfs label in row 2 and a sub-header in row 3).
    """
    for row_idx in range(below_row - 1, 0, -1):
        val = _perf_cell_value_with_merge(ws, row_idx, col)
        if val is None:
            continue
        s = str(val).strip()
        if not s:
            continue
        for fdfs_lbl in all_fdfs:
            if _perf_fdfs_matches_header(fdfs_lbl, s):
                return (fdfs_lbl, s)
    return ("", "")


def _perf_normalize_date(v) -> str:
    """Normalise a cell value to a ``YYYY-MM-DD`` string, or return ``""``."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Accept YYYY/MM/DD or YYYY-MM-DD (first 10 chars)
    s = s[:10].replace("/", "-")
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return s
    return ""


# Standard IEC 60086-2 discharge-test-condition template order per battery family.
# Each entry is a condition label as it typically appears in the Excel report template.
# Matching against actual DB labels uses _perf_fdfs_matches_template (fuzzy + bracket-norm).
_TEMPLATE_CONDITION_ORDER: dict[str, list[str]] = {
    "LR6": [
        "10ohm 24h/d-0.9V",
        "1000mA 24h/d-0.9V",
        "(1500mW2s,650mW28s)10T/h,24h/d-1.05V",
        "(1500mW2s,650mW28s)10T/h,24h/d-1.0V",
        "3.9ohm 1h/d-0.8V",
        "3.9ohm 4m/h 8h/d-0.9V",
        "250mA 1h/d-0.9V",
        "3.9ohm 24h/d-0.8V",
        "1000mA 10s/m 1h/d-0.9V",
        "100mA 1h/d-0.9V",
        "50mA 1h/8h 24h/d-1V",
        "750mA 2m/h 8h/d-1.1V",
        "(450mW5s,45mW175s) 3h/124h-1.1V",
        "(1ohm,0.25s.3.0ohm,19.75s),10m/h,1h/12h-1.0V",
    ],
    "LR03": [
        "20ohm 24h/d-0.9V",
        "600mA 24h/d-0.9V",
        "5.1ohm 1h/d-0.8V",
        "5.1ohm 4m/h 8h/d-0.9V",
        "600mA 10s/m 1h/d-0.9V",
        "50mA 1h/12h-0.9V",
        "250mA 5m/h 12h/d-1.1V",
        "100mA 1h/d-0.9V",
        "24ohm 15s/m 8h/d-1V",
        "3.9ohm 24h/d-0.8V",
        "75mA 1h/12h 24h/d-0.9V",
    ],
    "LR61": [
        "35mA 24h/d-0.9V",
        "5.1ohm 5m/d-0.9V",
        "75ohm 1h/d-0.9V",
        "75ohm 1h/d-1.1V",
    ],
    "9V": [
        "35mA 24h/d-5.4V",
        "180ohm 4h/d-6.8V",
        "270ohm 1h/d-5.4V",
        "620ohm 2h/d-5.4V",
        "620ohm+10Kohm 1s/60m.24h/d-7.5V",
    ],
}

# Normalise bracket style so that DMP-software labels using {} are treated the
# same as template labels that use () when doing template-order lookups.
_BRACKET_NORM_TABLE = str.maketrans("{}", "()")


def _perf_fdfs_matches_template(cond: str, tmpl: str) -> bool:
    """Return True if *cond* (a DB condition label) matches *tmpl* (a template entry).

    Stricter than ``_perf_fdfs_matches_header``: the leading-token fallback is
    intentionally omitted so that ``1000mA 10s/m 1h/d-0.9V`` does not
    incorrectly match ``1000mA 24h/d-0.9V`` via the shared ``1000mA`` prefix.
    Also normalises ``{}`` → ``()`` on both sides before comparing.
    """
    c_norm = cond.translate(_BRACKET_NORM_TABLE)
    t_norm = tmpl.translate(_BRACKET_NORM_TABLE)

    # Strip trailing unit annotation like "(h)", "(m)", "(t)"
    _unit_re = re.compile(r"\s*\([hHmMtT]\)\s*$")
    f = re.sub(_unit_re, "", c_norm).strip().lower()
    h = re.sub(_unit_re, "", t_norm).strip().lower()
    if not f or not h:
        return False

    # Exact match after normalisation
    if f == h:
        return True

    # Whitespace-normalised match (spacing differences)
    f_no_ws = re.sub(r'\s+', '', f)
    h_no_ws = re.sub(r'\s+', '', h)
    if f_no_ws and h_no_ws and f_no_ws == h_no_ws:
        return True

    # Strip trailing endpoint-voltage suffix from one or both sides.
    # Handles both "-0.9V" (template / DMP jstj format) and " to 0.900V"
    # (older DM2000 label format) suffixes so that "10ohm 24h/d-0.900V"
    # and "10ohm 24h/d-0.9V" are considered equivalent.
    _vsuf = re.compile(r'(\s*-\s*|\s+to\s+)\d+\.?\d*\s*[vV]\s*$', re.IGNORECASE)
    f_no_v = re.sub(r'\s+', '', _vsuf.sub('', f))
    h_no_v = re.sub(r'\s+', '', _vsuf.sub('', h))
    if f_no_v and h_no_ws and f_no_v == h_no_ws:
        return True
    if f_no_ws and h_no_v and f_no_ws == h_no_v:
        return True
    # Both sides voltage-stripped (handles differing decimal precision, e.g.
    # "10ohm 24h/d-0.900V" vs "10ohm 24h/d-0.9V")
    if f_no_v and h_no_v and f_no_v == h_no_v:
        return True

    return False


def _template_condition_sort_key(cond: str, battery_type: str) -> tuple:
    """Return a sort key ``(template_pos, cond)`` for a condition label.

    *template_pos* is the 0-based position in the IEC template for *battery_type*,
    or ``len(template)`` when the condition is not found (unrecognised conditions
    are appended at the end, preserving their relative insertion order via *cond*).
    """
    battery_type_upper = battery_type.strip().upper()
    template = _TEMPLATE_CONDITION_ORDER.get(battery_type_upper, [])
    for i, tmpl_entry in enumerate(template):
        if _perf_fdfs_matches_template(cond, tmpl_entry):
            return (i, cond)
    return (len(template), cond)


def _perf_fdfs_matches_header(fdfs: str, header: str) -> bool:
    """Return True if *fdfs* label is a reasonable match for *header* text.

    The header may contain a unit suffix such as ``(h)`` / ``(m)`` / ``(t)``
    which is ignored for matching purposes.

    Uses whole-word boundary matching to avoid false positives such as
    "10ohm" incorrectly matching a "100ohm" column header.
    """
    if not fdfs or not header:
        return False
    # Strip trailing unit annotation like "(h)", "(m)", "(t)"
    clean_header = re.sub(r"\s*\([hHmMtT]\)\s*$", "", header).strip()
    f = fdfs.lower().strip()
    h = clean_header.lower().strip()
    if not f or not h:
        return False
    # Exact match after normalisation
    if f == h:
        return True
    # Whitespace-normalized match: ignore internal spacing differences.
    # e.g. "(1500mW2s,650mW28s)10T/h,24h/d-1.05V" (DB) vs
    #      "(1500mW2s,650mW28s) 10T/h,24h/d-1.05V" (template header with space)
    f_no_ws = re.sub(r'\s+', '', f)
    h_no_ws = re.sub(r'\s+', '', h)
    if f_no_ws and h_no_ws and f_no_ws == h_no_ws:
        return True
    # Match when one side carries a trailing endpoint-voltage suffix ("-X.XXV")
    # and the other does not.  The most common case is the fdfs label coming from
    # para_pub.jstj which embeds the voltage (e.g. "(1500mW2s,650mW28s)10T/h,24h/d-1.05V")
    # while the Excel template header omits it (e.g. "(1500mW2s,650mW28s) 10T/h,24h/d").
    # Only one side is stripped at a time so two labels with *different* voltages
    # (e.g. "-1.05V" vs "-0.9V") are never incorrectly considered equal.
    _voltage_suffix_pattern = re.compile(r'\s*-\s*\d+\.?\d*\s*[vV]\s*$')
    f_no_v = re.sub(r'\s+', '', _voltage_suffix_pattern.sub('', f))
    h_no_v = re.sub(r'\s+', '', _voltage_suffix_pattern.sub('', h))
    if f_no_v and h_no_ws and f_no_v == h_no_ws:
        return True
    if f_no_ws and h_no_v and f_no_ws == h_no_v:
        return True
    # Whole-word boundary check — prevents "10ohm" from matching "100ohm".
    # A word boundary here means the match is not immediately preceded or
    # followed by an alphanumeric character or a forward-slash.
    _wb = r'(?<![0-9A-Za-z/])'
    _we = r'(?![0-9A-Za-z/])'

    def _whole_word(needle: str, haystack: str) -> bool:
        return bool(re.search(_wb + re.escape(needle) + _we, haystack))

    if _whole_word(f, h) or _whole_word(h, f):
        return True
    # Match on leading token (e.g. "10ohm" vs "10ohm 24h/d-0.9V")
    f_tok = f.split()[0] if f.split() else ""
    h_tok = h.split()[0] if h.split() else ""
    if f_tok and h_tok and f_tok == h_tok:
        return True
    return False


def _render_perf_template(template_path: str, groups: dict) -> bytes:
    """Fill a user-uploaded perf template Excel with performance data.

    **Template tag reference** — place these tags in the worksheet cells:

    * ``{{PERF_SHEET_NAME}}`` — replaced with the battery-series sheet name
      (e.g. "LR6 501").  Can be placed in any cell (title row, etc.).

    **Mode A — row-expansion (``{{#PERF_ROWS}}``):**
      Any cell in a row that contains ``{{#PERF_ROWS}}`` marks that row as the
      template for data rows.  That row is expanded (one copy per data point).
      Tags available inside the row:

      * ``{{DATE}}``       — test date (``YYYY-MM-DD``)
      * ``{{TYPE}}``       — battery grade (``HP`` / ``UD`` / ``UD+``)
      * ``{{RESULT_0}}``   — avg discharge result for the 1st fdfs condition
      * ``{{RATE_0}}``     — uniform rate (%) for the 1st fdfs condition
      * ``{{RESULT_1}}``, ``{{RATE_1}}`` — 2nd fdfs condition, … and so on.

    **Mode B — date-based cell lookup (no ``{{#PERF_ROWS}}``):**
      When no ``{{#PERF_ROWS}}`` marker is found the engine falls back to
      date-based row matching.  The template must have dates pre-filled in
      column A (format ``YYYY/MM/DD`` or ``YYYY-MM-DD``).  Place the following
      tags in any row that contains a matching date in column A:

      * ``{{TYPE}}``         — battery grade
      * ``{{RESULT_0}}`` … — result for the 1st (alphabetically) fdfs condition.
        The engine also inspects column headers *above* each ``{{RESULT_N}}``
        tag and tries to match them against the archive's discharge-pattern
        (fdfs) label, allowing the correct column to be filled regardless of
        the number of conditions exported in one call.
      * ``{{RATE_0}}`` …    — uniform rate for the matching fdfs condition.

      The unit written for ``{{RESULT_N}}`` is chosen from the column header
      suffix.  **Convention by data source:**

      * **DM2000** reports time-based metrics:

        * header ending with ``(h)``  → average discharge time in **hours**
        * header ending with ``(m)``  → average discharge time in **minutes**
        * header ending with ``(t)``  → also minutes (legacy alias for ``(m)``)

      * **DMP** reports cycle-count metrics for ``(t)`` columns:

        * header ending with ``(h)``  → average discharge time in **hours**
        * header ending with ``(t)``  → average **number of discharge cycles**
          ("số lần phóng điện") — an integer count, *not* a time value
        * header ending with ``(m)`` is **not used** for DMP and will fall back
          to hours; if a DMP column is meant to track cycles, mark it with
          ``(t)`` in the header.

      **Rate cells:** Any cell whose ``{{RATE_N}}`` tag (or destination cell
      in date-lookup mode) is formatted as a percentage (``0.00%``) receives
      the uniform rate as a 0-1 fraction; otherwise it receives the value in
      0-100 form.  The percentage format is detected from the destination
      cell, so the data row that actually holds the value must be formatted
      as ``0.00%``.

      **Column-header naming guide** (so values land in the correct column):

      * The text **before** the ``(h)``/``(m)``/``(t)`` suffix must contain
        the discharge-condition (fdfs) label exactly as stored in the source
        DB (``para_pub.fdfs`` for DMP, ``ls_jb_cs.fdfs`` for DM2000), or
        contain the leading token (e.g. ``10ohm``, ``1000mA``,
        ``(1500mW2s,650mW28s)10T/h,24h/d``).  Matching is case-insensitive and
        uses whole-word boundaries to avoid false hits like ``10ohm`` matching
        ``100ohm``.
      * If a header text cannot be matched to any condition exported in the
        request, **no value is written** to that column (the cell is left as
        defined in the template).  This prevents wrong-column writes when the
        source DB has empty ``fdfs``/``jstj`` fields.

    Sheet matching: the key in *groups* may contain ``|``-separated candidate
    names (e.g. ``"LR6 Voniko|LR6 501|LR6"``).  The engine tries each candidate
    in order and uses the first one that exists in the workbook.  If none match,
    the entry is skipped and the template sheet is left unchanged.
    """
    wb = load_workbook(template_path)

    for sheet_name_key, date_type_map in groups.items():
        # Resolve the first candidate that exists in the workbook.
        candidates = [s for c in sheet_name_key.split("|") if (s := c.strip())]
        sheet_name = next((c for c in candidates if c in wb.sheetnames), None)
        if sheet_name is None:
            continue
        ws = wb[sheet_name]

        # Collect sorted fdfs labels for this sheet
        all_fdfs: list[str] = []
        seen_fdfs: set[str] = set()
        for row_data in date_type_map.values():
            for lbl in row_data:
                if lbl not in seen_fdfs:
                    seen_fdfs.add(lbl)
                    all_fdfs.append(lbl)
        all_fdfs.sort()

        # Sort data rows by (date, battery_type)
        sorted_keys = sorted(date_type_map.keys(), key=lambda k: (k[0], k[1]))

        # ── Mode A: find the PERF_ROWS template row ───────────────────────────
        perf_row_idx: int | None = None
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{#PERF_ROWS}}" in cell.value:
                    perf_row_idx = cell.row
                    break
            if perf_row_idx is not None:
                break

        if perf_row_idx is not None and sorted_keys:
            template_row = list(ws.iter_rows(min_row=perf_row_idx, max_row=perf_row_idx))[0]

            # Pre-scan template row to detect which RATE_j columns are percentage-formatted.
            # A percentage-formatted cell (number_format containing "%") expects a 0-1
            # decimal fraction; the uniform_rate values are in the 0-100 range, so they
            # must be divided by 100 before being placed in such cells.
            #
            # Also detect the unit suffix of each RESULT_j column header so DMP
            # cycle-count columns ("(t)") are populated with ``avg_count`` rather
            # than ``avg_hours``.
            rate_is_pct: dict[int, bool] = {}  # j → True if RATE_j cell is pct-formatted
            result_unit: dict[int, str] = {}   # j → "h" | "m" | "t" | "" (lowercased suffix; "" = no explicit unit)
            for rate_tmpl_cell in template_row:
                if isinstance(rate_tmpl_cell.value, str):
                    rate_match = re.fullmatch(r"\{\{RATE_(\d+)\}\}", rate_tmpl_cell.value.strip())
                    if rate_match:
                        rate_fmt = rate_tmpl_cell.number_format or ""
                        rate_is_pct[int(rate_match.group(1))] = "%" in rate_fmt
                    result_match = re.fullmatch(r"\{\{RESULT_(\d+)\}\}", rate_tmpl_cell.value.strip())
                    if result_match:
                        # Inspect column header above this RESULT_j cell to pick the unit
                        hdr = _perf_col_header(ws, rate_tmpl_cell.column, perf_row_idx)
                        h_lower = (hdr or "").lower()
                        if h_lower.endswith("(t)"):
                            result_unit[int(result_match.group(1))] = "t"
                        elif h_lower.endswith("(m)"):
                            result_unit[int(result_match.group(1))] = "m"
                        elif h_lower.endswith("(h)"):
                            result_unit[int(result_match.group(1))] = "h"
                        else:
                            # No explicit unit suffix — resolved per data source at write time
                            result_unit[int(result_match.group(1))] = ""

            # Insert extra rows so there is one row per data point
            if len(sorted_keys) > 1:
                ws.insert_rows(perf_row_idx + 1, len(sorted_keys) - 1)

            for i, row_key in enumerate(sorted_keys):
                date_str, btype = row_key
                row_data = date_type_map[row_key]
                target_row_idx = perf_row_idx + i

                ctx: dict = {"DATE": date_str, "TYPE": btype}
                for j, fdfs_lbl in enumerate(all_fdfs):
                    entry = row_data.get(fdfs_lbl, {})
                    avg_h = entry.get("avg_hours")
                    avg_m = entry.get("avg_minutes")
                    avg_c = entry.get("avg_count")
                    ur = entry.get("uniform_rate")
                    is_dmp = bool(entry.get("is_dmp"))

                    unit = result_unit.get(j, "")
                    if unit == "t":
                        # DMP "(t)" → cycle count; DM2000 "(t)" → minutes (legacy)
                        val = avg_c if is_dmp else avg_m
                    elif unit == "m":
                        val = avg_m
                    elif unit == "h":
                        val = avg_h
                    else:
                        # No explicit unit suffix: DMP → cycle count, DM2000 → hours
                        val = avg_c if is_dmp else avg_h
                    ctx[f"RESULT_{j}"] = val if val is not None else ""

                    ur_val = ur if ur is not None else ""
                    if isinstance(ur_val, (int, float)) and rate_is_pct.get(j, False):
                        ur_val = ur_val / 100.0
                    ctx[f"RATE_{j}"] = ur_val

                for tmpl_cell in template_row:
                    target_cell = ws.cell(row=target_row_idx, column=tmpl_cell.column)
                    if i > 0:
                        target_cell._style = tmpl_cell._style  # type: ignore[attr-defined]
                    raw = tmpl_cell.value
                    if isinstance(raw, str):
                        raw = raw.replace("{{#PERF_ROWS}}", "").strip()
                    target_cell.value = _interpolate_cell(raw, ctx) if raw else raw

        else:
            # ── Mode B: date-based cell lookup ────────────────────────────────
            # Step 1: locate the template tag row — the first row that contains
            #   {{TYPE}}, {{RESULT_N}}, or {{RATE_N}} in any cell.
            type_col: int | None = None
            result_cols: dict[int, int] = {}   # N → column index
            rate_cols: dict[int, int] = {}     # N → column index
            tag_row_idx: int | None = None

            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    v = cell.value.strip()
                    if v == "{{TYPE}}":
                        type_col = cell.column
                        tag_row_idx = cell.row
                    m = re.fullmatch(r"\{\{RESULT_(\d+)\}\}", v)
                    if m:
                        result_cols[int(m.group(1))] = cell.column
                        tag_row_idx = cell.row
                    m = re.fullmatch(r"\{\{RATE_(\d+)\}\}", v)
                    if m:
                        rate_cols[int(m.group(1))] = cell.column
                        tag_row_idx = cell.row
                if tag_row_idx is not None:
                    break  # only scan up to the first tagged row

            if tag_row_idx is not None and (type_col is not None or result_cols):
                # Step 2: build fdfs → (result_col, rate_col, header_text) map.
                # Scan all rows above the tag row for each RESULT column and
                # try to match the cell text against a known fdfs label.  This
                # handles templates where a merged fdfs label sits in one row
                # (e.g. row 2) while a sub-header ("Kết quả") occupies row 3.
                fdfs_col_map: dict[str, tuple[int | None, int | None, str]] = {}
                for n, r_col in result_cols.items():
                    ur_col = rate_cols.get(n)
                    matched_fdfs, header = _perf_find_fdfs_for_col(ws, r_col, tag_row_idx, all_fdfs)
                    if matched_fdfs and matched_fdfs not in fdfs_col_map:
                        fdfs_col_map[matched_fdfs] = (r_col, ur_col, header)
                # Position-based fallback for unmatched fdfs labels.
                # DMP batches commonly leave para_pub.fdfs and para_pub.jstj
                # empty, so the fdfs_label falls back to grp.loai (e.g. "UD").
                # That value never matches a column header, so without this
                # fallback the entire data row would be skipped and only the
                # {{TYPE}} cell would be written.  Header-matched columns keep
                # priority; position-based entries only fill the remaining gaps.
                _unmatched = [
                    lbl for lbl in all_fdfs if lbl not in fdfs_col_map
                ]
                if _unmatched:
                    claimed_cols = {v[0] for v in fdfs_col_map.values()}
                    free_result_cols = sorted(
                        [
                            (n, r_col, rate_cols.get(n))
                            for n, r_col in result_cols.items()
                            if r_col not in claimed_cols
                        ],
                        key=lambda t: t[1],
                    )
                    for lbl, (n, r_col, ur_col) in zip(sorted(_unmatched), free_result_cols):
                        hdr = _perf_col_header(ws, r_col, tag_row_idx)
                        fdfs_col_map[lbl] = (r_col, ur_col, hdr or "")
                        logger.info(
                            "_render_perf_template[%s]: positional fallback assigned"
                            " fdfs label %r → col %d (header: %r)",
                            sheet_name, lbl, r_col, hdr,
                        )

                # Step 3: build date/label → row index map from column A.
                # Also captures special row labels like "6020", "3 THÁNG", "6 THÁNG".
                _SPECIAL_LABELS = {"6020", "3 THÁNG", "6 THÁNG", "3 THANG", "6 THANG"}
                date_row_map: dict[str, int] = {}
                for drow in ws.iter_rows(min_col=1, max_col=1):
                    for dcell in drow:
                        ds = _perf_normalize_date(dcell.value)
                        if ds:
                            date_row_map[ds] = dcell.row
                        elif dcell.value is not None:
                            # Capture special row labels (case-insensitive strip)
                            raw_label = str(dcell.value).strip()
                            up_label = raw_label.upper()
                            if up_label in {s.upper() for s in _SPECIAL_LABELS}:
                                date_row_map[up_label] = dcell.row
                            elif raw_label:
                                # Store any non-date text label as-is (upper) for lookup
                                date_row_map[up_label] = dcell.row

                # Step 4: write data to the matching date rows
                for row_key in sorted_keys:
                    date_str, btype = row_key
                    row_data = date_type_map[row_key]
                    norm_date = _perf_normalize_date(date_str) or date_str.replace("/", "-")
                    # Try date first, then fall back to upper-cased label lookup
                    target_row = date_row_map.get(norm_date) or date_row_map.get(date_str.upper())
                    if target_row is None:
                        continue

                    if type_col is not None:
                        ws.cell(row=target_row, column=type_col).value = btype

                    for fdfs_lbl, (r_col, ur_col, header) in fdfs_col_map.items():
                        entry = row_data.get(fdfs_lbl, {})
                        avg_h = entry.get("avg_hours")
                        avg_m = entry.get("avg_minutes")
                        avg_c = entry.get("avg_count")
                        ur = entry.get("uniform_rate")
                        is_dmp = bool(entry.get("is_dmp"))

                        if r_col is not None:
                            # Choose unit from the column header suffix:
                            #   (h)            → hours
                            #   (m)            → minutes  (DM2000 only)
                            #   (t) for DMP    → cycle count ("số lần phóng điện")
                            #   (t) for DM2000 → minutes (legacy alias of (m))
                            #   no suffix      → DMP → cycle count, DM2000 → hours
                            h_lower = header.lower()
                            if h_lower.endswith("(t)"):
                                val = avg_c if is_dmp else avg_m
                            elif h_lower.endswith("(m)"):
                                val = avg_m
                            elif h_lower.endswith("(h)"):
                                val = avg_h
                            else:
                                # No explicit unit suffix: DMP → cycle count, DM2000 → hours
                                val = avg_c if is_dmp else avg_h
                            ws.cell(row=target_row, column=r_col).value = (
                                val if val is not None else ""
                            )
                        if ur_col is not None:
                            ur_write: float | str = ur if ur is not None else ""
                            if isinstance(ur_write, (int, float)):
                                # Detect percentage formatting from the actual
                                # destination cell — the template tag row often
                                # uses "General" while the data rows below it
                                # are formatted as "0.00%".  Excel expects a
                                # 0-1 fraction in percentage-formatted cells,
                                # so divide accordingly.
                                ur_fmt = (
                                    ws.cell(row=target_row, column=ur_col).number_format
                                    or ""
                                )
                                if "%" not in ur_fmt:
                                    # Fall back to the tag row's format for
                                    # backward compatibility with templates
                                    # that only set the format on the tag row.
                                    ur_fmt = (
                                        ws.cell(row=tag_row_idx, column=ur_col).number_format
                                        or ""
                                    )
                                if "%" in ur_fmt:
                                    ur_write = ur_write / 100.0
                            ws.cell(row=target_row, column=ur_col).value = ur_write

                # Step 5: clear template tags from the tag row if that row's
                # date was not among the exported data dates (so stale tags
                # don't appear in the output).
                exported_dates = {
                    _perf_normalize_date(ds) or ds.replace("/", "-")
                    for ds, _ in sorted_keys
                }
                tag_row_date = _perf_normalize_date(ws.cell(row=tag_row_idx, column=1).value)
                if tag_row_date not in exported_dates:
                    for trow in ws.iter_rows(min_row=tag_row_idx, max_row=tag_row_idx):
                        for tcell in trow:
                            if isinstance(tcell.value, str) and (
                                "{{TYPE}}" in tcell.value
                                or re.search(r"\{\{(RESULT|RATE)_\d+\}\}", tcell.value)
                            ):
                                tcell.value = ""

        # Replace {{PERF_SHEET_NAME}} anywhere in the sheet
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{PERF_SHEET_NAME}}" in cell.value:
                    cell.value = cell.value.replace("{{PERF_SHEET_NAME}}", sheet_name)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def compute_dm2000_stats(rows: list[dict]) -> dict:
    """Compute DM2000 curve statistics. TIM is already in minutes."""
    def safe_float(v):
        if v is None or v == "--" or v == "":
            return None
        try:
            f = float(v)
            return None if math.isnan(f) else f
        except (TypeError, ValueError):
            return None

    volt_vals = [safe_float(r.get("VOLT") or r.get("volt")) for r in rows]
    volt_vals = [v for v in volt_vals if v is not None]

    tim_vals = [safe_float(r.get("TIM") or r.get("tim")) for r in rows]
    tim_vals = [t for t in tim_vals if t is not None]

    def agg(vals):
        if not vals:
            return {"max": None, "min": None, "avg": None}
        return {
            "max": round(max(vals), 4),
            "min": round(min(vals), 4),
            "avg": round(sum(vals) / len(vals), 4),
        }

    v = agg(volt_vals)
    duration = max(tim_vals) if tim_vals else None
    return {
        "VOLT_MAX": v["max"],
        "VOLT_MIN": v["min"],
        "VOLT_AVG": v["avg"],
        "DURATION_MIN": duration,
    }


def _compute_average_curve(rows: list[dict]) -> list[dict]:
    """Group by TIM and average VOLT values."""
    from collections import defaultdict
    tim_groups: dict[float, list[float]] = defaultdict(list)
    for row in rows:
        tim = row.get("TIM") or row.get("tim")
        volt = row.get("VOLT") or row.get("volt")
        try:
            t = float(tim)
            v = float(volt)
            if not math.isnan(t) and not math.isnan(v):
                tim_groups[t].append(v)
        except (TypeError, ValueError):
            continue
    return [
        {"TIM": t, "VOLT": round(sum(vs) / len(vs), 6)}
        for t, vs in sorted(tim_groups.items())
    ]


def _compute_sot_mah_from_tav(tav_rows: list, load_r_ohm: float, fcv=None) -> float | None:
    """Compute approximate SOt mAh by trapezoidal integration over voltage thresholds.

    tav_rows: list of dicts with 'sj'/'SJ' (voltage threshold, V) and
              'minutes'/'MINUTES' (cumulative time from discharge start, min).
    load_r_ohm: load resistance in Ohms.
    fcv: final closed-circuit voltage in V (optional). When provided, an extra
         segment from FCV at t=0 to the first threshold is included, which gives
         a more accurate total capacity.

    Returns SOt in mAh, or None if the data is insufficient.
    """
    if not tav_rows or not load_r_ohm or load_r_ohm <= 0:
        return None

    points: list[tuple] = []
    for row in tav_rows:
        sj = row.get("sj") or row.get("SJ")
        mins = row.get("minutes") or row.get("MINUTES")
        try:
            v = float(sj)
            t = float(mins)
            if not math.isnan(v) and not math.isnan(t) and t >= 0:
                points.append((v, t))
        except (TypeError, ValueError):
            continue

    if len(points) < 2:
        return None

    points.sort(key=lambda x: x[1])  # sort by time ascending

    total_mah = 0.0

    # Include initial segment from FCV (at t=0) to the first threshold
    if fcv is not None:
        try:
            fcv_f = float(fcv)
            if not math.isnan(fcv_f) and points[0][1] > 0 and fcv_f > points[0][0]:
                dt_hours = points[0][1] / 60.0
                v_avg = (fcv_f + points[0][0]) / 2.0
                total_mah += (v_avg / load_r_ohm) * 1000.0 * dt_hours
        except (TypeError, ValueError):
            pass

    # Sum all threshold-to-threshold segments
    for i in range(len(points) - 1):
        v1, t1 = points[i]
        v2, t2 = points[i + 1]
        if t2 > t1:
            dt_hours = (t2 - t1) / 60.0
            v_avg = (v1 + v2) / 2.0
            total_mah += (v_avg / load_r_ohm) * 1000.0 * dt_hours

    return total_mah if total_mah > 0 else None


def _derive_thresholds_from_curves(
    curves_by_baty: dict, endpoint_voltage: Optional[float] = None
) -> list[float]:
    """Derive a sensible set of voltage thresholds for DMP discharge curves.

    Returns thresholds in descending order, snapped to 0.05 V steps from
    the maximum starting voltage down to ``endpoint_voltage`` (or the lowest
    voltage observed across all curves when no endpoint is provided).
    """
    max_v: Optional[float] = None
    min_v: Optional[float] = None
    for rows in curves_by_baty.values():
        for r in rows or []:
            v = r.get("VOLT")
            if v is None or v == "" or v == "--":
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            if math.isnan(fv):
                continue
            if max_v is None or fv > max_v:
                max_v = fv
            if min_v is None or fv < min_v:
                min_v = fv
    if max_v is None or min_v is None:
        return []
    low_bound = endpoint_voltage if endpoint_voltage is not None else min_v
    if low_bound > max_v:
        low_bound = min_v
    # Snap to 0.05 V grid for clean display rows
    step = 0.05
    top = math.floor(max_v / step) * step
    bot = math.ceil(low_bound / step) * step
    if bot > top:
        bot = top
    thresholds: list[float] = []
    v = top
    # Cap the number of rows so the report stays readable for very long curves
    max_rows = 80
    while v >= bot - 1e-9 and len(thresholds) < max_rows:
        thresholds.append(round(v, 3))
        v -= step
    return thresholds


def _tav_from_dmp_telemetry(telemetry: list[dict], thresholds: list[float]) -> list[dict]:
    """Build time-at-voltage rows from DMP telemetry for the given thresholds.

    For each threshold V the function returns the cumulative discharge time
    (in minutes) at which the curve first drops to or below V, interpolating
    linearly between the two surrounding telemetry samples.

    Returns rows shaped like ls_vtime: ``[{"sj": V, "minutes": M}, ...]``.
    Thresholds that are never reached (curve still above V at the end of the
    run) are omitted.
    """
    points: list[tuple[float, float]] = []  # (TIM hours, VOLT)
    for r in telemetry or []:
        t = r.get("TIM")
        v = r.get("VOLT") or r.get("volt") or r.get("Volt")
        if t is None or v is None or t == "--" or v == "--":
            continue
        try:
            tf = float(t)
            vf = float(v)
        except (TypeError, ValueError):
            continue
        if math.isnan(tf) or math.isnan(vf):
            continue
        points.append((tf, vf))
    if len(points) < 2:
        return []
    points.sort(key=lambda p: p[0])

    rows: list[dict] = []
    for thr in thresholds:
        cross_t: Optional[float] = None
        for i in range(1, len(points)):
            t1, v1 = points[i - 1]
            t2, v2 = points[i]
            if v1 >= thr and v2 <= thr:
                if v1 == v2:
                    cross_t = t1
                else:
                    # Linear interpolation: solve for t where v(t) = thr
                    cross_t = t1 + (t2 - t1) * ((v1 - thr) / (v1 - v2))
                break
        if cross_t is not None and cross_t >= 0:
            rows.append({"sj": round(thr, 3), "minutes": round(cross_t * 60.0, 4)})
    return rows


def _count_at_volt_from_dmp_telemetry(
    telemetry: list[dict], thresholds: list[float]
) -> list[dict]:
    """Build sample-count rows from DMP telemetry for the given thresholds.

    For each threshold V, returns the 1-based position (along the time-sorted
    telemetry) of the first sample whose voltage is ``<= V``. Equivalently,
    the number of "times" / discharge samples accumulated from the start of
    the run up to and including the sample that crosses the threshold. This
    matches the DMP-1 sample report's "Unit: times" column.

    Returns rows shaped as ``[{"sj": V, "count": N}, ...]``. Thresholds that
    are never reached (curve still above V at end of run) are omitted.
    """
    points: list[tuple[float, float]] = []  # (TIM, VOLT)
    for r in telemetry or []:
        t = r.get("TIM")
        v = r.get("VOLT") or r.get("volt") or r.get("Volt")
        if t is None or v is None or t == "--" or v == "--":
            continue
        try:
            tf = float(t)
            vf = float(v)
        except (TypeError, ValueError):
            continue
        if math.isnan(tf) or math.isnan(vf):
            continue
        points.append((tf, vf))
    if len(points) < 2:
        return []
    points.sort(key=lambda p: p[0])

    rows: list[dict] = []
    for thr in thresholds:
        cross_idx: Optional[int] = None
        for i in range(1, len(points)):
            v1 = points[i - 1][1]
            v2 = points[i][1]
            if v1 >= thr and v2 <= thr:
                cross_idx = i + 1  # 1-based ordinal of the sample at/after crossing
                break
        if cross_idx is not None:
            rows.append({"sj": round(thr, 3), "count": cross_idx})
    return rows



def _get_pam2_ocv_fcv(archname: str, baty: int) -> dict | None:
    """Get OCV (open-circuit voltage) and FCV (final closed-circuit voltage) for
    a single battery.

    Tries ls_pam2 first (archname-based schema where ocv/fcv columns exist),
    then falls back to ls_evolt (cdid-based schema) where the DM2000 stores the
    initial OCV and FCV measurements as dedicated rows with dy='OCV' / dy='FCV'
    and per-pin voltages in volt1..volt9 columns.

    Returns a dict with VOLT_MAX (OCV) and VOLT_MIN (FCV), or None if the
    data is unavailable.
    """
    def safe_float(v):
        if v is None or v in ("--", ""):
            return None
        try:
            f = float(v)
            return None if math.isnan(f) else f
        except (TypeError, ValueError):
            return None

    ocv = None
    fcv = None

    # --- Try ls_pam2 first (archname-based schema) ---
    # Query all pam2 rows for this archive and filter by battery position in
    # Python to handle both integer and text storage of the gpp/baty column.
    try:
        rows = _read_dm2000_ls_multi([
            ("SELECT * FROM ls_pam2 WHERE cdid = ?", (archname,)),
            ("SELECT * FROM ls_pam2 WHERE archname = ?", (archname,)),
        ])
        for r in rows:
            gpp_val = _dm2000_get_value(r, "gpp", "baty")
            try:
                if gpp_val is not None and int(float(str(gpp_val))) == baty:
                    ocv = safe_float(_dm2000_get_value(r, "ocv", "OCV"))
                    fcv = safe_float(_dm2000_get_value(r, "fcv", "FCV"))
                    break
            except (TypeError, ValueError):
                pass
    except (pyodbc.Error, HTTPException):
        pass

    # --- Fallback: read from ls_evolt (cdid-based schema) ---
    # In the cdid-based schema the DM2000 instrument stores the open-circuit
    # voltage (measured before discharge) and the loaded FCV (measured at the
    # start of discharge) as special rows in ls_evolt with dy='OCV' / dy='FCV'.
    # Each row has per-pin voltages in volt1..volt9.
    if (ocv is None or fcv is None) and 1 <= baty <= 9:
        volt_col = f"volt{baty}"
        try:
            evolt_rows = _read_dm2000_ls(
                f"SELECT dy, {volt_col} AS val FROM ls_evolt"
                f" WHERE cdid = ? AND (dy = 'OCV' OR dy = 'FCV')",
                (archname,),
            )
            for er in evolt_rows:
                dy = str(er.get("dy") or "").strip().upper()
                val = safe_float(er.get("val"))
                if dy == "OCV" and ocv is None:
                    ocv = val
                elif dy == "FCV" and fcv is None:
                    fcv = val
        except (pyodbc.Error, HTTPException):
            pass

    if ocv is None and fcv is None:
        return None

    return {
        "VOLT_MAX": ocv,
        "VOLT_MIN": fcv,
    }


def _dm2000_get_value(row: dict, *keys):
    # DM2000 Access databases use "--" as a null/empty indicator for many fields.
    # The DM2000 software also writes the string "None" for empty/unconfigured
    # fields (e.g. sbmc='None'), so treat that string as null as well.
    _NULL_LIKE = (None, "", "--", "None", "none")
    for key in keys:
        if key in row and row.get(key) not in _NULL_LIKE:
            return row.get(key)
    lowered = {str(k).lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value not in _NULL_LIKE:
            return value
    return None


def _is_valid_template_name(name: str) -> bool:
    if not isinstance(name, str):
        return False
    if "/" in name or "\\" in name:
        return False
    if not name.endswith(".xlsx"):
        return False
    stem = name[:-5]
    return bool(stem) and all(ch.isalnum() or ch in "_-" for ch in stem)


def _validate_dm2000_archname(archname: str) -> None:
    # Allow alphanumeric, underscore, dot, hyphen, and forward slash (archname-based
    # schema archives can include dates like "QC2026/4/18").  Null bytes, backslashes,
    # and other shell-special characters are still rejected.
    if not re.match(r'^[A-Za-z0-9_./ \-]+$', archname):
        raise HTTPException(status_code=400, detail="Invalid archname")


def _cache_set_with_cap(cache: dict, key, value) -> None:
    cache[key] = value
    if len(cache) <= _DM2000_CACHE_MAX_ENTRIES:
        return
    oldest_key = min(cache.items(), key=lambda item: item[1][1])[0]
    cache.pop(oldest_key, None)


def _read_dm2000_curve_rows(archname: str, baty: int) -> list[dict]:
    cache_key = (archname, baty)
    with _DM2000_CURVE_CACHE_LOCK:
        cached = _DM2000_CURVE_CACHE.get(cache_key)
        if cached is not None:
            rows, ts = cached
            if time.time() - ts < _DM2000_CURVE_CACHE_TTL:
                return rows

    if baty <= 0 or baty > 99:
        raise HTTPException(status_code=400, detail="Invalid baty")
    time_col = f"time{baty}"
    # Try both schema variants in a single shadow copy to avoid duplicate file
    # copies when the primary archname-based schema is not present.
    try:
        raw = _read_dm2000_ls_multi([
            ("SELECT TIM, VOLT FROM ls_vtime WHERE archname = ? AND baty = ? ORDER BY TIM ASC", (archname, baty)),
            (f"SELECT dy, {time_col} AS TIM FROM ls_vtime WHERE cdid = ? ORDER BY {time_col} ASC", (archname,)),
        ])
    except (pyodbc.Error, HTTPException):
        raw = []

    # Detect which schema was returned and normalise to {TIM, VOLT} dicts.
    if raw and "dy" in raw[0]:
        # cdid-based schema: dy = voltage threshold, TIM = time{baty} for this battery.
        # ls_vtime may have duplicate rows for the same voltage threshold (multiple
        # discharge sessions stored in the same archive).  Group by voltage (dy) and
        # average the TIM values so that each threshold appears exactly once, giving
        # a single smooth curve instead of multiple overlapping lines.
        volt_groups: dict[float, list[float]] = {}
        for row in raw:
            tim = row.get("TIM")
            volt = row.get("dy")
            try:
                t = float(tim)
                v = float(volt)
                if not math.isnan(t) and not math.isnan(v):
                    volt_groups.setdefault(v, []).append(t)
            except (TypeError, ValueError):
                continue
        rows = sorted(
            [{"TIM": sum(ts) / len(ts), "VOLT": v} for v, ts in volt_groups.items()],
            key=lambda r: r["TIM"],
        )
    else:
        # archname-based schema: full time-series rows with TIM and VOLT columns.
        # Apply the same deduplication used for the average curve: group by TIM and
        # average VOLT to collapse any duplicate measurements from multiple sessions.
        rows = _compute_average_curve(raw)

    with _DM2000_CURVE_CACHE_LOCK:
        _cache_set_with_cap(_DM2000_CURVE_CACHE, cache_key, (rows, time.time()))
    return rows


def _read_dm2000_average_curve_rows(archname: str) -> list[dict]:
    cache_key = ("avg", archname)
    with _DM2000_CURVE_CACHE_LOCK:
        cached = _DM2000_CURVE_CACHE.get(cache_key)
        if cached is not None:
            rows, ts = cached
            if time.time() - ts < _DM2000_CURVE_CACHE_TTL:
                return rows

    # Try both schema variants in a single shadow copy to avoid two separate file copies.
    try:
        raw = _read_dm2000_ls_multi([
            ("SELECT baty, TIM, VOLT FROM ls_vtime WHERE archname = ? ORDER BY baty ASC, TIM ASC", (archname,)),
            (
                "SELECT dy, time1, time2, time3, time4, time5, time6, time7, time8, time9 FROM ls_vtime WHERE cdid = ?",
                (archname,),
            ),
        ])
    except (pyodbc.Error, HTTPException):
        raw = []

    # Detect which schema was returned: cdid-based rows contain a "dy" voltage column.
    if raw and "dy" in raw[0]:
        flattened: list[dict] = []
        for row in raw:
            volt = row.get("dy")
            for idx in range(1, 10):
                tim = row.get(f"time{idx}")
                flattened.append({"TIM": tim, "VOLT": volt})
        avg_rows = _compute_average_curve(flattened)
    else:
        avg_rows = _compute_average_curve(raw)

    with _DM2000_CURVE_CACHE_LOCK:
        _cache_set_with_cap(_DM2000_CURVE_CACHE, cache_key, (avg_rows, time.time()))
    return avg_rows


def _parse_iso_date_param(value: str | None, field_name: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}, expected YYYY-MM-DD") from exc


def _query_vidata_by_channel(mdb_path: str, channel: int) -> list[dict]:
    """
    Query vidata with fallback strategy for Access ODBC parameter quirks.
    """
    base_sql = "SELECT baty, TIM, VOLT, Im FROM vidata WHERE baty = {placeholder} ORDER BY TIM ASC"

    try:
        return query_mdb(mdb_path, base_sql.format(placeholder="?"), (channel,))
    except HTTPException:
        raise
    except Exception as exc1:
        logger.debug("_query_vidata attempt1 (int param) failed: %s", exc1)

    try:
        return query_mdb(mdb_path, base_sql.format(placeholder="?"), (str(channel),))
    except HTTPException:
        raise
    except Exception as exc2:
        logger.debug("_query_vidata attempt2 (str param) failed: %s", exc2)

    try:
        return query_mdb(
            mdb_path,
            f"SELECT baty, TIM, VOLT, Im FROM vidata WHERE baty = {int(channel)} ORDER BY TIM ASC",
        )
    except HTTPException:
        raise
    except Exception as exc3:
        logger.debug("_query_vidata attempt3 (inline int) failed: %s", exc3)

    try:
        return query_mdb(
            mdb_path,
            f"SELECT baty, TIM, VOLT, Im FROM vidata WHERE baty = '{int(channel)}' ORDER BY TIM ASC",
        )
    except HTTPException:
        raise
    except Exception as exc4:
        logger.debug("_query_vidata attempt4 (inline str) failed: %s", exc4)
        raise HTTPException(
            status_code=500,
            detail=f"Cannot query vidata for channel {channel}: {exc4}",
        ) from exc4


def _read_telemetry(cdmc: str, channel: int) -> list[dict]:
    cache_key = (cdmc, channel)
    with _TELEMETRY_CACHE_LOCK:
        cached = _TELEMETRY_CACHE.get(cache_key)
        if cached is not None:
            rows, ts = cached
            if time.time() - ts < _TELEMETRY_CACHE_TTL:
                logger.debug("_read_telemetry cache hit: cdmc=%r channel=%d", cdmc, channel)
                return rows

    mdb_path = resolve_data_file(cdmc)
    try:
        with shadow_copy(mdb_path) as copied:
            rows = _query_vidata_by_channel(copied, channel)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("_read_telemetry: unexpected error for cdmc=%r channel=%d", cdmc, channel)
        raise HTTPException(status_code=500, detail=f"Unexpected error reading telemetry: {exc}") from exc
    for row in rows:
        tim = row.get("TIM")
        if tim is not None and tim != "--":
            try:
                row["TIM"] = round(float(tim) / 3600, 6)
            except (TypeError, ValueError):
                pass
    with _TELEMETRY_CACHE_LOCK:
        _TELEMETRY_CACHE[cache_key] = (rows, time.time())
    return rows


def _get_table_columns(mdb_path: str, table_name: str) -> set[str]:
    if table_name.lower() not in _SCHEMA_TABLE_WHITELIST:
        raise HTTPException(
            status_code=400,
            detail=f"Table '{table_name}' is not whitelisted for schema lookup",
        )
    conn_str = (
        r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
        f"DBQ={mdb_path};"
    )
    with _acquire_query_lock():
        with pyodbc.connect(conn_str, timeout=10) as conn:
            cursor = conn.cursor()
            return {
                str(row.column_name).lower()
                for row in cursor.columns(table=table_name)
                if getattr(row, "column_name", None)
            }


@app.get("/")
def health_check():
    return {"status": "ok", "service": "DMP Battery Data Bridge"}


@app.get("/batches/years")
def get_batch_years():
    """Return a sorted list of distinct years found in para_pub.fdrq."""
    try:
        rows = _read_dmpdata("SELECT fdrq FROM para_pub")
    except (pyodbc.Error, HTTPException) as exc:
        logger.warning("get_batch_years: query failed: %s", exc)
        return {"years": []}

    years: set[int] = set()
    for row in rows:
        fdrq = row.get("fdrq")
        if fdrq is None:
            continue
        try:
            if hasattr(fdrq, "year"):
                years.add(int(fdrq.year))
            else:
                yr = int(str(fdrq)[:4])
                if 1990 <= yr <= 2100:
                    years.add(yr)
        except (TypeError, ValueError):
            pass
    return {"years": sorted(years, reverse=True)}


@app.get("/batches")
def get_batches(year: Optional[int] = None):
    try:
        rows = _read_dmpdata("SELECT * FROM para_pub ORDER BY fdrq DESC")
    except pyodbc.Error:
        try:
            rows = _read_dmpdata("SELECT id, dcxh, fdrq, fdfs FROM para_pub ORDER BY fdrq DESC")
        except pyodbc.Error as exc:
            logger.error("get_batches: fallback query also failed: %s", exc)
            raise HTTPException(status_code=500, detail="Database query failed") from exc

    # DMP databases use a single dash "-" as a null/empty placeholder for string
    # fields (distinct from the DM2000 double-dash "--").  The DMP software also
    # writes the integer/string 0 for flag fields like para_singl.smark when no
    # remark has been set, and the string "None" for unconfigured text fields.
    # This helper returns True when a value should be treated as absent so that
    # fallback logic can run.
    def _dmp_is_empty(v) -> bool:
        if v is None:
            return True
        return str(v).strip() in ("", "-", "--", "0", "None", "none")

    # Build a channel-count map from para_singl in a single query so every batch
    # row can be annotated without an N+1 per-batch lookup.
    # COUNT(*) is used so the query never depends on a specific column name.
    # Some Access ODBC versions do not honour the alias and return the aggregate
    # column under a generated name (e.g. "Expr1000").  When the alias lookup
    # fails we fall back to the second column by ordinal position.
    # _dm2000_get_value is used for all column lookups on pyodbc result rows so
    # that Access databases that return column names in uppercase are handled
    # identically to lowercase schemas.
    channel_counts: dict[str, int] = {}
    singl_cdmc_by_sid: dict[str, str] = {}
    try:
        cc_rows = _read_dmpdata("SELECT sid, COUNT(*) AS channel_count FROM para_singl GROUP BY sid")
        for cc in cc_rows:
            sid = _dm2000_get_value(cc, "sid")
            cnt = _dm2000_get_value(cc, "channel_count")
            # Fallback: some Access ODBC versions ignore the alias and return the
            # aggregate under a generated name; get it by ordinal position instead.
            if cnt is None:
                vals = list(cc.values())
                if len(vals) >= 2:
                    cnt = vals[1]
            if sid is not None and cnt is not None:
                channel_counts[str(sid)] = int(cnt)
    except Exception as exc:  # noqa: BLE001
        logger.warning("get_batches: could not load channel counts from para_singl: %s", exc)

    # Build a first-cdmc lookup from para_singl so the database path can be
    # derived for batches where para_pub.cdmc is NULL.
    # Uses a plain SELECT (no GROUP BY / MIN aggregate) to avoid a Microsoft
    # Access ODBC driver quirk where aggregate column aliases (e.g.
    # "MIN(cdmc) AS cdmc") are silently ignored and the column is returned
    # under its expression form ("Min(cdmc)") rather than the alias.
    try:
        cdmc_rows = _read_dmpdata(
            "SELECT sid, cdmc FROM para_singl WHERE sid IS NOT NULL AND cdmc IS NOT NULL",
        )
        for cr in cdmc_rows:
            sid = _dm2000_get_value(cr, "sid")
            cdmc = _dm2000_get_value(cr, "cdmc")
            # Skip dash-placeholder values used by DMP software for empty fields.
            if sid is not None and cdmc and not _dmp_is_empty(cdmc):
                sid_str = str(sid)
                if sid_str not in singl_cdmc_by_sid:  # keep first occurrence per sid
                    singl_cdmc_by_sid[sid_str] = str(cdmc).strip()
    except Exception as exc:  # noqa: BLE001
        logger.warning("get_batches: could not load para_singl cdmc: %s", exc)

    # Load per-channel extras from para_singl.
    # cdmc  = session archive file name (often equals para_pub.id in many schemas)
    # scdw  = manufacturer, dcph = serial/battery id, dcmc = battery model name,
    # scrq  = sample/start date, smark = per-channel remark.
    # para_singl stores one row per channel; the first non-null value per sid is
    # chosen in Python to avoid the Access ODBC aggregate-alias quirk.
    # Note: jstj is a para_pub column (not para_singl) and is already present
    # in each batch row from the SELECT * above.
    #
    # Two lookup dicts are built from the same rows:
    #   singl_extras_by_sid  — keyed by str(para_singl.sid)
    #   singl_extras_by_cdmc — keyed by para_singl.cdmc (archive name).
    # Many DMP installations use a sequential integer for para_singl.sid that
    # does NOT match para_pub.id.  In those schemas cdmc (the session archive
    # filename) is the same value as para_pub.id, so the secondary lookup by
    # cdmc can resolve the extras even when the sid lookup fails.
    singl_extras_by_sid: dict[str, dict] = {}
    singl_extras_by_cdmc: dict[str, dict] = {}
    try:
        extras_rows = _read_dmpdata(
            "SELECT sid, cdmc, scdw, dcph, dcmc, scrq, smark"
            " FROM para_singl WHERE sid IS NOT NULL",
        )
        for er in extras_rows:
            sid = _dm2000_get_value(er, "sid")
            if sid is None:
                continue
            sid_str = str(sid)

            # Session archive name — used to key the secondary lookup.
            cdmc_er = _dm2000_get_value(er, "cdmc")

            # --- primary lookup: keyed by para_singl.sid ---
            if sid_str not in singl_extras_by_sid:
                singl_extras_by_sid[sid_str] = {}
            entry = singl_extras_by_sid[sid_str]
            for field_name in ("scdw", "dcph", "dcmc", "scrq", "smark"):
                if not entry.get(field_name):
                    field_value = _dm2000_get_value(er, field_name)
                    # Ignore dash-placeholder values; treat them as absent so that
                    # a later channel with a real value can fill the slot.
                    if field_value is not None and not _dmp_is_empty(field_value):
                        entry[field_name] = field_value

            # --- secondary lookup: keyed by para_singl.cdmc (archive name) ---
            if cdmc_er and not _dmp_is_empty(cdmc_er):
                cdmc_key = str(cdmc_er).strip()
                if cdmc_key not in singl_extras_by_cdmc:
                    singl_extras_by_cdmc[cdmc_key] = {}
                entry_c = singl_extras_by_cdmc[cdmc_key]
                for field_name in ("scdw", "dcph", "dcmc", "scrq", "smark"):
                    if not entry_c.get(field_name):
                        field_value = _dm2000_get_value(er, field_name)
                        if field_value is not None and not _dmp_is_empty(field_value):
                            entry_c[field_name] = field_value
    except Exception as exc:  # noqa: BLE001
        logger.warning("get_batches: could not load para_singl extras: %s", exc)

    # Derive channel_counts_by_cdmc from the already-built dicts without an
    # extra query.  This allows the channel count to be found via the archive
    # name when para_singl.sid does not match para_pub.id.
    channel_counts_by_cdmc: dict[str, int] = {
        cdmc_v: channel_counts[sid_k]
        for sid_k, cdmc_v in singl_cdmc_by_sid.items()
        if sid_k in channel_counts
    }

    # Date fields in para_pub that need serialisation to string
    _DATE_FIELDS = ("fdrq", "madedate", "scrq")

    def _to_date_str(value):
        if value is None:
            return None
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        s = str(value)[:10]
        return s if s and s != "None" else None

    result = []
    for row in rows:
        fdrq = row.get("fdrq")
        if fdrq is None:
            row_year = None
            row["fdrq"] = None
        elif hasattr(fdrq, "strftime"):
            row_year = int(fdrq.year)
            row["fdrq"] = fdrq.strftime("%Y-%m-%d")
        else:
            date_str = str(fdrq)[:10]
            row["fdrq"] = date_str
            try:
                row_year = int(date_str[:4])
            except (TypeError, ValueError):
                row_year = None

        # Normalise other date columns that may arrive as datetime objects
        for _df in _DATE_FIELDS:
            if _df == "fdrq":
                continue
            val = row.get(_df)
            if val is not None and hasattr(val, "strftime"):
                row[_df] = val.strftime("%Y-%m-%d")

        # Merge fields from para_singl extras: manufacturer, serial number,
        # battery model name, and sample date.
        batch_id = _dm2000_get_value(row, "id")
        singl_ext = singl_extras_by_sid.get(str(batch_id)) if batch_id is not None else None
        # Fallback 1: some schemas store para_singl.sid as the cdmc session filename,
        # not the numeric batch id — try a cdmc-keyed lookup when the id lookup fails.
        if singl_ext is None:
            _cdmc_key = str(_dm2000_get_value(row, "cdmc") or "").strip()
            if _cdmc_key:
                singl_ext = singl_extras_by_sid.get(_cdmc_key)
        # Fallback 2: try the secondary cdmc-keyed extras dict.  In schemas where
        # para_singl.cdmc (archive name) equals para_pub.id this resolves extras
        # even when para_singl.sid is a different sequential number.
        if singl_ext is None and batch_id is not None:
            singl_ext = singl_extras_by_cdmc.get(str(batch_id))
        if singl_ext:
            scdw_val = _dm2000_get_value(singl_ext, "scdw")
            if scdw_val is not None and not _dmp_is_empty(scdw_val):
                row["manufacturer"] = str(scdw_val).strip()
            # dcmc is the battery model name stored per-channel; use it as the
            # batch name when no name has been set yet.
            dcmc_val_singl = _dm2000_get_value(singl_ext, "dcmc")
            if dcmc_val_singl is not None and not _dmp_is_empty(dcmc_val_singl) and _dmp_is_empty(row.get("name")):
                row["name"] = str(dcmc_val_singl).strip()
            # scrq is the sample/start date from para_singl; use it as madedate.
            scrq_val = _dm2000_get_value(singl_ext, "scrq")
            if scrq_val is not None and _dmp_is_empty(row.get("madedate")):
                row["madedate"] = _to_date_str(scrq_val)

        # Serial No comes from para_singl.dcph.
        if singl_ext:
            dcph_val = _dm2000_get_value(singl_ext, "dcph")
            if dcph_val is not None and not _dmp_is_empty(dcph_val) and _dmp_is_empty(row.get("serialno")):
                row["serialno"] = str(dcph_val).strip()

        # Remark: prefer para_pub.bz (batch-level remark); fall back to
        # para_singl.smark (per-channel remark used by many DMP installations).
        bz_pub = _dm2000_get_value(row, "bz")
        if not _dmp_is_empty(bz_pub):
            row["remarks"] = str(bz_pub).strip()
        elif singl_ext:
            smark_val = _dm2000_get_value(singl_ext, "smark")
            if not _dmp_is_empty(smark_val):
                row["remarks"] = str(smark_val).strip()

        # para_pub does not have a cdmc column.  The session .mdb file name is
        # stored in para_singl.cdmc and was pre-fetched into singl_cdmc_by_sid.
        # Use that as the session identifier; fall back to the batch id itself.
        cdmc_val = singl_cdmc_by_sid.get(str(batch_id), "") if batch_id is not None else ""
        # Treat a dash placeholder from the database as absent so the batch-id
        # fallback below can provide a meaningful session identifier.
        if _dmp_is_empty(cdmc_val) and batch_id is not None:
            cdmc_val = str(batch_id)
        if cdmc_val:
            if _dmp_is_empty(row.get("name")):
                row["name"] = cdmc_val
            if _dmp_is_empty(row.get("database")):
                row["database"] = str(Path(DMP_DATA_DIR) / f"{cdmc_val}.mdb")

        # Expose archname (= batch id) so the frontend keyword search can use it.
        if not row.get("archname") and batch_id is not None:
            row["archname"] = str(batch_id)

        # Attach the pre-computed channel count for this batch.
        # Try by numeric batch id first; fall back to cdmc key for schemas where
        # para_singl.sid stores the session filename rather than the numeric id;
        # finally try channel_counts_by_cdmc when cdmc (archive name) == batch_id.
        _cc = channel_counts.get(str(batch_id)) if batch_id is not None else None
        if _cc is None and cdmc_val:
            _cc = channel_counts.get(cdmc_val)
        if _cc is None and batch_id is not None:
            _cc = channel_counts_by_cdmc.get(str(batch_id))
        row["channel_count"] = _cc

        if year is not None and row_year != year:
            continue
        result.append(row)
    return {"batches": result}


@app.get("/batches/{batch_id}/channels")
def get_channels(batch_id: str):
    """
    Get the channel list for a batch.

    DMP Access schema:
    - para_pub.id = batch_id (batch identifier)
    - para_singl.sid = para_pub.id (JOIN key)
    - para_singl.baty = channel number
    - para_singl.cdmc = session .mdb file name (the source of telemetry data)
    """
    rows: list[dict] = []
    last_error: Exception | None = None
    try:
        rows = _read_dmpdata(
            "SELECT baty, cdmc FROM para_singl WHERE sid = ?",
            (batch_id,),
        )
    except Exception as exc:
        logger.debug("get_channels: query with cdmc failed for %r: %s", batch_id, exc)
        try:
            rows = _read_dmpdata(
                "SELECT baty FROM para_singl WHERE sid = ?",
                (batch_id,),
            )
        except Exception as exc2:
            logger.debug("get_channels: fallback query failed for %r: %s", batch_id, exc2)
            last_error = exc2

    if not rows and last_error is not None:
        raise HTTPException(status_code=500, detail=f"Database query failed: {str(last_error)}")

    return {"channels": rows}


@app.get("/changes")
def get_changes(since: float = 0):
    with _WATCH_LOCK:
        changed = [
            (stem, changed_at)
            for stem, changed_at in _WATCHED_CHANGES.items()
            if changed_at > since
        ]
    changed.sort(key=lambda item: item[1])
    return {"changes": [stem for stem, _ in changed], "timestamp": time.time()}


@app.get("/telemetry")
def get_telemetry(cdmc: str, channel: int):
    return {"telemetry": _read_telemetry(cdmc, channel)}


@app.get("/stats")
def get_stats(cdmc: str, channel: int):
    rows = _read_telemetry(cdmc, channel)
    return compute_stats(rows)


@app.get("/templates")
def get_templates():
    templates_dir = Path(DMP_TEMPLATES_DIR).resolve()
    if not templates_dir.exists():
        return {"templates": []}
    templates = sorted([f.name for f in templates_dir.iterdir() if f.is_file() and f.suffix.lower() == ".xlsx"])
    return {"templates": templates}


@app.post("/report")
def generate_report(payload: ReportRequest):
    batch_rows = _read_dmpdata(
        "SELECT id, dcxh, fdrq, fdfs FROM para_pub WHERE id=?",
        (payload.batch_id,),
    )
    if not batch_rows:
        raise HTTPException(status_code=404, detail="Batch not found")

    telemetry = _read_telemetry(payload.cdmc, payload.channel)
    stats = compute_stats(telemetry)
    batch = batch_rows[0]

    context = {
        "BATCH_ID": batch.get("id"),
        "MODEL": batch.get("dcxh"),
        "DATE": str(batch.get("fdrq")),
        "DISCHARGE_PATTERN": batch.get("fdfs"),
        "CHANNEL": payload.channel,
        **stats,
        "HISTORY_DATA": telemetry,
    }

    template_path = _resolve_template_path(payload.template_name)
    report_bytes = render_excel_template(template_path, context)
    filename = f"dmp_report_{payload.batch_id}_{payload.channel}.xlsx"

    return StreamingResponse(
        BytesIO(report_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/report-simple")
def generate_dmp_simple_report(payload: DMPSimpleReportRequest):
    """Generate a "Battery Discharge Curve" Excel report for one or more DMP
    channels.

    Two modes are supported:

    * **Multi-battery (DM2000-style preview)** — the request supplies ``batys``
      (1 or more channel numbers). The output uses the shared
      ``_build_preview_workbook`` so the file is identical in shape to the
      DM2000 simple report (archive info header, OCV/FCV/SOt mAh per battery,
      "Duration of Series Designated Voltage" rows, plus the discharge-curve
      chart).
    * **Single-channel (legacy)** — when ``batys`` is empty and ``channel`` is
      supplied, the original tabular dump (stats + raw telemetry) is generated
      so older callers continue to work.
    """
    safe_id = re.sub(r'[^\w\-]', '_', str(payload.batch_id))

    # ── Legacy single-channel mode ─────────────────────────────────────────
    if not payload.batys and payload.channel is not None:
        if not (1 <= payload.channel <= 99):
            raise HTTPException(status_code=400, detail="Invalid channel")
        if not payload.cdmc:
            raise HTTPException(status_code=400, detail="cdmc is required for single-channel report")

        batch_rows = _read_dmpdata(
            "SELECT id, dcxh, fdrq, fdfs FROM para_pub WHERE id = ?",
            (payload.batch_id,),
        )
        if not batch_rows:
            raise HTTPException(status_code=404, detail="Batch not found")
        batch = batch_rows[0]

        telemetry = _read_telemetry(payload.cdmc, payload.channel)
        stats = compute_stats(telemetry)

        duration = None
        if telemetry:
            try:
                duration = max(
                    float(r.get("TIM") or 0)
                    for r in telemetry
                    if r.get("TIM") not in (None, "", "--")
                )
            except (TypeError, ValueError):
                duration = None

        wb = Workbook()
        ws = wb.active
        ws.title = f"CH{payload.channel}"

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 16

        for label, value in [
            ("Batch ID", str(batch.get("id") or "")),
            ("Type / Model", str(batch.get("dcxh") or "")),
            ("Date", str(batch.get("fdrq") or "")),
            ("Discharge Pattern", str(batch.get("fdfs") or "")),
            ("Channel", str(payload.channel)),
        ]:
            ws.append([label, value])
        ws.append([])

        ws.append(["Statistics"])
        for label, value in [
            ("Duration (h)", round(duration, 4) if duration is not None else None),
            ("Voltage Max (V)", stats.get("VOLT_MAX")),
            ("Voltage Min (V)", stats.get("VOLT_MIN")),
            ("Voltage Avg (V)", stats.get("VOLT_AVG")),
            ("Current Max (mA)", stats.get("IM_MAX")),
            ("Current Min (mA)", stats.get("IM_MIN")),
            ("Current Avg (mA)", stats.get("IM_AVG")),
        ]:
            ws.append([label, value])
        ws.append([])

        ws.append(["#", "Time (h)", "Voltage (V)", "Current (mA)"])
        for idx, row in enumerate(telemetry, 1):
            tim = row.get("TIM")
            volt = row.get("VOLT") or row.get("volt") or row.get("Volt")
            im = row.get("Im") or row.get("IM") or row.get("im")
            ws.append([idx, tim, volt, im])

        buf = BytesIO()
        wb.save(buf)
        filename = f"dmp_report_{safe_id}_{payload.channel}.xlsx"
        return StreamingResponse(
            BytesIO(buf.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── Multi-battery mode (DM2000-style preview) ──────────────────────────
    batys = sorted({int(b) for b in payload.batys if isinstance(b, int) and 1 <= int(b) <= 99})
    if not batys:
        raise HTTPException(status_code=400, detail="batys must not be empty")

    # Fetch batch metadata from para_pub. SELECT * so all available columns are
    # available to map onto archive_fields irrespective of schema variations.
    try:
        batch_rows = _read_dmpdata(
            "SELECT * FROM para_pub WHERE id = ?",
            (payload.batch_id,),
        )
    except pyodbc.Error:
        batch_rows = _read_dmpdata(
            "SELECT id, dcxh, fdrq, fdfs FROM para_pub WHERE id = ?",
            (payload.batch_id,),
        )
    if not batch_rows:
        raise HTTPException(status_code=404, detail="Batch not found")
    batch = batch_rows[0]

    # Fetch para_singl rows for this batch to discover per-channel cdmc and
    # additional metadata (manufacturer, dcph serial, dcmc model name, scrq date).
    try:
        singl_rows = _read_dmpdata(
            "SELECT baty, cdmc, scdw, dcph, dcmc, scrq FROM para_singl WHERE sid = ?",
            (payload.batch_id,),
        )
    except pyodbc.Error:
        singl_rows = []

    cdmc_by_baty: dict[int, str] = {}
    first_singl: dict = {}
    for row in singl_rows:
        try:
            b = int(float(str(row.get("baty"))))
        except (TypeError, ValueError):
            continue
        if b <= 0:
            continue
        cdmc_val = row.get("cdmc")
        if cdmc_val:
            cdmc_by_baty[b] = str(cdmc_val).strip()
        if not first_singl:
            first_singl = {k: v for k, v in row.items() if v is not None}

    def _to_date_text(v):
        if v and hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10] if v not in (None, "") else ""

    def _apply_override(db_val, override_val):
        return override_val if override_val is not None and str(override_val).strip() != "" else db_val

    # Map para_pub / para_singl columns onto the archive_fields dict consumed
    # by _build_preview_workbook.
    archive_fields = {
        "archname": str(batch.get("id") or payload.batch_id),
        "name": str(batch.get("name") or first_singl.get("dcmc") or ""),
        "startdate": _to_date_text(batch.get("fdrq")),
        "enddate": _to_date_text(batch.get("jsrq") or batch.get("fdjssj")),
        "dcxh": str(_apply_override(batch.get("dcxh"), payload.override_battery_type) or ""),
        "fdfs": str(batch.get("fdfs") or ""),
        # DMP does not store a precomputed uniform rate; let the workbook
        # builder compute it from time-at-voltage data.
        "unifrate": "",
        "manufacturer": str(_apply_override(first_singl.get("scdw"), payload.override_manufacturer) or ""),
        "madedate": _to_date_text(first_singl.get("scrq")),
        "serialno": str(first_singl.get("dcph") or ""),
        "remarks": str(batch.get("bz") or ""),
        # Voltage type — para_pub.dylx; falls back to fdlx/jstj for older schemas.
        "voltage_type": str(batch.get("dylx") or batch.get("fdlx") or ""),
        "trademark": str(batch.get("sbmc") or batch.get("trademark") or ""),
        "load_resistance": str(batch.get("fzdz") or batch.get("fz2") or ""),
        "endpoint_voltage": str(batch.get("zzdy") or ""),
        "dis_condition": str(batch.get("jstj") or batch.get("hjwd") or batch.get("wd") or ""),
        "min_duration": str(batch.get("fdts") or ""),
    }

    # Fetch telemetry once per battery; reuse for stats, OCV/FCV, SOt mAh and
    # synthesised time-at-voltage rows.
    telemetry_by_baty: dict[int, list[dict]] = {}
    for b in batys:
        cdmc = cdmc_by_baty.get(b) or payload.cdmc
        if not cdmc:
            telemetry_by_baty[b] = []
            continue
        try:
            telemetry_by_baty[b] = _read_telemetry(cdmc, b)
        except HTTPException:
            telemetry_by_baty[b] = []
        except Exception as exc:  # noqa: BLE001
            logger.debug("DMP report: telemetry read failed for baty=%d: %s", b, exc)
            telemetry_by_baty[b] = []

    # Endpoint voltage parsed from para_pub.zzdy, used to bound thresholds.
    ep_voltage: Optional[float] = None
    raw_ep = (archive_fields["endpoint_voltage"] or "").strip()
    if raw_ep:
        parts = raw_ep.split()
        token = (parts[0] if parts else raw_ep)[:32]
        # Strip any trailing unit characters (e.g. "0.9V" → "0.9") without a
        # regex to keep CodeQL happy and avoid any backtracking concern.
        while token and token[-1] not in "0123456789.-":
            token = token[:-1]
        try:
            ep_voltage = float(token)
        except (TypeError, ValueError):
            ep_voltage = None

    thresholds = _derive_thresholds_from_curves(telemetry_by_baty, ep_voltage)

    # Build OCV/FCV/SOt mAh and time-at-voltage data per battery.
    stats_map: dict[int, dict] = {}
    time_at_volt_map: dict[int, list[dict]] = {}
    count_at_volt_map: dict[int, list[dict]] = {}
    battery_params: dict[int, dict] = {}

    try:
        load_r = float(archive_fields["load_resistance"]) if archive_fields["load_resistance"] else None
    except (TypeError, ValueError):
        load_r = None

    for b in batys:
        rows = telemetry_by_baty.get(b) or []
        s = compute_stats(rows)
        # OCV = first voltage sample (open circuit, before discharge starts);
        # FCV = last voltage sample (final closed-circuit). Fall back to
        # max/min when the first/last samples are missing.
        ocv: Optional[float] = None
        fcv: Optional[float] = None
        for r in rows:
            v = r.get("VOLT") or r.get("volt") or r.get("Volt")
            try:
                fv = float(v)
                if not math.isnan(fv):
                    ocv = fv
                    break
            except (TypeError, ValueError):
                continue
        for r in reversed(rows):
            v = r.get("VOLT") or r.get("volt") or r.get("Volt")
            try:
                fv = float(v)
                if not math.isnan(fv):
                    fcv = fv
                    break
            except (TypeError, ValueError):
                continue
        if ocv is None:
            ocv = s.get("VOLT_MAX")
        if fcv is None:
            fcv = s.get("VOLT_MIN")
        s["OCV"] = ocv
        s["FCV"] = fcv
        stats_map[b] = s

        tav = _tav_from_dmp_telemetry(rows, thresholds)
        time_at_volt_map[b] = tav
        count_at_volt_map[b] = _count_at_volt_from_dmp_telemetry(rows, thresholds)
        # Compute SOt mAh: prefer integration of measured current (Im in mA)
        # over time, which is the most accurate. Fall back to TAV-based
        # trapezoidal integration when Im samples are unavailable but a load
        # resistance is configured.
        sot_mah: Optional[float] = None
        # Method 1: ∫ Im dt (mA · h = mAh).  Telemetry TIM is already in hours.
        cur_points: list[tuple[float, float]] = []
        for r in rows:
            t = r.get("TIM")
            im = r.get("Im") or r.get("IM") or r.get("im")
            try:
                tf = float(t)
                im_f = float(im)
            except (TypeError, ValueError):
                continue
            if math.isnan(tf) or math.isnan(im_f):
                continue
            cur_points.append((tf, im_f))
        if len(cur_points) >= 2:
            cur_points.sort(key=lambda p: p[0])
            integ = 0.0
            for i in range(len(cur_points) - 1):
                t1, i1 = cur_points[i]
                t2, i2 = cur_points[i + 1]
                if t2 > t1:
                    integ += (i1 + i2) / 2.0 * (t2 - t1)
            if integ > 0:
                sot_mah = round(integ, 3)
        if sot_mah is None and load_r and load_r > 0:
            sot_mah = _compute_sot_mah_from_tav(tav, load_r, fcv)

        battery_params[b] = {
            "baty": b,
            "ocv": ocv,
            "fcv": fcv,
            "sot_mah": sot_mah,
        }

    try:
        workbook_bytes = _build_preview_workbook(
            archive_fields=archive_fields,
            company=DM2000_COMPANY_NAME or "",
            batys=batys,
            stats_map=stats_map,
            time_at_volt_map=time_at_volt_map,
            battery_params=battery_params,
            endpoint_cutoff=payload.endpoint_cutoff,
            report_kind="dmp",
            telemetry_by_baty=telemetry_by_baty,
            count_at_volt_map=count_at_volt_map,
        )
    except Exception as exc:
        logger.exception("Error building DMP preview workbook for batch=%s: %s", payload.batch_id, exc)
        raise HTTPException(status_code=500, detail=f"Failed to build report: {exc}") from exc

    filename = f"dmp_report_{safe_id}.xlsx"
    return StreamingResponse(
        BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── DM2000 Historic Database Routes ────────────────────────────────────────

@app.get("/dm2000/archives")
def get_dm2000_archives(
    date_from: str = None,
    date_to: str = None,
    type_filter: str = None,
    name_filter: str = None,
    mfr_filter: str = None,
    serial_filter: str = None,
    keyword: str = None,
    limit: Optional[int] = None,
):
    table_names_to_try = ["ls_jb_cs", "ls_pam2", "ls_cs", "ls_jbcs", "ls_archive"]
    rows = None
    cache_key = "all_archives"
    with _DM2000_ARCHIVES_CACHE_LOCK:
        cached = _DM2000_ARCHIVES_CACHE.get(cache_key)
        if cached is not None:
            cached_rows, ts = cached
            if time.time() - ts < _DM2000_ARCHIVES_CACHE_TTL:
                rows = cached_rows

    if rows is None:
        # Use a single shadow copy to try every candidate table name, avoiding
        # the O(n) shadow-copy cost that previously caused proxy timeouts.
        try:
            rows = _read_dm2000_ls_multi([
                (f"SELECT * FROM {t}", ()) for t in table_names_to_try
            ])
        except (pyodbc.Error, HTTPException):
            rows = None
        if rows is None:
            return {"archives": [], "total": 0, "warning": "Archive table not found in dmdata_ls.mdb"}
        with _DM2000_ARCHIVES_CACHE_LOCK:
            _cache_set_with_cap(_DM2000_ARCHIVES_CACHE, cache_key, (rows, time.time()))

    date_from_parsed = _parse_iso_date_param(date_from, "date_from")
    date_to_parsed = _parse_iso_date_param(date_to, "date_to")

    def _to_date_text(value):
        if value and hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return str(value)[:10] if value not in (None, "") else ""

    archives = []
    for row in rows:
        item = {
            "archname": _dm2000_get_value(row, "archname", "cdid", "id"),
            "startdate": _to_date_text(_dm2000_get_value(row, "startdate", "fdrq")),
            "enddate": _to_date_text(_dm2000_get_value(row, "enddate", "fzdq", "jssj", "endrq", "endate", "end_date", "fzrq", "stopdate", "fdend")),
            "dcxh": _dm2000_get_value(row, "dcxh"),
            "name": _dm2000_get_value(row, "name", "dcmc"),
            "fdfs": _dm2000_get_value(row, "fdfs"),
            "duration": _dm2000_get_value(row, "duration", "fdts"),
            # unifrate: try percentage column (hl=合格率) first, then the integer
            # index (yfws=匀放系数 0-9).  DM2000 may store either depending on version.
            "unifrate": _dm2000_get_value(row, "unifrate", "hl", "hlfd", "yfws_pct", "yfws"),
            "manufacturer": _dm2000_get_value(row, "manufacturer", "scdw"),
            "madedate": _to_date_text(_dm2000_get_value(row, "madedate", "scrq")),
            "serialno": _dm2000_get_value(row, "serialno", "dcph", "ph", "scph", "pch", "lot", "lot_no", "batchno", "batch_no"),
            "remarks": _dm2000_get_value(row, "remarks", "remark", "bz", "note", "memo", "bzh"),
            # Additional fields for report preview.
            # Multiple aliases cover different DM2000 schema versions.
            "voltage_type": _dm2000_get_value(
                row,
                "dylx", "voltage_type", "bcdv", "dcdy", "dxy", "edy",
                "nominal_voltage", "dianxin_leixing", "dianxin", "nominal_v",
                "lxdy", "vtype", "battv", "lx", "dctype", "v_type", "jstj",
            ),
            "trademark": _dm2000_get_value(row, "trademark", "shangbiao", "sbmc", "pinpai"),
            "load_resistance": _dm2000_get_value(row, "load_resistance", "fzdz", "fzlkdz", "dw"),
            "endpoint_voltage": _dm2000_get_value(
                row,
                "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
                "endpoint_v", "vcut", "cutoffv", "cutoff_v",
                "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
            ),
            "dis_condition": _dm2000_get_value(
                row,
                "dis_condition", "wd", "fdwd", "hjwd", "wendu",
                "fdtj", "hjtj", "temperature", "temp_c",
                "temp", "hjt", "qw", "t", "csh", "jchj",
            ),
            "min_duration": _dm2000_get_value(row, "min_duration", "zdts", "min_ts", "minduration", "zdsc", "zxfdts"),
            "company": DM2000_COMPANY_NAME or None,
        }
        # Build database file path from archname and data directory
        archname_val = item.get("archname") or ""
        item["database"] = str(Path(DM2000_DATA_DIR) / f"{archname_val}.mdb") if archname_val else None
        archives.append(item)

    def _contains(value, pattern):
        if not pattern:
            return True
        return pattern.lower() in str(value or "").lower()

    filtered = []
    for row in archives:
        row_date = None
        if row["startdate"]:
            try:
                row_date = date.fromisoformat(str(row["startdate"])[:10])
            except ValueError:
                row_date = None
        if date_from_parsed and row_date and row_date < date_from_parsed:
            continue
        if date_to_parsed and row_date and row_date > date_to_parsed:
            continue
        if not _contains(row.get("dcxh"), type_filter):
            continue
        if not _contains(row.get("name"), name_filter):
            continue
        if not _contains(row.get("manufacturer"), mfr_filter):
            continue
        if not _contains(row.get("serialno"), serial_filter):
            continue
        if keyword:
            kw = keyword.lower()
            if not any(
                kw in str(row.get(field) or "").lower()
                for field in ("dcxh", "name", "manufacturer", "serialno", "archname", "remarks")
            ):
                continue
        filtered.append(row)

    filtered.sort(key=lambda r: str(r.get("startdate") or ""), reverse=True)
    if limit is not None and limit > 0:
        filtered = filtered[:limit]
    return {"archives": filtered, "total": len(filtered)}


def _derive_dm2000_batteries_from_vtime(archname: str) -> list[dict]:
    """Fallback: derive active battery channels from ls_vtime time1..time9 columns.

    For cdid-based ls_vtime, each row holds one voltage threshold with elapsed
    time per battery in time{n}. A battery channel is considered active when at
    least one of its time{n} values is non-null/non-empty.

    For archname-based ls_vtime (which uses a baty column instead of time1..time9),
    falls back to querying DISTINCT baty values directly.
    """
    select_cols = ", ".join(f"time{i}" for i in range(1, 10))
    try:
        rows = _read_dm2000_ls(
            f"SELECT {select_cols} FROM ls_vtime WHERE cdid = ?",
            (archname,),
        )
    except pyodbc.Error:
        # cdid-based schema failed; try archname-based schema
        try:
            baty_rows = _read_dm2000_ls(
                "SELECT DISTINCT baty FROM ls_vtime WHERE archname = ? ORDER BY baty ASC",
                (archname,),
            )
        except pyodbc.Error:
            return []
        result = []
        for row in baty_rows:
            val = _dm2000_get_value(row, "baty")
            try:
                num = int(float(val))
                if num > 0:
                    result.append({"baty": num})
            except (TypeError, ValueError):
                pass
        return result

    active: set[int] = set()
    for row in rows:
        for i in range(1, 10):
            value = row.get(f"time{i}")
            if value in (None, "", "--"):
                continue
            try:
                num = float(value)
            except (TypeError, ValueError):
                continue
            if math.isnan(num):
                continue
            active.add(i)
    return [{"baty": i} for i in sorted(active)]


@app.get("/dm2000/archives/{archname}/batteries")
def get_dm2000_batteries(archname: str):
    _validate_dm2000_archname(archname)

    with _DM2000_BATTERIES_CACHE_LOCK:
        cached = _DM2000_BATTERIES_CACHE.get(archname)
        if cached is not None:
            rows, ts = cached
            if time.time() - ts < _DM2000_BATTERIES_CACHE_TTL:
                return {"batteries": rows, "archname": archname}

    try:
        rows = _read_dm2000_ls_multi([
            ("SELECT * FROM ls_pam2 WHERE archname = ? ORDER BY baty ASC", (archname,)),
            ("SELECT * FROM ls_pam2 WHERE cdid = ? ORDER BY gpp ASC", (archname,)),
        ])
    except (pyodbc.Error, HTTPException):
        rows = []
    for row in rows:
        if "baty" not in row:
            row["baty"] = _dm2000_get_value(row, "baty", "gpp")
        # Normalise OCV / FCV / SOt to well-known keys so the frontend can
        # find them regardless of the original Access column name casing or
        # abbreviation variant used by this DM2000 version.
        if "ocv" not in row:
            row["ocv"] = _dm2000_get_value(row, "ocv", "OCV")
        if "fcv" not in row:
            row["fcv"] = _dm2000_get_value(row, "fcv", "FCV")
        # SOt (actual discharged capacity in mAh).  DM2000 uses several
        # column names across versions: sh (实耗), fdrl (放电容量), rl (容量),
        # rql, sot, capacity, sl, sc, dcrl, fdl.
        row["sot_mah"] = _dm2000_get_value(
            row,
            "sh", "sot", "SOT", "sot_mah", "sotmah",
            "rql", "fdrl", "rl", "dcrl", "capacity", "sl", "sc", "fdl",
            "fdsh", "sh_mah", "fdmah", "actual_cap", "shrc", "fdrc",
        )

    # If ls_pam2 has no usable rows (e.g. discharge in progress, or pam2 not
    # populated for this cdid), derive the battery list from ls_vtime so the
    # dropdown still shows each individual pin instead of being empty.
    def _baty_int(value):
        try:
            num = int(float(value))
            return num if num > 0 else None
        except (TypeError, ValueError):
            return None

    has_battery = any(_baty_int(row.get("baty")) is not None for row in rows)
    if not has_battery:
        rows = _derive_dm2000_batteries_from_vtime(archname)

    # Supplement OCV / FCV from ls_evolt for any battery still missing them.
    # In the cdid-based schema the DM2000 stores the pre-discharge OCV and
    # the initial loaded FCV as dedicated rows in ls_evolt (dy='OCV'/'FCV')
    # with per-pin voltages in volt1..volt9.  ls_pam2 does not carry these
    # per-pin measured values in the cdid schema.
    try:
        evolt_ocv_fcv = _read_dm2000_ls(
            "SELECT dy, volt1, volt2, volt3, volt4, volt5, volt6, volt7, volt8, volt9"
            " FROM ls_evolt WHERE cdid = ? AND (dy = 'OCV' OR dy = 'FCV')",
            (archname,),
        )
        evolt_map: dict[int, dict[str, float]] = {}
        for er in evolt_ocv_fcv:
            dy = str(er.get("dy") or "").strip().upper()
            if dy not in ("OCV", "FCV"):
                continue
            for i in range(1, 10):
                raw = er.get(f"volt{i}")
                if raw in (None, "", "--"):
                    continue
                try:
                    fv = float(raw)
                    if not math.isnan(fv):
                        evolt_map.setdefault(i, {})[dy] = fv
                except (TypeError, ValueError):
                    pass
    except (pyodbc.Error, HTTPException):
        evolt_map = {}

    if evolt_map:
        for row in rows:
            b = _baty_int(row.get("baty"))
            if b is None or b not in evolt_map:
                continue
            if row.get("ocv") is None:
                row["ocv"] = evolt_map[b].get("OCV")
            if row.get("fcv") is None:
                row["fcv"] = evolt_map[b].get("FCV")

    with _DM2000_BATTERIES_CACHE_LOCK:
        _cache_set_with_cap(_DM2000_BATTERIES_CACHE, archname, (rows, time.time()))
    return {"batteries": rows, "archname": archname}


@app.get("/dm2000/archives/{archname}/curve")
def get_dm2000_curve(archname: str, baty: int):
    _validate_dm2000_archname(archname)
    rows = _read_dm2000_curve_rows(archname, baty)
    # Prepend the initial loaded voltage (FCV) at t=0 so the chart shows the
    # full discharge curve starting from the actual measured starting voltage.
    if not rows or rows[0].get("TIM", 0) > 0:
        pam2 = _get_pam2_ocv_fcv(archname, baty)
        if pam2:
            fcv = pam2.get("VOLT_MIN")  # _get_pam2_ocv_fcv maps FCV → VOLT_MIN, OCV → VOLT_MAX
            if fcv is not None:
                rows = [{"TIM": 0.0, "VOLT": round(float(fcv), 6)}] + rows
    return {"curve": rows, "archname": archname, "baty": baty, "time_unit": "minutes"}


@app.get("/dm2000/archives/{archname}/average-curve")
def get_dm2000_average_curve(archname: str):
    _validate_dm2000_archname(archname)
    avg = _read_dm2000_average_curve_rows(archname)
    return {"curve": avg, "archname": archname, "baty": "average", "time_unit": "minutes"}


@app.get("/dm2000/archives/{archname}/stats")
def get_dm2000_stats(archname: str, baty: int = 0):
    _validate_dm2000_archname(archname)
    cache_key = (archname, baty)
    with _DM2000_STATS_CACHE_LOCK:
        cached = _DM2000_STATS_CACHE.get(cache_key)
        if cached is not None:
            result, ts = cached
            if time.time() - ts < _DM2000_STATS_CACHE_TTL:
                return result

    if baty == 0:
        rows = _read_dm2000_average_curve_rows(archname)
    else:
        rows = _read_dm2000_curve_rows(archname, baty)
    stats = compute_dm2000_stats(rows)
    # Override VOLT_MAX/VOLT_MIN with the true OCV/FCV stored in ls_pam2.
    # ls_pam2 holds the instrument-measured open-circuit voltage (OCV, before
    # discharge) and loaded voltage (FCV, at discharge start) — the values
    # DM2000 reports in its own Excel export.  The ls_vtime curve data uses
    # fixed voltage thresholds, not real measurements.
    if baty > 0:
        pam2_stats = _get_pam2_ocv_fcv(archname, baty)
        if pam2_stats:
            stats.update(pam2_stats)
            # Expose as dedicated keys so the frontend and templates can use
            # OCV/FCV independently of the curve VOLT_MAX/VOLT_MIN stats.
            stats["OCV"] = pam2_stats["VOLT_MAX"]
            stats["FCV"] = pam2_stats["VOLT_MIN"]

    with _DM2000_STATS_CACHE_LOCK:
        _cache_set_with_cap(_DM2000_STATS_CACHE, cache_key, (stats, time.time()))
    return stats


@app.get("/dm2000/archives/{archname}/daily-voltage")
def get_dm2000_daily_voltage(archname: str, baty: int):
    _validate_dm2000_archname(archname)
    try:
        rows = _read_dm2000_ls(
            "SELECT * FROM ls_evolt WHERE archname = ? AND baty = ? ORDER BY date ASC",
            (archname, baty),
        )
        for row in rows:
            d = row.get("date")
            if d and hasattr(d, "strftime"):
                row["date"] = d.strftime("%Y-%m-%d")
        return {"daily_voltage": rows, "archname": archname, "baty": baty}
    except pyodbc.Error:
        if baty <= 0 or baty > 99:
            raise HTTPException(status_code=400, detail="Invalid baty")
        volt_col = f"volt{baty}"
        try:
            rows = _read_dm2000_ls(
                f"SELECT daytime, {volt_col} AS volt FROM ls_evolt WHERE cdid = ? ORDER BY daytime ASC",
                (archname,),
            )
        except pyodbc.Error:
            return {"daily_voltage": [], "archname": archname, "baty": baty}
        normalized = []
        for row in rows:
            d = row.get("daytime")
            v = row.get("volt")
            if d and hasattr(d, "strftime"):
                d = d.strftime("%Y-%m-%d")
            normalized.append({"date": str(d)[:10] if d not in (None, "") else None, "volt": v})
        return {"daily_voltage": normalized, "archname": archname, "baty": baty}


@app.get("/dm2000/archives/{archname}/time-at-voltage")
def get_dm2000_time_at_voltage(archname: str, baty: int):
    _validate_dm2000_archname(archname)
    try:
        rows = _read_dm2000_ls(
            "SELECT * FROM ls_timev WHERE archname = ? AND baty = ?",
            (archname, baty),
        )
        if rows:
            return {"time_at_voltage": rows, "archname": archname, "baty": baty}
    except pyodbc.Error:
        pass

    if baty <= 0 or baty > 99:
        raise HTTPException(status_code=400, detail="Invalid baty")
    tim_col = f"tim_vot{baty}"
    try:
        rows = _read_dm2000_ls(
            f"SELECT sj, {tim_col} AS minutes FROM ls_timev WHERE cdid = ? ORDER BY sj DESC",
            (archname,),
        )
        if rows:
            return {"time_at_voltage": rows, "archname": archname, "baty": baty}
    except pyodbc.Error:
        pass

    # Final fallback: ls_vtime stores the same time-at-voltage data for the cdid-based
    # schema using dy (voltage threshold) and time1..time9 columns (values in minutes).
    time_col = f"time{baty}"
    try:
        rows = _read_dm2000_ls(
            f"SELECT dy AS sj, {time_col} AS minutes FROM ls_vtime WHERE cdid = ? ORDER BY dy DESC",
            (archname,),
        )
        return {"time_at_voltage": rows, "archname": archname, "baty": baty}
    except pyodbc.Error:
        return {"time_at_voltage": [], "archname": archname, "baty": baty}


@app.get("/dm2000/config")
def get_dm2000_config():
    """Return station-level configuration (e.g. company name) for use in report previews."""
    return {"company": DM2000_COMPANY_NAME or ""}


@app.get("/dm2000/archives/{archname}/schema")
def get_dm2000_archive_schema(archname: str):
    """Return raw column names and non-null values for an archive row from ls_jb_cs.

    This diagnostic endpoint helps identify the actual column names used in the
    local DM2000 database when expected fields (e.g. voltage_type, endpoint_voltage)
    are missing from report previews.
    """
    _validate_dm2000_archname(archname)
    try:
        rows = _read_dm2000_ls_multi([
            ("SELECT * FROM ls_jb_cs WHERE cdid = ?", (archname,)),
            ("SELECT * FROM ls_jb_cs WHERE archname = ?", (archname,)),
        ])
    except (pyodbc.Error, HTTPException) as exc:
        raise HTTPException(status_code=500, detail=f"Schema query failed: {exc}") from exc
    if not rows:
        raise HTTPException(status_code=404, detail="Archive not found in ls_jb_cs")
    row = rows[0]
    columns = []
    for k, v in row.items():
        columns.append({
            "column": k,
            "value": str(v) if v not in (None, "") else None,
            "is_null_like": v in (None, "", "--"),
        })
    return {"archname": archname, "columns": columns}


@app.post("/dm2000/refresh-archives")
def refresh_dm2000_archives():
    """Force-refresh the DM2000 archives in-memory cache.

    Clears the in-memory archives cache and attempts to re-copy the source
    database file. If the copy fails (e.g. the file is currently locked by
    Microsoft Access), the service falls back to reading directly from the
    source path so updated records are visible immediately.

    Call this endpoint after manually editing dmdata_ls.mdb in Access to see
    changes without waiting for the next auto-refresh cycle.
    """
    # Attempt a forced file-copy refresh (skips the mtime short-circuit).
    _dm2000_refresh_ls_cache(force=True)
    # Always clear the in-memory archives cache regardless of whether the
    # file copy succeeded, so the very next query reads fresh data from DB.
    with _DM2000_ARCHIVES_CACHE_LOCK:
        _DM2000_ARCHIVES_CACHE.clear()
    return {"status": "ok", "cache_path": _DM2000_LS_CACHE_PATH}
def get_dm2000_templates():
    templates_dir = Path(DM2000_TEMPLATES_DIR).resolve()
    if not templates_dir.exists():
        return {"templates": []}
    templates = sorted([
        f.name for f in templates_dir.iterdir()
        if f.is_file() and f.suffix.lower() == ".xlsx"
    ])
    return {"templates": templates}


@app.get("/dm2000/perf-templates")
def get_dm2000_perf_templates():
    templates_dir = Path(DM2000_PERF_TEMPLATES_DIR).resolve()
    if not templates_dir.exists():
        return {"templates": []}
    templates = sorted([
        f.name for f in templates_dir.iterdir()
        if f.is_file() and f.suffix.lower() == ".xlsx"
    ])
    return {"templates": templates}


@app.post("/dm2000/perf-template/upload")
async def upload_dm2000_perf_template(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    # Extract only the basename (no directory components)
    raw_name = Path(file.filename).name
    # Sanitize: replace any characters not in [A-Za-z0-9_-] in the stem with underscores
    stem = raw_name[:-5]
    sanitized_stem = re.sub(r'[^A-Za-z0-9_\-]', '_', stem)
    # Remove consecutive underscores and strip leading/trailing underscores
    sanitized_stem = re.sub(r'_+', '_', sanitized_stem).strip('_')
    if not sanitized_stem:
        sanitized_stem = "template"
    safe_name = sanitized_stem + ".xlsx"
    templates_dir = Path(DM2000_PERF_TEMPLATES_DIR).resolve()
    templates_dir.mkdir(parents=True, exist_ok=True)
    # Re-join using only the validated basename to prevent any path traversal
    dest = templates_dir / safe_name
    if not str(dest).startswith(str(templates_dir)):
        raise HTTPException(status_code=400, detail="Path traversal detected")
    contents = await file.read()
    dest.write_bytes(contents)
    return {"ok": True, "name": safe_name}


@app.post("/dm2000/report")
def generate_dm2000_report(payload: DM2000ReportRequest):
    _validate_dm2000_archname(payload.archname)

    try:
        archive_rows = _read_dm2000_ls("SELECT * FROM ls_jb_cs WHERE cdid = ?", (payload.archname,))
    except pyodbc.Error:
        archive_rows = []
    if not archive_rows:
        try:
            archive_rows = _read_dm2000_ls("SELECT * FROM ls_jb_cs WHERE archname = ?", (payload.archname,))
        except pyodbc.Error:
            archive_rows = []
    if not archive_rows:
        raise HTTPException(status_code=404, detail="Archive not found")
    archive = archive_rows[0]

    if payload.baty == 0:
        curve_data = _read_dm2000_average_curve_rows(payload.archname)
        baty_label = "Average"
    else:
        curve_data = _read_dm2000_curve_rows(payload.archname, payload.baty)
        baty_label = str(payload.baty)

    def _apply_override(db_val, override_val):
        return override_val if override_val is not None and str(override_val).strip() != "" else db_val

    stats = compute_dm2000_stats(curve_data)
    # Override VOLT_MAX/VOLT_MIN with the true OCV/FCV stored in ls_pam2.
    # ls_pam2 holds the instrument-measured open-circuit voltage (OCV, before
    # discharge) and loaded voltage (FCV, at discharge start) — the values
    # DM2000 reports in its own Excel export.  The ls_vtime curve data uses
    # fixed voltage thresholds, not real measurements.
    if payload.baty > 0:
        pam2_stats = _get_pam2_ocv_fcv(payload.archname, payload.baty)
        if pam2_stats:
            stats.update(pam2_stats)
            # Expose as dedicated keys for templates using {{OCV}} / {{FCV}}
            stats["OCV"] = pam2_stats["VOLT_MAX"]
            stats["FCV"] = pam2_stats["VOLT_MIN"]
    context = {
        "ARCHNAME": _apply_override(_dm2000_get_value(archive, "archname", "cdid", "id"), payload.override_archname),
        "START_DATE": _apply_override(str(_dm2000_get_value(archive, "startdate", "fdrq", "fdkssj", "qyrq", "fdrq") or ""), payload.override_start_date),
        "BATTERY_TYPE": _apply_override(_dm2000_get_value(archive, "dcxh"), payload.override_battery_type),
        "BATCH_NAME": _apply_override(_dm2000_get_value(archive, "name", "dcmc"), payload.override_batch_name),
        "DISCHARGE_CONDITION": _apply_override(_dm2000_get_value(archive, "fdfs"), payload.override_discharge_condition),
        "DURATION": _dm2000_get_value(archive, "duration", "fdts"),
        "UNIFORMITY_RATE": _dm2000_get_value(archive, "unifrate", "hl", "hlfd", "yfws_pct", "yfws"),
        "MANUFACTURER": _apply_override(_dm2000_get_value(archive, "manufacturer", "scdw"), payload.override_manufacturer),
        "MADE_DATE": _apply_override(str(_dm2000_get_value(archive, "madedate", "scrq") or ""), payload.override_made_date),
        "SERIAL_NO": _apply_override(_dm2000_get_value(archive, "serialno", "dcph", "ph", "scph", "pch", "lot", "lot_no", "batchno", "batch_no"), payload.override_serial_no),
        "REMARKS": _apply_override(_dm2000_get_value(archive, "remarks", "remark", "bz", "note", "memo", "bzh"), payload.override_remarks),
        "BATTERY_NO": baty_label,
        # Extra fields for full template support
        "COMPANY": DM2000_COMPANY_NAME or "",
        "END_DATE": str(_dm2000_get_value(archive, "enddate", "fzdq", "jssj", "endrq", "endate", "end_date", "fzrq", "stopdate", "fdend") or ""),
        "VOLTAGE_TYPE": str(_dm2000_get_value(
            archive,
            "dylx", "voltage_type", "bcdv", "dcdy", "dxy", "edy",
            "nominal_voltage", "dianxin_leixing", "dianxin", "nominal_v",
            "lxdy", "vtype", "battv", "lx", "dctype", "v_type", "jstj",
        ) or ""),
        "TRADEMARK": str(_dm2000_get_value(archive, "trademark", "shangbiao", "sbmc", "pinpai") or ""),
        "LOAD_RESISTANCE": _append_unit(str(_dm2000_get_value(archive, "load_resistance", "fzdz", "fzlkdz", "dw") or ""), "ohm"),
        "ENDPOINT_VOLTAGE": _append_unit(str(_dm2000_get_value(
            archive,
            "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
            "endpoint_v", "vcut", "cutoffv", "cutoff_v",
            "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
        ) or ""), "V"),
        "DIS_CONDITION": str(_dm2000_get_value(
            archive,
            "dis_condition", "wd", "fdwd", "hjwd", "wendu",
            "fdtj", "hjtj", "temperature", "temp_c",
            "temp", "hjt", "qw", "t", "csh", "jchj",
        ) or ""),
        "MIN_DURATION": str(_dm2000_get_value(archive, "min_duration", "zdts", "min_ts", "minduration", "zdsc", "zxfdts") or ""),
        **stats,
        "HISTORY_DATA": curve_data,
    }

    template_path = _resolve_dm2000_template_path(payload.template_name)
    report_bytes = render_excel_template(template_path, context)
    filename = f"dm2000_report_{payload.archname}_{baty_label}.xlsx"
    return StreamingResponse(
        BytesIO(report_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _compute_uniform_rate_from_tav(
    endpoint_voltage_str: str,
    time_at_volt_map: dict,
    batys: list,
) -> Optional[float]:
    """Compute Uniform Rate = (1 - (Max - Min) / Avg) * 100 at endpoint voltage.

    Uses the time-at-voltage data (ls_vtime/ls_timev) at the endpoint voltage
    threshold for all active batteries.  Returns a percentage rounded to 2 decimal
    places, or None when insufficient data is available.
    """
    try:
        ep = float(str(endpoint_voltage_str or "").strip().split()[0])
    except (IndexError, TypeError, ValueError):
        return None

    times: list[float] = []
    for b in batys:
        rows = time_at_volt_map.get(b) or []
        for row in rows:
            sj = row.get("sj") or row.get("SJ")
            try:
                if sj is not None and abs(float(sj) - ep) < 0.001:
                    mins = row.get("minutes") or row.get("MINUTES")
                    if mins is not None:
                        f = float(mins)
                        if not math.isnan(f) and f >= 0:
                            times.append(f)
                    break
            except (TypeError, ValueError):
                pass

    if len(times) < 2:
        return None

    max_t = max(times)
    min_t = min(times)
    avg_t = sum(times) / len(times)
    if avg_t <= 0:
        return None

    return round((1.0 - (max_t - min_t) / avg_t) * 100.0, 2)


def _append_unit(val: str, unit: str) -> str:
    """Append a physical unit to a value string if not already present."""
    if not val or val == "-":
        return val or ""
    val_stripped = val.strip()
    if val_stripped.lower().endswith(unit.lower()):
        return val_stripped
    return f"{val_stripped} {unit}"


def _build_preview_workbook(  # noqa: C901
    archive_fields: dict,
    company: str,
    batys: list,
    stats_map: dict,
    time_at_volt_map: dict,
    battery_params: dict,
    endpoint_cutoff: Optional[float] = None,
    *,
    report_kind: str = "dm2000",
    telemetry_by_baty: Optional[dict] = None,
    count_at_volt_map: Optional[dict] = None,
) -> bytes:
    """Build a simple Excel workbook matching the ReportPreview HTML format.

    ``report_kind`` selects the layout. ``"dm2000"`` (default) keeps the legacy
    DM2000 layout (OCV/FCV/SOt mAh rows, "Unit: hour" durations and an average
    discharge curve). ``"dmp"`` switches to the DMP variant matching the
    DMP-1 sample report: only OCV/CCV rows, "Unit: times" with integer
    sample counts, DMP-specific archive labels and a per-channel chart built
    from ``telemetry_by_baty``.
    """
    is_dmp = report_kind == "dmp"
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter as _get_col_letter

    # Derive voltage thresholds dynamically from actual time-at-voltage data,
    # sorted descending, so all available rows are shown regardless of range.
    sj_set: set[float] = set()
    for tav_rows in time_at_volt_map.values():
        for row in (tav_rows or []):
            sj_raw = row.get("sj") or row.get("SJ")
            try:
                if sj_raw is not None:
                    sj_set.add(round(float(sj_raw), 4))
            except (TypeError, ValueError):
                pass
    if is_dmp and count_at_volt_map:
        for cnt_rows in count_at_volt_map.values():
            for row in (cnt_rows or []):
                sj_raw = row.get("sj") or row.get("SJ")
                try:
                    if sj_raw is not None:
                        sj_set.add(round(float(sj_raw), 4))
                except (TypeError, ValueError):
                    pass
    THRESHOLDS = sorted(sj_set, reverse=True)  # descending
    # Apply optional cutoff: only show thresholds >= endpoint_cutoff
    if endpoint_cutoff is not None:
        THRESHOLDS = [t for t in THRESHOLDS if t >= endpoint_cutoff - 0.0001]
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill("solid", fgColor="FAFAFA")
    fill_label = PatternFill("solid", fgColor="F5F5F5")
    fill_section = PatternFill("solid", fgColor="EFF4FF")

    wb = _Workbook()
    ws = wb.active
    ws.title = "Battery Discharge Curve"

    num_batys = len(batys)
    total_cols = 1 + num_batys + 3  # label + No.1..N + Max + Min + Avge

    def _safe_float(v):
        if v is None or v in ("", "--"):
            return None
        try:
            f = float(v)
            return None if math.isnan(f) else f
        except (TypeError, ValueError):
            return None

    def _fmt_num(v, decimals=3):
        f = _safe_float(v)
        return round(f, decimals) if f is not None else "-"

    def _agg(vals):
        vs = [_safe_float(v) for v in vals]
        vs = [x for x in vs if x is not None]
        if not vs:
            return "-", "-", "-"
        return round(max(vs), 3), round(min(vs), 3), round(sum(vs) / len(vs), 3)

    def _set(row, col, value, bold=False, fill=None, align="center", merge_to=None, italic=False, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.border = bdr
        c.font = Font(bold=bold, italic=italic, name="Arial", size=size)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if fill:
            c.fill = fill
        if merge_to and merge_to > col:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        return c

    half = total_cols // 2
    r = 1

    # Title
    _set(r, 1, "Battery Discharge Curve", bold=True, merge_to=total_cols, size=14)
    r += 1

    # Company
    if company:
        _set(r, 1, company, merge_to=total_cols)
        r += 1

    # Archive info pairs
    # Compute Uniform Rate from time-at-voltage data when the stored value is not
    # already a proper percentage (the raw DB column yfws holds an integer 0-9).
    _unifrate_raw = archive_fields.get("unifrate") or ""
    try:
        _unifrate_val = float(str(_unifrate_raw).replace("%", "").strip())
        # yfws is stored as a whole integer in [0, 9]; any other value is a percentage
        _unifrate_is_pct = not (_unifrate_val == int(_unifrate_val) and 0 <= _unifrate_val <= 9)
    except (TypeError, ValueError):
        _unifrate_val = None
        _unifrate_is_pct = False

    if not _unifrate_is_pct:
        _computed = _compute_uniform_rate_from_tav(
            archive_fields.get("endpoint_voltage") or "",
            time_at_volt_map,
            batys,
        )
        if _computed is not None:
            _unifrate_display = f"{_computed:.2f} %"
        elif _unifrate_raw:
            _unifrate_display = str(_unifrate_raw)
        else:
            _unifrate_display = "-"
    else:
        _unifrate_display = f"{_unifrate_val:.2f} %" if _unifrate_val is not None else str(_unifrate_raw)

    _voltage_type_val = archive_fields.get("voltage_type") or "-"
    _load_resistance_val = _append_unit(archive_fields.get("load_resistance") or "", "ohm")
    _endpoint_voltage_val = _append_unit(archive_fields.get("endpoint_voltage") or "", "V")

    if is_dmp:
        info_rows_data = [
            ("Battery Name", archive_fields.get("name") or "-", "Archive Name", archive_fields.get("archname") or "-"),
            ("Battery Type", archive_fields.get("dcxh") or "-", "Voltage Type", _voltage_type_val),
            ("Dis-Pattern", archive_fields.get("fdfs") or "-", "Uniformity", _unifrate_display),
            ("Trademark", archive_fields.get("trademark") or "-", "Manufacturer", archive_fields.get("manufacturer") or "-"),
            ("Start Date", archive_fields.get("startdate") or "-", "Made date", archive_fields.get("madedate") or "-"),
            ("Last Date", archive_fields.get("enddate") or "-", "Minimum Duration", archive_fields.get("min_duration") or "-"),
            ("Dis-Surroundings", archive_fields.get("dis_condition") or "-", "End-point Voltage", _endpoint_voltage_val),
        ]
    else:
        info_rows_data = [
            ("Name", archive_fields.get("name") or "-", "Record Name", archive_fields.get("archname") or "-"),
            ("Type", archive_fields.get("dcxh") or "-", "Discharge Pattern", archive_fields.get("fdfs") or "-"),
            ("Voltage Type", _voltage_type_val, "Load Resistance", _load_resistance_val),
            ("Trademark", archive_fields.get("trademark") or "-", "End-point Voltage", _endpoint_voltage_val),
            ("Serial No", archive_fields.get("serialno") or "-", "Uniform Rate", _unifrate_display),
            ("Manufacturer", archive_fields.get("manufacturer") or "-", "Start Date", archive_fields.get("startdate") or "-"),
            ("Made date", archive_fields.get("madedate") or "-", "Last Date", archive_fields.get("enddate") or "-"),
            ("Minimum Duration", archive_fields.get("min_duration") or "-", "Dis-condition", archive_fields.get("dis_condition") or "-"),
        ]
    for left_lbl, left_val, right_lbl, right_val in info_rows_data:
        _set(r, 1, left_lbl, fill=fill_label, align="left")
        _set(r, 2, left_val, align="left", merge_to=half)
        _set(r, half + 1, right_lbl, fill=fill_label, align="left")
        _set(r, half + 2, right_val, align="left", merge_to=total_cols)
        r += 1

    # Instrument row (full width, italic)
    if is_dmp:
        _set(r, 1, "Testing equipment: Type DMP-1 Power Discharge Analyzer (V7.00)",
             italic=True, align="left", merge_to=total_cols)
    else:
        _set(r, 1, "Measure Instrument: Type DM2000 Automatic Discharge Test System (V6.22)",
             italic=True, align="left", merge_to=total_cols)
    r += 1

    # Battery column headers
    _set(r, 1, "", fill=fill_header)
    for i, b in enumerate(batys):
        _set(r, 2 + i, f"No.{b}", bold=True, fill=fill_header)
    _set(r, 2 + num_batys, "Max", bold=True, fill=fill_header)
    _set(r, 3 + num_batys, "Min", bold=True, fill=fill_header)
    _set(r, 4 + num_batys, "Avge", bold=True, fill=fill_header)
    r += 1

    def _get_batt_field(baty, *keys):
        row = battery_params.get(baty) or {}
        for k in keys:
            for kk in (k, k.upper(), k.lower()):
                v = row.get(kk)
                if v not in (None, "", "--"):
                    return v
        return None

    # OCV row
    ocv_vals = [_safe_float(_get_batt_field(b, "ocv", "OCV")) for b in batys]
    mx, mn, av = _agg(ocv_vals)
    _set(r, 1, "OCV V", fill=fill_label, align="left")
    for i, v in enumerate(ocv_vals):
        _set(r, 2 + i, _fmt_num(v))
    _set(r, 2 + num_batys, mx); _set(r, 3 + num_batys, mn); _set(r, 4 + num_batys, av)
    r += 1

    # FCV/CCV row
    fcv_vals = [_safe_float(_get_batt_field(b, "fcv", "FCV")) for b in batys]
    mx, mn, av = _agg(fcv_vals)
    _set(r, 1, "CCV V" if is_dmp else "FCV V", fill=fill_label, align="left")
    for i, v in enumerate(fcv_vals):
        _set(r, 2 + i, _fmt_num(v))
    _set(r, 2 + num_batys, mx); _set(r, 3 + num_batys, mn); _set(r, 4 + num_batys, av)
    r += 1

    # SOt mAh row (omitted for DMP per sample report)
    if not is_dmp:
        sot_vals = [_safe_float(_get_batt_field(b, "sot_mah", "sot", "SOT", "sh", "rql", "fdrl")) for b in batys]
        mx, mn, av = _agg(sot_vals)
        _set(r, 1, "SOt mAh", fill=fill_label, align="left")
        for i, v in enumerate(sot_vals):
            _set(r, 2 + i, _fmt_num(v))
        _set(r, 2 + num_batys, mx); _set(r, 3 + num_batys, mn); _set(r, 4 + num_batys, av)
        r += 1

    # Duration section header
    _duration_unit_label = "times" if is_dmp else "hour"
    _set(r, 1, f"The Duration of Series Designated Voltage (Unit: {_duration_unit_label})",
         fill=fill_section, merge_to=total_cols, italic=True, align="left")
    r += 1

    def _get_tav(baty, threshold):
        rows = time_at_volt_map.get(baty) or []
        for row in rows:
            sj = row.get("sj") or row.get("SJ")
            try:
                if sj is not None and abs(float(sj) - threshold) < 0.001:
                    mins = row.get("minutes") or row.get("MINUTES")
                    if mins is not None:
                        f = float(mins)
                        return None if math.isnan(f) else f / 60.0
            except (TypeError, ValueError):
                pass
        return None

    def _get_count(baty, threshold):
        rows = (count_at_volt_map or {}).get(baty) or []
        for row in rows:
            sj = row.get("sj") or row.get("SJ")
            try:
                if sj is not None and abs(float(sj) - threshold) < 0.001:
                    cnt = row.get("count") or row.get("COUNT")
                    if cnt is not None:
                        return int(cnt)
            except (TypeError, ValueError):
                pass
        return None

    for threshold in THRESHOLDS:
        if is_dmp:
            cell_vals = [_get_count(b, threshold) for b in batys]
        else:
            cell_vals = [_get_tav(b, threshold) for b in batys]
        # Skip rows where no battery has data for this threshold
        if all(v is None for v in cell_vals):
            continue
        if is_dmp:
            numeric = [v for v in cell_vals if v is not None]
            if numeric:
                mx = max(numeric)
                mn = min(numeric)
                av = int(round(sum(numeric) / len(numeric)))
            else:
                mx, mn, av = "-", "-", "-"
            _set(r, 1, round(threshold, 3), fill=fill_label, align="left")
            for i, v in enumerate(cell_vals):
                _set(r, 2 + i, v if v is not None else "-")
            _set(r, 2 + num_batys, mx); _set(r, 3 + num_batys, mn); _set(r, 4 + num_batys, av)
        else:
            mx, mn, av = _agg(cell_vals)
            _set(r, 1, round(threshold, 3), fill=fill_label, align="left")
            for i, v in enumerate(cell_vals):
                _set(r, 2 + i, round(v, 3) if v is not None else "-")
            _set(r, 2 + num_batys, mx); _set(r, 3 + num_batys, mn); _set(r, 4 + num_batys, av)
        r += 1

    # Remarks
    _set(r, 1, "Remark", fill=fill_label, align="left")
    _set(r, 2, archive_fields.get("remarks") or "-", align="left", merge_to=total_cols)

    # ── Discharge-curve chart ────────────────────────────────────────────────
    chart_img_buf: Optional[BytesIO] = None
    if is_dmp:
        # DMP: plot every channel's raw VOLT vs TIM curve so the embedded chart
        # matches the on-screen "all channels" view.
        series: list[tuple[int, list[float], list[float]]] = []
        for b in batys:
            rows = (telemetry_by_baty or {}).get(b) or []
            pts: list[tuple[float, float]] = []
            for rec in rows:
                t = rec.get("TIM")
                v = rec.get("VOLT") or rec.get("volt") or rec.get("Volt")
                try:
                    tf = float(t)
                    vf = float(v)
                except (TypeError, ValueError):
                    continue
                if math.isnan(tf) or math.isnan(vf):
                    continue
                pts.append((tf, vf))
            if len(pts) >= 2:
                pts.sort(key=lambda p: p[0])
                series.append((b, [p[0] for p in pts], [p[1] for p in pts]))

        if series:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            from openpyxl.drawing.image import Image as _XLImage

            colors = [
                "#1677ff", "#f5222d", "#52c41a", "#faad14", "#722ed1",
                "#13c2c2", "#eb2f96", "#fa8c16", "#a0d911", "#2f54eb",
            ]
            fig, ax = plt.subplots(figsize=(8, 4.5))
            for idx, (b, xs, ys) in enumerate(series):
                ax.plot(
                    xs, ys,
                    color=colors[idx % len(colors)],
                    linewidth=1.2,
                    label=f"No.{b}",
                )
            ax.set_xlabel("Time (h)")
            ax.set_ylabel("Voltage (V)")
            ax.set_title("Battery Discharge Curve")
            ax.grid(True, alpha=0.3)
            ax.legend(loc="best", fontsize=8, ncol=min(len(series), 5))
            fig.tight_layout()

            chart_img_buf = BytesIO()
            fig.savefig(chart_img_buf, format="png", dpi=100)
            plt.close(fig)
            chart_img_buf.seek(0)

            xl_img = _XLImage(chart_img_buf)
            xl_img.width = 640
            xl_img.height = 360
            ws.add_image(xl_img, f"A{r + 2}")
    else:
        # DM2000 (legacy): single average-curve plot using TAV aggregates.
        chart_points: list[tuple[float, float]] = []
        _, _, avg_fcv_raw = _agg(fcv_vals)
        avg_fcv = _safe_float(avg_fcv_raw)
        if avg_fcv is not None:
            chart_points.append((round(avg_fcv, 4), 0.0))
        for threshold in THRESHOLDS:
            tav_vals = [_get_tav(b, threshold) for b in batys]
            if all(v is None for v in tav_vals):
                continue
            _, _, av = _agg(tav_vals)
            if isinstance(av, (int, float)):
                chart_points.append((round(threshold, 3), round(av, 4)))

        if len(chart_points) >= 2:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            from openpyxl.drawing.image import Image as _XLImage

            xs = [p[1] for p in chart_points]
            ys = [p[0] for p in chart_points]

            fig, ax = plt.subplots(figsize=(6, 4))
            ax.plot(xs, ys, color="#1677ff", linewidth=1.5, marker="o", markersize=3)
            ax.set_xlabel("Hour")
            ax.set_ylabel("Voltage (V)")
            ax.set_title("The Duration of Series Designated Voltage")
            ax.grid(True, alpha=0.3)
            fig.tight_layout()

            chart_img_buf = BytesIO()
            fig.savefig(chart_img_buf, format="png", dpi=100)
            plt.close(fig)
            chart_img_buf.seek(0)

            xl_img = _XLImage(chart_img_buf)
            xl_img.width = 480
            xl_img.height = 320
            ws.add_image(xl_img, f"A{r + 2}")

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, total_cols + 1):
        ltr = _get_col_letter(col_idx)
        ws.column_dimensions[ltr].width = 10

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/dm2000/report-simple")
def generate_dm2000_simple_report(payload: DM2000SimpleReportRequest):
    """Generate a preview-style Excel report without requiring a template file."""
    _validate_dm2000_archname(payload.archname)

    if not payload.batys:
        raise HTTPException(status_code=400, detail="batys must not be empty")

    batys = [b for b in payload.batys if isinstance(b, int) and 1 <= b <= 99]
    if not batys:
        raise HTTPException(status_code=400, detail="No valid battery numbers provided")

    # Fetch archive metadata
    try:
        archive_rows = _read_dm2000_ls_multi([
            ("SELECT * FROM ls_jb_cs WHERE cdid = ?", (payload.archname,)),
            ("SELECT * FROM ls_jb_cs WHERE archname = ?", (payload.archname,)),
        ])
    except (pyodbc.Error, HTTPException):
        archive_rows = []
    archive = archive_rows[0] if archive_rows else {}

    def _to_date_text(v):
        if v and hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10] if v not in (None, "") else ""

    def _apply_override(db_val, override_val):
        return override_val if override_val is not None and str(override_val).strip() != "" else db_val

    archive_fields = {
        "archname": str(_dm2000_get_value(archive, "archname", "cdid", "id") or payload.archname),
        "name": str(_dm2000_get_value(archive, "name", "dcmc") or ""),
        "startdate": _to_date_text(_dm2000_get_value(archive, "startdate", "fdrq")),
        "enddate": _to_date_text(_dm2000_get_value(archive, "enddate", "fzdq", "jssj", "endrq", "endate", "end_date", "fzrq", "stopdate", "fdend")),
        "dcxh": str(_apply_override(_dm2000_get_value(archive, "dcxh"), payload.override_battery_type) or ""),
        "fdfs": str(_dm2000_get_value(archive, "fdfs") or ""),
        "unifrate": str(_dm2000_get_value(archive, "unifrate", "hl", "hlfd", "yfws_pct", "yfws") or ""),
        "manufacturer": str(_apply_override(_dm2000_get_value(archive, "manufacturer", "scdw"), payload.override_manufacturer) or ""),
        "madedate": _to_date_text(_dm2000_get_value(archive, "madedate", "scrq")),
        "serialno": str(_dm2000_get_value(archive, "serialno", "dcph", "ph", "scph", "pch", "lot", "lot_no", "batchno", "batch_no") or ""),
        "remarks": str(_dm2000_get_value(archive, "remarks", "remark", "bz", "note", "memo", "bzh") or ""),
        "voltage_type": str(_dm2000_get_value(
            archive,
            "dylx", "voltage_type", "bcdv", "dcdy", "dxy", "edy",
            "nominal_voltage", "dianxin_leixing", "dianxin", "nominal_v",
            "lxdy", "vtype", "battv", "lx", "dctype", "v_type", "jstj",
        ) or ""),
        "trademark": str(_dm2000_get_value(archive, "trademark", "shangbiao", "sbmc", "pinpai") or ""),
        "load_resistance": str(_dm2000_get_value(archive, "load_resistance", "fzdz", "fzlkdz", "dw") or ""),
        "endpoint_voltage": str(_dm2000_get_value(
            archive,
            "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
            "endpoint_v", "vcut", "cutoffv", "cutoff_v",
            "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
        ) or ""),
        "dis_condition": str(_dm2000_get_value(
            archive,
            "dis_condition", "wd", "fdwd", "hjwd", "wendu",
            "fdtj", "hjtj", "temperature", "temp_c",
            "temp", "hjt", "qw", "t", "csh", "jchj",
        ) or ""),
        "min_duration": str(_dm2000_get_value(archive, "min_duration", "zdts", "min_ts", "minduration", "zdsc", "zxfdts") or ""),
    }

    # Fetch per-battery params (OCV/FCV/SOt)
    try:
        batt_rows = _read_dm2000_ls_multi([
            ("SELECT * FROM ls_pam2 WHERE archname = ? ORDER BY baty ASC", (payload.archname,)),
            ("SELECT * FROM ls_pam2 WHERE cdid = ? ORDER BY gpp ASC", (payload.archname,)),
        ])
    except (pyodbc.Error, HTTPException):
        batt_rows = []

    battery_params: dict = {}
    for row in batt_rows:
        b_raw = _dm2000_get_value(row, "baty", "gpp")
        try:
            b = int(float(str(b_raw)))
        except (TypeError, ValueError):
            continue
        if b <= 0:
            continue
        if "ocv" not in row:
            row["ocv"] = _dm2000_get_value(row, "ocv", "OCV")
        if "fcv" not in row:
            row["fcv"] = _dm2000_get_value(row, "fcv", "FCV")
        row["sot_mah"] = _dm2000_get_value(row, "sh", "sot", "SOT", "sot_mah", "sotmah", "rql", "fdrl", "rl", "dcrl", "capacity", "sl", "sc", "fdl", "fdsh", "sh_mah", "fdmah", "actual_cap", "shrc", "fdrc")
        battery_params[b] = row

    # Supplement OCV/FCV from ls_evolt
    try:
        evolt_rows = _read_dm2000_ls(
            "SELECT dy, volt1, volt2, volt3, volt4, volt5, volt6, volt7, volt8, volt9"
            " FROM ls_evolt WHERE cdid = ? AND (dy = 'OCV' OR dy = 'FCV')",
            (payload.archname,),
        )
        evolt_map: dict = {}
        for er in evolt_rows:
            dy = str(er.get("dy") or "").strip().upper()
            if dy not in ("OCV", "FCV"):
                continue
            for i in range(1, 10):
                raw_v = er.get(f"volt{i}")
                if raw_v in (None, "", "--"):
                    continue
                try:
                    fv = float(raw_v)
                    if not math.isnan(fv):
                        evolt_map.setdefault(i, {})[dy] = fv
                except (TypeError, ValueError):
                    pass
        for b, evolt_data in evolt_map.items():
            row = battery_params.setdefault(b, {"baty": b})
            if row.get("ocv") is None:
                row["ocv"] = evolt_data.get("OCV")
            if row.get("fcv") is None:
                row["fcv"] = evolt_data.get("FCV")
    except (pyodbc.Error, HTTPException):
        pass

    stats_map: dict = {}
    time_at_volt_map: dict = {}
    for b in batys:
        try:
            stats_map[b] = compute_dm2000_stats(_read_dm2000_curve_rows(payload.archname, b))
            pam2 = _get_pam2_ocv_fcv(payload.archname, b)
            if pam2:
                stats_map[b].update(pam2)
                stats_map[b]["OCV"] = pam2["VOLT_MAX"]
                stats_map[b]["FCV"] = pam2["VOLT_MIN"]
        except (pyodbc.Error, HTTPException):
            stats_map[b] = {}

        row = battery_params.setdefault(b, {"baty": b})
        if row.get("ocv") is None and stats_map[b].get("OCV") is not None:
            row["ocv"] = stats_map[b]["OCV"]
        if row.get("fcv") is None and stats_map[b].get("FCV") is not None:
            row["fcv"] = stats_map[b]["FCV"]

        tav: list = []
        try:
            tav = _read_dm2000_ls("SELECT * FROM ls_timev WHERE archname = ? AND baty = ?", (payload.archname, b))
        except (pyodbc.Error, HTTPException):
            pass
        if not tav:
            tim_col = f"tim_vot{b}"
            try:
                tav = _read_dm2000_ls(
                    f"SELECT sj, {tim_col} AS minutes FROM ls_timev WHERE cdid = ? ORDER BY sj DESC",
                    (payload.archname,),
                )
            except (pyodbc.Error, HTTPException):
                pass
        if not tav:
            time_col = f"time{b}"
            try:
                tav = _read_dm2000_ls(
                    f"SELECT dy AS sj, {time_col} AS minutes FROM ls_vtime WHERE cdid = ? ORDER BY dy DESC",
                    (payload.archname,),
                )
            except (pyodbc.Error, HTTPException):
                pass
        time_at_volt_map[b] = tav

    # Compute SOt mAh from time-at-voltage data for batteries missing it in ls_pam2
    load_r_str = archive_fields.get("load_resistance", "")
    try:
        load_r = float(load_r_str) if load_r_str else None
    except (TypeError, ValueError):
        load_r = None
    if load_r and load_r > 0:
        for b in batys:
            row = battery_params.setdefault(b, {"baty": b})
            if row.get("sot_mah") is None:
                fcv = row.get("fcv")
                sot = _compute_sot_mah_from_tav(time_at_volt_map.get(b, []), load_r, fcv)
                if sot is not None:
                    row["sot_mah"] = sot

    try:
        workbook_bytes = _build_preview_workbook(
            archive_fields=archive_fields,
            company=DM2000_COMPANY_NAME or "",
            batys=batys,
            stats_map=stats_map,
            time_at_volt_map=time_at_volt_map,
            battery_params=battery_params,
            endpoint_cutoff=payload.endpoint_cutoff,
        )
    except Exception as exc:
        logger.exception("Error building preview workbook for archname=%s: %s", payload.archname, exc)
        raise HTTPException(status_code=500, detail=f"Failed to build report: {exc}") from exc
    filename = f"dm2000_preview_{payload.archname}.xlsx"
    return StreamingResponse(
        BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Performance Monitoring Report ──────────────────────────────────────────

def _get_batys_for_archive(archname: str) -> list[int]:
    """Return sorted list of battery numbers that have data in the archive."""
    rows = _derive_dm2000_batteries_from_vtime(archname)
    if not rows:
        try:
            rows = _read_dm2000_ls_multi([
                ("SELECT * FROM ls_pam2 WHERE archname = ? ORDER BY baty ASC", (archname,)),
                ("SELECT * FROM ls_pam2 WHERE cdid = ? ORDER BY gpp ASC", (archname,)),
            ])
        except (pyodbc.Error, HTTPException):
            rows = []
    result: list[int] = []
    for row in rows:
        raw = _dm2000_get_value(row, "baty", "gpp")
        try:
            b = int(float(str(raw)))
            if b > 0:
                result.append(b)
        except (TypeError, ValueError):
            pass
    return sorted(set(result))


def _get_tav_for_batteries(archname: str, batys: list[int]) -> dict[int, list[dict]]:
    """Return time-at-voltage data for each battery number in batys."""
    result: dict[int, list[dict]] = {}
    for b in batys:
        tav: list[dict] = []
        try:
            tav = _read_dm2000_ls(
                "SELECT * FROM ls_timev WHERE archname = ? AND baty = ?",
                (archname, b),
            )
        except (pyodbc.Error, HTTPException):
            pass
        if not tav:
            tim_col = f"tim_vot{b}"
            try:
                tav = _read_dm2000_ls(
                    f"SELECT sj, {tim_col} AS minutes FROM ls_timev WHERE cdid = ? ORDER BY sj DESC",
                    (archname,),
                )
            except (pyodbc.Error, HTTPException):
                pass
        if not tav:
            time_col = f"time{b}"
            try:
                tav = _read_dm2000_ls(
                    f"SELECT dy AS sj, {time_col} AS minutes FROM ls_vtime WHERE cdid = ? ORDER BY dy DESC",
                    (archname,),
                )
            except (pyodbc.Error, HTTPException):
                pass
        result[b] = tav
    return result


def _compute_perf_values(
    endpoint_voltage_str: str,
    tav_map: dict[int, list[dict]],
    batys: list[int],
) -> dict:
    """Compute average discharge time at endpoint voltage and uniform rate.

    Returns a dict with keys:
      ``avg_hours``    — average time in hours (None when unavailable)
      ``avg_minutes``  — average time in minutes (None when unavailable)
      ``uniform_rate`` — uniform rate % (None when < 2 batteries)

    When *endpoint_voltage_str* is empty or cannot be parsed the function
    falls back to the **minimum (last) voltage threshold** present in the
    time-at-voltage data, which corresponds to the final milestone of
    "The Duration of Series Designated Voltage".
    """
    # Parse endpoint voltage, stripping any unit character (e.g. "0.9V" → 0.9)
    ep: Optional[float] = None
    raw_ep = str(endpoint_voltage_str or "").strip()
    if raw_ep:
        token = raw_ep.split()[0] if raw_ep.split() else raw_ep
        # Remove all trailing non-numeric chars (e.g. "mV", "Volts" → numeric part)
        token = re.sub(r"[^0-9.\-]+$", "", token)
        try:
            ep = float(token)
        except (TypeError, ValueError):
            ep = None

    # Fallback: use the minimum voltage that has at least one battery with valid data.
    # Using plain min(all_sj) would pick the lowest threshold voltage even when all
    # batteries have null data there (e.g. a 0.600V row where no battery reached that
    # voltage), which then causes the tolerance check to fail for batteries whose curve
    # only goes down to 0.900V.  We therefore find the lowest voltage where at least
    # one battery actually has a finite, non-negative discharge time recorded.
    if ep is None:
        _fb_ep: Optional[float] = None
        for _fb_b in batys:
            for _fb_row in (tav_map.get(_fb_b) or []):
                _fb_sj = _fb_row.get("sj") or _fb_row.get("SJ")
                _fb_m = _fb_row.get("minutes") or _fb_row.get("MINUTES")
                try:
                    _fb_sv = float(_fb_sj)
                    _fb_mv = float(_fb_m)
                    if not math.isnan(_fb_mv) and _fb_mv >= 0:
                        if _fb_ep is None or _fb_sv < _fb_ep:
                            _fb_ep = _fb_sv
                except (TypeError, ValueError):
                    pass
        if _fb_ep is not None:
            ep = _fb_ep
        else:
            return {
                "avg_hours": None,
                "avg_minutes": None,
                "avg_count": None,
                "uniform_rate": None,
                "is_dmp": False,
            }

    # For each battery find the TAV row whose voltage is closest to ep.
    # A tolerance of 0.05 V is used to handle minor rounding differences
    # between the stored endpoint voltage (e.g. "0.9") and the TAV row
    # voltage thresholds (which may be stored as 0.90, 0.900, etc.).
    TOLERANCE = 0.05
    minutes_list: list[float] = []
    for b in batys:
        rows = tav_map.get(b) or []
        best_mins: Optional[float] = None
        best_diff: float = float("inf")
        for row in rows:
            sj = row.get("sj") or row.get("SJ")
            try:
                diff = abs(float(sj) - ep)
                if diff < best_diff:
                    mins = row.get("minutes") or row.get("MINUTES")
                    if mins is not None:
                        f = float(mins)
                        if not math.isnan(f) and f >= 0:
                            best_diff = diff
                            best_mins = f
            except (TypeError, ValueError):
                pass
        if best_mins is not None and best_diff <= TOLERANCE:
            minutes_list.append(best_mins)

    if not minutes_list:
        return {
            "avg_hours": None,
            "avg_minutes": None,
            "avg_count": None,
            "uniform_rate": None,
            "is_dmp": False,
        }

    avg_min = sum(minutes_list) / len(minutes_list)
    avg_hours = round(avg_min / 60.0, 3)
    avg_minutes = round(avg_min, 3)

    uniform_rate: Optional[float] = None
    if len(minutes_list) >= 2:
        max_t = max(minutes_list)
        min_t = min(minutes_list)
        if avg_min > 0:
            uniform_rate = round((1.0 - (max_t - min_t) / avg_min) * 100.0, 2)

    return {
        "avg_hours": avg_hours,
        "avg_minutes": avg_minutes,
        "avg_count": None,
        "uniform_rate": uniform_rate,
        "is_dmp": False,
    }


def _build_perf_workbook(groups: dict) -> bytes:  # noqa: C901
    """Build the performance monitoring Excel workbook.

    ``groups`` structure::

        {
            sheet_name: {                           # e.g. "LR6 501"
                (date_str, battery_type): {         # e.g. ("2026-01-06", "UD")
                    fdfs_label: {
                        "avg_hours": float | None,
                        "avg_minutes": float | None,
                        "avg_count": int | None,        # cycle count (DMP only)
                        "uniform_rate": float | None,
                        "is_dmp": bool,                 # True for DMP, False for DM2000
                    },
                    ...
                },
                ...
            },
            ...
        }
    """
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter as _gcl

    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_title = PatternFill("solid", fgColor="1F4E79")   # dark blue
    fill_header1 = PatternFill("solid", fgColor="2E75B6")  # blue
    fill_header2 = PatternFill("solid", fgColor="BDD7EE")  # light blue
    fill_hp = PatternFill("solid", fgColor="FFF2CC")       # light yellow
    fill_ud = PatternFill("solid", fgColor="E2EFDA")       # light green
    fill_udplus = PatternFill("solid", fgColor="FCE4D6")   # light orange

    TYPE_FILLS = {"HP": fill_hp, "UD": fill_ud, "UD+": fill_udplus}

    wb = _Workbook()
    wb.remove(wb.active)

    for sheet_name_key, date_type_map in groups.items():
        # Use the first candidate (before any "|") as the sheet title
        sheet_name = sheet_name_key.split("|")[0].strip()[:_EXCEL_MAX_SHEET_NAME]
        ws = wb.create_sheet(title=sheet_name)

        # Collect all unique fdfs labels in this sheet (sorted for consistency)
        all_fdfs: list[str] = []
        seen_fdfs: set[str] = set()
        for row_data in date_type_map.values():
            for lbl in row_data:
                if lbl not in seen_fdfs:
                    seen_fdfs.add(lbl)
                    all_fdfs.append(lbl)
        all_fdfs.sort()

        # Column layout: col1=Date, col2=Type, then 2 cols per fdfs
        num_fdfs = len(all_fdfs)
        total_cols = 2 + num_fdfs * 2

        def _cell(row, col, value="", bold=False, fill=None, font_color="000000",
                  align="center", merge_to_col=None, wrap=False, size=10, italic=False):
            c = ws.cell(row=row, column=col, value=value)
            c.border = bdr
            c.font = Font(bold=bold, italic=italic, name="Arial", size=size,
                          color=font_color)
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=wrap)
            if fill:
                c.fill = fill
            if merge_to_col and merge_to_col > col:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=merge_to_col)
            return c

        # Row 1: Title
        title_text = f"BẢNG THEO DÕI HIỆU SUẤT PIN - {sheet_name}"
        _cell(1, 1, title_text, bold=True, fill=fill_title, font_color="FFFFFF",
              size=13, merge_to_col=total_cols)
        ws.row_dimensions[1].height = 24

        # Row 2: "Ngày", "Loại", then fdfs headers (each spans 2 cols)
        _cell(2, 1, "Ngày", bold=True, fill=fill_header1, font_color="FFFFFF", size=11)
        _cell(2, 2, "Loại", bold=True, fill=fill_header1, font_color="FFFFFF", size=11)
        for i, lbl in enumerate(all_fdfs):
            col_start = 3 + i * 2
            _cell(2, col_start, lbl, bold=True, fill=fill_header1, font_color="FFFFFF",
                  size=9, wrap=True, merge_to_col=col_start + 1)
        ws.row_dimensions[2].height = 40

        # Row 3: Sub-headers ("Kết quả (h)", "Tỉ lệ (%)" per fdfs)
        _cell(3, 1, "", fill=fill_header2)
        _cell(3, 2, "", fill=fill_header2)
        for i in range(num_fdfs):
            col_r = 3 + i * 2
            col_u = col_r + 1
            _cell(3, col_r, "Kết quả (h)", bold=True, fill=fill_header2, size=9)
            _cell(3, col_u, "Tỉ lệ (%)", bold=True, fill=fill_header2, size=9)
        ws.row_dimensions[3].height = 18

        # Data rows — sort by (date, battery_type)
        sorted_keys = sorted(date_type_map.keys(), key=lambda k: (k[0], k[1]))
        for row_key in sorted_keys:
            row_data = date_type_map[row_key]
            date_str, btype = row_key
            excel_row = ws.max_row + 1

            row_fill = TYPE_FILLS.get(btype.upper()) if btype.upper() in TYPE_FILLS else None
            _cell(excel_row, 1, date_str, fill=row_fill, align="left")
            _cell(excel_row, 2, btype, fill=row_fill)

            for i, lbl in enumerate(all_fdfs):
                col_r = 3 + i * 2
                col_u = col_r + 1
                entry = row_data.get(lbl, {})
                avg_h = entry.get("avg_hours")
                ur = entry.get("uniform_rate")
                _cell(excel_row, col_r, avg_h if avg_h is not None else "", fill=row_fill)
                _cell(excel_row, col_u, ur if ur is not None else "", fill=row_fill)

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 8
        for i in range(num_fdfs):
            col_r = 3 + i * 2
            col_u = col_r + 1
            ws.column_dimensions[_gcl(col_r)].width = 13
            ws.column_dimensions[_gcl(col_u)].width = 10

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/dm2000/perf-report")
def generate_dm2000_perf_report(payload: PerfReportRequest):  # noqa: C901
    """Generate a performance monitoring report (Bảng theo dõi hiệu suất pin).

    For each entry the caller specifies the archive name, battery type label
    (HP/UD/UD+) and optionally a subset of battery channel numbers.  The
    endpoint computes the average discharge hours at the endpoint voltage and
    the uniform rate for the chosen batteries, then generates a multi-sheet
    Excel workbook grouped by sheet_name.

    Sheet name resolution (when not manually specified):
      1. ``dcxh + trademark``   — "[Type] [Manufacturer]" as required by template
      2. ``dcxh + serialno``    — legacy / production-line-number format
      3. ``dcxh + manufacturer``
      4. ``dcxh`` alone
    Multiple candidates are stored "|"-joined so the template engine can try
    each one against the actual workbook sheet names.

    Test-method (fdfs) resolution:
      Uses the ``fdfs`` column from the archive record.  When that is empty,
      falls back to ``load_resistance`` so the column-header matching logic
      (whole-word check on leading token) can still find the right column.
    """
    if not payload.entries:
        raise HTTPException(status_code=400, detail="entries must not be empty")

    # groups[sheet_name][(date_str, battery_type)][fdfs_label] =
    #   {avg_hours, avg_minutes, avg_count, uniform_rate, is_dmp}
    groups: dict[str, dict] = {}

    for entry in payload.entries:
        _validate_dm2000_archname(entry.archname)

        # Fetch archive metadata
        try:
            archive_rows = _read_dm2000_ls(
                "SELECT * FROM ls_jb_cs WHERE cdid = ?", (entry.archname,)
            )
        except pyodbc.Error:
            archive_rows = []
        if not archive_rows:
            try:
                archive_rows = _read_dm2000_ls(
                    "SELECT * FROM ls_jb_cs WHERE archname = ?", (entry.archname,)
                )
            except pyodbc.Error:
                archive_rows = []
        if not archive_rows:
            raise HTTPException(
                status_code=404, detail=f"Archive not found: {entry.archname}"
            )
        archive = archive_rows[0]

        def _to_date_text(v):
            if v and hasattr(v, "strftime"):
                return v.strftime("%Y-%m-%d")
            return str(v)[:10] if v not in (None, "") else ""

        startdate = _to_date_text(
            _dm2000_get_value(archive, "startdate", "fdrq")
        )
        fdfs_raw = str(_dm2000_get_value(archive, "fdfs") or "").strip()
        dcxh = str(_dm2000_get_value(archive, "dcxh") or "").strip()
        serialno = str(_dm2000_get_value(archive, "serialno", "dcph") or "").strip()
        trademark = str(_dm2000_get_value(archive, "trademark", "shangbiao", "sb") or "").strip()
        manufacturer_db = str(_dm2000_get_value(archive, "manufacturer", "changshang", "cs") or "").strip()
        load_resistance = str(_dm2000_get_value(
            archive,
            "load_resistance", "fzdz", "fzlkdz", "dw", "fddl", "fdz", "resistance", "load_r", "r_ohm",
        ) or "").strip()
        endpoint_voltage_raw = _dm2000_get_value(
            archive,
            "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
            "endpoint_v", "vcut", "cutoffv", "cutoff_v",
            "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
        )
        endpoint_voltage_str = str(endpoint_voltage_raw or "").strip()

        # fdfs label used to identify the test-method column in the template.
        # Priority: stored fdfs → load_resistance (leading-token match) → archname
        if fdfs_raw:
            fdfs = fdfs_raw
        elif load_resistance:
            # Use load resistance as a partial fdfs key; _perf_fdfs_matches_header
            # will match it against the column header via whole-word check.
            fdfs = load_resistance
        else:
            fdfs = entry.archname

        # Sheet name candidates in priority order:
        #   1. dcxh + trademark  ("Type Manufacturer" per user requirement)
        #   2. dcxh + serialno   (production-line / legacy format)
        #   3. dcxh + manufacturer_db
        #   4. dcxh alone
        # Multiple candidates are stored "|"-joined so _render_perf_template and
        # _build_perf_workbook can try them in order.
        if entry.sheet_name:
            sheet_name = entry.sheet_name.strip()[:_EXCEL_MAX_SHEET_NAME]
        else:
            seen_cands: set[str] = set()
            raw_cands: list[str] = []
            for suffix in (trademark, serialno, manufacturer_db):
                if suffix:
                    cand = f"{dcxh} {suffix}".strip()[:_EXCEL_MAX_SHEET_NAME] if dcxh else suffix[:_EXCEL_MAX_SHEET_NAME]
                    if cand and cand not in seen_cands:
                        seen_cands.add(cand)
                        raw_cands.append(cand)
            if dcxh and dcxh[:_EXCEL_MAX_SHEET_NAME] not in seen_cands:
                raw_cands.append(dcxh[:_EXCEL_MAX_SHEET_NAME])
            if not raw_cands:
                raw_cands = [entry.archname[:_EXCEL_MAX_SHEET_NAME]]
            sheet_name = "|".join(raw_cands)

        # Resolve battery list
        batys = [b for b in entry.batys if isinstance(b, int) and 1 <= b <= MAX_BATTERY_NUMBER]
        if not batys:
            batys = _get_batys_for_archive(entry.archname)
        if not batys:
            continue

        # Get time-at-voltage data
        tav_map = _get_tav_for_batteries(entry.archname, batys)

        # Compute performance values
        perf = _compute_perf_values(endpoint_voltage_str, tav_map, batys)

        # Insert into groups structure
        sheet_group = groups.setdefault(sheet_name, {})
        row_key = (startdate, entry.battery_type)
        row_data = sheet_group.setdefault(row_key, {})
        row_data[fdfs] = perf

    if not groups:
        raise HTTPException(
            status_code=422, detail="No data could be extracted for any entry"
        )

    if payload.template_name:
        template_path = _resolve_perf_template_path(payload.template_name)
        workbook_bytes = _render_perf_template(template_path, groups)
        filename = payload.template_name
    else:
        workbook_bytes = _build_perf_workbook(groups)
        filename = "perf_report.xlsx"
    return StreamingResponse(
        BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── DMP Performance Report (Bảng theo dõi hiệu suất pin) ────────────────────


def _resolve_dmp_perf_template_path(template_name: str) -> str:
    if not _is_valid_template_name(template_name):
        raise HTTPException(status_code=400, detail="Invalid template name")
    templates_dir = Path(DMP_PERF_TEMPLATES_DIR).resolve()
    allowed = {
        f.name for f in templates_dir.iterdir()
        if f.is_file() and _is_valid_template_name(f.name)
    } if templates_dir.exists() else set()
    if template_name not in allowed:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_name}")
    result = (templates_dir / template_name).resolve()
    if not str(result).startswith(str(templates_dir)):
        raise HTTPException(status_code=400, detail="Path traversal detected")
    return str(result)


# Tray assignment by group count (fixed convention)
_DMP_TRAY_ASSIGNMENT: dict[int, list[list[int]]] = {
    1: [list(range(1, 10))],              # 9 trays
    2: [list(range(1, 5)), list(range(6, 10))],   # 4 + 4, skip tray 5
    3: [list(range(1, 4)), list(range(4, 7)), list(range(7, 10))],  # 3 + 3 + 3
}

# Map special_type → row label for column-A matching in Excel template
_SPECIAL_TYPE_LABEL: dict[str, str] = {
    "6020": "6020",
    "3thang": "3 THÁNG",
    "6thang": "6 THÁNG",
}


def _dm2000_archive_matches_chuyen(arch_meta: dict, chuyen: str) -> bool:
    """Return True if *chuyen* (production-line code) is contained in the
    archive's **remark (bz)**, **archive name**, or **serial number** fields.

    Deliberately does *not* check the manufacturer/cs field (``scdw``): that
    field often covers a range of production lines (e.g. ``"501-502"``) and
    would match archives for a completely different line — making the match
    ambiguous when two archives share the same manufacturer range.

    Handles exact tokens (e.g. ``"501"`` in ``"UDP501"``) and bare numeric
    tokens (e.g. ``"501"``).  Tokens with a non-digit prefix are stripped
    (e.g. ``"HP503"`` → ``"503"``).

    Returns ``False`` when *chuyen* is empty so that groups without a
    production-line filter are never incorrectly matched to all archives.
    """
    chuyen = str(chuyen).strip()
    if not chuyen:
        return False  # cannot match without a production-line code
    for field_val in (
        _dm2000_get_value(arch_meta, "bz", "remark"),           # remark / ghi chú
        _dm2000_get_value(arch_meta, "archname", "cdmc"),        # archive name
        _dm2000_get_value(arch_meta, "serialno", "dcph"),        # serial number
    ):
        if not field_val:
            continue
        for tok in re.split(r"[\-,/\s]+", str(field_val).strip()):
            tok = tok.strip()
            if not tok:
                continue
            if tok == chuyen:
                return True
            # Strip leading non-digit prefix (e.g. "HP503" → "503", "UDP501" → "501")
            stripped = re.sub(r"^[^0-9]+", "", tok)
            if stripped and stripped == chuyen:
                return True
    return False


def _build_dm2000_condition_label(fdfs_raw: str, load_res: str, ep_str: str, fallback: str) -> str:
    """Build a combined DM2000 condition label like '10ohm 24h/d-0.900V'.

    Uses the same format as DMP's para_pub.jstj (e.g. "(1500mW2s,650mW28s)10T/h,24h/d-1.05V")
    so that _perf_fdfs_matches_template can correctly locate the label in the IEC
    60086-2 template and assign the right column order.

      resistance = load_res appended with "ohm" when it is a bare number
      prefix     = resistance + fdfs joined by a space (empty parts omitted)
      suffix     = "-<ep_str>V" when endpoint voltage is present
    """
    resistance = ""
    if load_res:
        resistance = f"{load_res}ohm" if re.match(r"^\d+(\.\d+)?$", load_res) else load_res
    parts = [p for p in (resistance, fdfs_raw) if p]
    prefix = " ".join(parts)
    suffix = f"-{ep_str}V" if ep_str else ""
    return (prefix + suffix) or fallback


def _dmp_compute_group_perf(
    batch_id: str,
    trays: list[int],
    endpoint_voltage: Optional[float],
) -> dict:
    """Compute avg discharge hours and uniformity rate for a set of DMP trays.

    Reads vidata for each tray from the cached DMPDATA.mdb (via para_singl to
    resolve the correct per-batch MDB path), then interpolates the discharge
    duration to *endpoint_voltage* for each tray.

    Returns a dict with keys ``avg_hours``, ``avg_minutes``, ``avg_count``,
    ``uniform_rate`` and ``is_dmp`` (always ``True``).

    The ``avg_count`` field is the average number of discharge cycles (1-based
    index of the first telemetry sample whose voltage drops to/below the
    endpoint).  This matches the on-screen DMP report (Unit: times) and is the
    value that should be written to template columns whose header ends with
    ``(t)`` (number of cycles).  See ``_render_perf_template`` for unit
    selection details.

    For DMP, ``uniform_rate`` is computed from the cycle counts so that the
    figure on the report is consistent with the displayed cycle averages.
    """
    if not trays:
        return {
            "avg_hours": None,
            "avg_minutes": None,
            "avg_count": None,
            "uniform_rate": None,
            "is_dmp": True,
        }

    # Look up cdmc (MDB path) for each tray from para_singl
    try:
        singl_rows = _read_dmpdata(
            "SELECT baty, cdmc FROM para_singl WHERE sid = ?", (batch_id,)
        )
    except pyodbc.Error as exc:
        logger.warning("_dmp_compute_group_perf: para_singl read failed: %s", exc)
        singl_rows = []

    cdmc_by_baty: dict[int, str] = {}
    for row in singl_rows:
        baty = _dm2000_get_value(row, "baty")
        cdmc = _dm2000_get_value(row, "cdmc")
        if baty is not None and cdmc:
            cdmc_by_baty[int(baty)] = str(cdmc).strip()

    # Compute time-at-endpoint and cycle-count-at-endpoint for each tray.
    # ``hours_list`` holds the interpolated discharge time (hours) for the
    # ``(h)``/``(m)`` columns, while ``count_list`` holds the cycle count
    # (1-based index of the first sample whose voltage drops to/below the
    # endpoint) used for DMP ``(t)`` columns ("số lần phóng điện").
    hours_list: list[float] = []
    count_list: list[int] = []
    for tray in trays:
        cdmc = cdmc_by_baty.get(tray)
        if not cdmc:
            logger.debug("_dmp_compute_group_perf: no cdmc for tray %d in batch %s", tray, batch_id)
            continue
        try:
            telemetry = _read_telemetry(cdmc, tray)  # TIM already in hours after this call
        except HTTPException as exc:
            logger.warning("_dmp_compute_group_perf: telemetry read failed tray=%d: %s", tray, exc)
            continue

        if not telemetry:
            continue

        # Determine effective endpoint voltage
        ep = endpoint_voltage
        if ep is None:
            # Fall back to minimum voltage in this curve
            volt_vals = [
                float(r.get("VOLT") or r.get("volt") or 0)
                for r in telemetry
                if r.get("VOLT") not in (None, "", "--")
            ]
            ep = min(volt_vals) if volt_vals else None
        if ep is None:
            continue

        # Interpolate time and capture cycle count when VOLT first crosses ep
        # from above.  Both metrics use the same time-sorted point list so the
        # results stay aligned with the on-screen DMP report (which uses the
        # 1-based index of the crossing sample as its cycle count).
        points = sorted(
            [
                (float(r["TIM"]), float(r.get("VOLT") or r.get("volt") or 0))
                for r in telemetry
                if r.get("TIM") not in (None, "", "--")
                and r.get("VOLT") not in (None, "", "--")
            ],
            key=lambda p: p[0],
        )
        crossed_h: Optional[float] = None
        cycle_count: Optional[int] = None
        for i in range(1, len(points)):
            t1, v1 = points[i - 1]
            t2, v2 = points[i]
            if v1 >= ep and v2 <= ep:
                if v1 == v2:
                    crossed_h = t1
                else:
                    crossed_h = t1 + (t2 - t1) * (v1 - ep) / (v1 - v2)
                # Match the frontend ``countAtVoltage`` convention: the cycle
                # count is the 1-based position of the crossing sample (i + 1
                # because ``i`` starts at 1 in this loop, indexing the second
                # point of the pair).
                cycle_count = i + 1
                break
        if crossed_h is not None:
            hours_list.append(crossed_h)
        if cycle_count is not None:
            count_list.append(cycle_count)

    if not hours_list and not count_list:
        return {
            "avg_hours": None,
            "avg_minutes": None,
            "avg_count": None,
            "uniform_rate": None,
            "is_dmp": True,
        }

    avg_h: Optional[float] = None
    avg_m: Optional[float] = None
    if hours_list:
        avg_h = sum(hours_list) / len(hours_list)
        avg_m = avg_h * 60.0

    avg_count: Optional[int] = None
    if count_list:
        # Round to the nearest integer to match the frontend display
        # (``Math.round`` of the mean cycle count).
        avg_count = int(round(sum(count_list) / len(count_list)))

    # Uniform rate for DMP is computed from the cycle counts so the figure
    # printed in the report stays consistent with the displayed cycle
    # averages ("số lần phóng điện").  Fall back to the time-based formula
    # when cycle counts are unavailable.
    uniform_rate: Optional[float] = None
    if len(count_list) >= 2:
        avg_c = sum(count_list) / len(count_list)
        if avg_c > 0:
            uniform_rate = round(
                (1.0 - (max(count_list) - min(count_list)) / avg_c) * 100.0, 2
            )
    elif len(hours_list) >= 2 and avg_h and avg_h > 0:
        uniform_rate = round(
            (1.0 - (max(hours_list) - min(hours_list)) / avg_h) * 100.0, 2
        )

    return {
        "avg_hours": round(avg_h, 3) if avg_h is not None else None,
        "avg_minutes": round(avg_m, 3) if avg_m is not None else None,
        "avg_count": avg_count,
        "uniform_rate": uniform_rate,
        "is_dmp": True,
    }


@app.get("/dmp-perf-templates")
def get_dmp_perf_templates():
    templates_dir = Path(DMP_PERF_TEMPLATES_DIR).resolve()
    if not templates_dir.exists():
        return {"templates": []}
    templates = sorted([
        f.name for f in templates_dir.iterdir()
        if f.is_file() and f.suffix.lower() == ".xlsx"
    ])
    return {"templates": templates}


@app.post("/dmp-perf-template/upload")
async def upload_dmp_perf_template(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    raw_name = Path(file.filename).name
    stem = raw_name[:-5]
    sanitized_stem = re.sub(r'[^A-Za-z0-9_\-]', '_', stem)
    sanitized_stem = re.sub(r'_+', '_', sanitized_stem).strip('_')
    if not sanitized_stem:
        sanitized_stem = "template"
    safe_name = sanitized_stem + ".xlsx"
    templates_dir = Path(DMP_PERF_TEMPLATES_DIR).resolve()
    templates_dir.mkdir(parents=True, exist_ok=True)
    dest = templates_dir / safe_name
    if not str(dest).startswith(str(templates_dir)):
        raise HTTPException(status_code=400, detail="Path traversal detected")
    contents = await file.read()
    dest.write_bytes(contents)
    return {"ok": True, "name": safe_name}


def _parse_bz_groups(bz: str) -> list[dict]:
    """Parse a DM2000 remark like ``'LR6 UDP501 HP503'`` into group tokens.

    Recognises the same token formats as the frontend ``parseRemark`` function:

    * ``UDP{n}``  →  ``{"loai": "UD+", "chuyen": "{n}"}``
    * ``HP{n}``   →  ``{"loai": "HP",  "chuyen": "{n}"}``
    * ``UD{n}``   →  ``{"loai": "UD",  "chuyen": "{n}"}``

    A leading 6-digit DDMMYY date token is skipped.  Unrecognised tokens (e.g.
    model names like ``'LR6'``) are silently ignored.  Returns ``[]`` when no
    group tokens are found.
    """
    groups: list[dict] = []
    tokens = (bz or "").strip().upper().split()
    start = 1 if (tokens and re.fullmatch(r"\d{6}", tokens[0])) else 0
    for tok in tokens[start:]:
        if re.fullmatch(r"UDP\d+", tok):
            groups.append({"loai": "UD+", "chuyen": tok[3:]})
        elif re.fullmatch(r"HP\d+", tok):
            groups.append({"loai": "HP", "chuyen": tok[2:]})
        elif re.fullmatch(r"UD\d+", tok):
            groups.append({"loai": "UD", "chuyen": tok[2:]})
    return groups


def _dm2000_all_archives_match_all_groups(arch_rows: list[dict], groups: list) -> bool:
    """Return True when every archive matches every group's chuyen AND all archives
    have distinct start dates.

    This signals the "same-bz, different-dates" scenario: multiple archives
    share an identical remark (e.g. "LR6 UDP501 HP503") and therefore each
    archive covers ALL production lines.  In this case each archive must be
    processed with all groups independently (using *_parse_bz_groups* +
    *_DMP_TRAY_ASSIGNMENT* battery splits) rather than a 1-archive-per-group
    positional pairing.

    When two or more archives share the same start date they represent
    same-day per-grade runs (e.g. a UD+ archive on channels 1-4 and an HP
    archive on channels 6-9, both with remark "LR6 UDP501 HP503").  In that
    case the positional pairing path must be used so each archive is matched
    to its own group and queried with its own actual batteries — not with
    the fixed [1-4]/[6-9] splits that *_DMP_TRAY_ASSIGNMENT* prescribes
    (which fail when DM2000 renumbers each archive's channels from 1).
    """
    if not arch_rows or not groups:
        return False
    if not all(
        all(_dm2000_archive_matches_chuyen(a, g.chuyen) for g in groups)
        for a in arch_rows
    ):
        return False
    # Require unique start dates.  When any two archives share the same date
    # this is a same-day per-grade scenario, not different-dates same-bz.
    dates: list[str] = []
    for a in arch_rows:
        raw_d = _dm2000_get_value(a, "startdate", "fdrq")
        if raw_d is None:
            d_str = ""
        elif hasattr(raw_d, "strftime"):
            d_str = raw_d.strftime("%Y-%m-%d")
        else:
            d_str = str(raw_d).strip()[:10]
        dates.append(d_str)
    non_empty = [d for d in dates if d]
    if non_empty and len(non_empty) != len(set(non_empty)):
        return False
    return True


def _pair_dm2000_archives_to_groups(
    arch_rows: list[dict],
    groups: list,
) -> list[tuple[int, int]]:
    """Return ``[(arch_idx, grp_idx), ...]`` pairs for multi-archive DM2000 data.

    **Strategy 1 — bijective chuyen matching** (preferred):
    Build a 1-to-1 mapping where each archive uniquely matches exactly one group
    via :func:`_dm2000_archive_matches_chuyen` and each group is matched by
    exactly one archive.  Used when manufacturers differ between archives (e.g.
    ``"501"`` vs ``"503"``).

    **Strategy 2 — positional fallback** (ambiguous case):
    When the chuyen-based mapping is ambiguous — e.g. both archives share the
    same manufacturer field (``"501-502"``), so both match the first group and
    neither matches the second — sort archives by their minimum battery number
    (ascending) and pair them with groups in declaration order.  This correctly
    assigns the archive whose channels start at 1 (batteries 1-4, UD+) to
    group 0 and the archive whose channels start at 6 (batteries 6-9, HP) to
    group 1.
    """
    n_archs = len(arch_rows)
    n_grps = len(groups)

    # Build per-archive match lists via chuyen
    arch_to_grp: dict[int, int] = {}   # ai → gi  (only when the match is unique)
    grp_to_archs: dict[int, list[int]] = {}
    for ai, a in enumerate(arch_rows):
        matches = [
            gi for gi, g in enumerate(groups)
            if _dm2000_archive_matches_chuyen(a, g.chuyen)
        ]
        if len(matches) == 1:
            arch_to_grp[ai] = matches[0]
            grp_to_archs.setdefault(matches[0], []).append(ai)

    # Bijective: every archive has exactly one match AND every matched group is
    # claimed by exactly one archive AND all groups are covered.
    is_bijective = (
        len(arch_to_grp) == n_archs
        and all(len(v) == 1 for v in grp_to_archs.values())
        and len(grp_to_archs) == min(n_archs, n_grps)
    )

    if is_bijective:
        return list(arch_to_grp.items())

    # Positional fallback — sort archives by min battery number then pair.
    # Archives with no battery data sort last (using MAX_BATTERY_NUMBER + 1).
    _NO_BATTERY_SORT_VALUE = MAX_BATTERY_NUMBER + 1

    def _min_battery_key(ai: int) -> tuple:
        cdid = str(_dm2000_get_value(arch_rows[ai], "cdid", "archname") or "").strip()
        batys = _get_batys_for_archive(cdid) if cdid else []
        return (min(batys) if batys else _NO_BATTERY_SORT_VALUE, cdid, ai)

    sorted_ais = sorted(range(n_archs), key=_min_battery_key)
    return [(sorted_ais[i], i) for i in range(min(len(sorted_ais), n_grps))]


def _compute_dmp_perf_groups(  # noqa: C901
    payload: DmpPerfReportRequest,
    skip_not_found: bool = False,
) -> "dict[str, dict]":
    """Compute performance groups dict from DMP entries.

    Returns ``groups[sheet_key][(row_label, loai)][fdfs_label] =
    {avg_hours, avg_minutes, avg_count, uniform_rate, is_dmp}``.

    This shared helper is used by both the Excel report generator and the JSON
    preview endpoint so the computation logic lives in exactly one place.
    """

    def _to_date(v) -> str:
        if v and hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        s = str(v or "").strip()[:10].replace("/", "-")
        return s if len(s) == 10 and s[4] == "-" and s[7] == "-" else ""

    groups: dict[str, dict] = {}

    for entry in payload.entries:
        if not (entry.batch_id or "").strip() and not entry.dm2000_archname:
            continue

        # ── DM2000 path: when dm2000_archname is set, read from DM2000 instead of DMP ──
        if entry.dm2000_archname and entry.dm2000_archname.strip():
            _dm2k_arch = entry.dm2000_archname.strip()
            try:
                _validate_dm2000_archname(_dm2k_arch)
            except HTTPException:
                logger.warning("_compute_dmp_perf_groups: invalid dm2000_archname: %s", _dm2k_arch)
                continue

            # Read archive metadata from ls_jb_cs.
            # Primary lookup: by bz/remarks column (the remark text the user enters).
            # Fallback lookups: by cdid then by archname for backward compatibility.
            _arch_rows: list[dict] = []
            try:
                _arch_rows = _read_dm2000_ls(
                    "SELECT * FROM ls_jb_cs WHERE bz = ?", (_dm2k_arch,)
                )
            except pyodbc.Error:
                pass
            if not _arch_rows:
                # Escape Access SQL LIKE wildcards so a remark containing % or _ is treated literally
                _bz_escaped = _dm2k_arch.replace("%", "[%]").replace("_", "[_]")
                _bz_like = f"%{_bz_escaped}%"
                try:
                    _arch_rows = _read_dm2000_ls(
                        "SELECT * FROM ls_jb_cs WHERE bz LIKE ?", (_bz_like,)
                    )
                except pyodbc.Error:
                    pass
            if not _arch_rows:
                try:
                    _arch_rows = _read_dm2000_ls(
                        "SELECT * FROM ls_jb_cs WHERE cdid = ?", (_dm2k_arch,)
                    )
                except pyodbc.Error:
                    pass
            if not _arch_rows:
                try:
                    _arch_rows = _read_dm2000_ls(
                        "SELECT * FROM ls_jb_cs WHERE archname = ?", (_dm2k_arch,)
                    )
                except pyodbc.Error:
                    pass
            if not _arch_rows:
                if skip_not_found:
                    continue
                _arch_meta: dict = {}
                _dm2k_resolved = _dm2k_arch
            else:
                # When multiple archives match and at least one can be paired with a
                # group by production-line (chuyen), process every archive separately
                # for its matching group so that:
                #   • each group reads data from its own archive (correct grade/line),
                #   • all discharge conditions across all archives are collected.
                # _pair_dm2000_archives_to_groups handles the case where archives share
                # the same manufacturer field (e.g. both "501-502") and chuyen matching
                # is therefore ambiguous — it falls back to positional sort by min battery.
                _dm2k_multi = len(_arch_rows) > 1 and len(entry.groups) > 0 and any(
                    _dm2000_archive_matches_chuyen(a, g.chuyen)
                    for a in _arch_rows for g in entry.groups
                )
                if _dm2k_multi:
                    _dm2k_model_upper_m = entry.model.strip().upper()
                    _dm2k_no_chuyen_m = {"LR61", "9V", "6LR61"}
                    # When every archive matches every group the archives share the same
                    # remark ("LR6 UDP501 HP503") and represent **different test dates**
                    # — each archive covers ALL production lines.  Process each archive
                    # with all groups via _parse_bz_groups + _DMP_TRAY_ASSIGNMENT (the
                    # same logic the single-archive path already applies correctly).
                    # Using 1-to-1 positional pairing in this case is wrong: it assigns
                    # different-date archives to different groups, producing incorrect
                    # averages (all-battery mix) and silently dropping some groups.
                    _dm2k_all_match_all = _dm2000_all_archives_match_all_groups(
                        _arch_rows, entry.groups
                    )
                    if _dm2k_all_match_all:
                        for _dm2k_a in _arch_rows:
                            _dm2k_a_resolved = (
                                str(_dm2000_get_value(_dm2k_a, "cdid", "archname") or _dm2k_arch).strip()
                                or _dm2k_arch
                            )
                            _dm2k_a_fdfs_raw = str(_dm2000_get_value(_dm2k_a, "fdfs") or "").strip()
                            _dm2k_a_load_res = str(_dm2000_get_value(
                                _dm2k_a, "load_resistance", "fzdz", "fzlkdz", "dw", "fddl", "fdz", "resistance", "load_r", "r_ohm",
                            ) or "").strip()
                            _dm2k_a_ep_raw = _dm2000_get_value(
                                _dm2k_a,
                                "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
                                "endpoint_v", "vcut", "cutoffv", "cutoff_v",
                                "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
                            )
                            _dm2k_a_ep_str = str(_dm2k_a_ep_raw or "").strip()
                            _dm2k_a_fdfs = _build_dm2000_condition_label(
                                _dm2k_a_fdfs_raw, _dm2k_a_load_res, _dm2k_a_ep_str, _dm2k_arch,
                            )
                            _dm2k_a_startdate = _dm2000_get_value(_dm2k_a, "startdate", "fdrq")
                            _dm2k_a_fdrq = _to_date(_dm2k_a_startdate) if _dm2k_a_startdate else ""
                            if entry.special_type in _SPECIAL_TYPE_LABEL:
                                _dm2k_a_row_label = _SPECIAL_TYPE_LABEL[entry.special_type]
                            elif _dm2k_a_fdrq:
                                _dm2k_a_row_label = _dm2k_a_fdrq
                            elif entry.report_date:
                                _dm2k_a_row_label = _to_date(entry.report_date) or entry.report_date
                            else:
                                _dm2k_a_row_label = _dm2k_arch
                            _dm2k_a_all_batys = _get_batys_for_archive(_dm2k_a_resolved)
                            _dm2k_a_bz_raw = str(_dm2000_get_value(
                                _dm2k_a, "remarks", "remark", "bz", "note", "memo", "bzh",
                            ) or "").strip()
                            _dm2k_a_bz_groups = _parse_bz_groups(_dm2k_a_bz_raw)
                            _dm2k_a_eff_groups = [
                                {"loai": grp.loai, "chuyen": grp.chuyen, "trays": list(grp.trays or [])}
                                for grp in entry.groups
                            ]
                            _dm2k_a_n = len(_dm2k_a_eff_groups)
                            if len(_dm2k_a_bz_groups) > len(_dm2k_a_eff_groups) and not any(
                                g["trays"] for g in _dm2k_a_eff_groups
                            ):
                                _dm2k_a_eff_groups = [
                                    {"loai": g["loai"], "chuyen": g["chuyen"], "trays": []}
                                    for g in _dm2k_a_bz_groups
                                ]
                                _dm2k_a_n = len(_dm2k_a_eff_groups)
                            _dm2k_a_auto_trays = _DMP_TRAY_ASSIGNMENT.get(
                                _dm2k_a_n, [_dm2k_a_all_batys]
                            )
                            for _dm2k_a_g_idx, _dm2k_a_eff_grp in enumerate(_dm2k_a_eff_groups):
                                _dm2k_a_grp_trays = _dm2k_a_eff_grp.get("trays") or []
                                if _dm2k_a_grp_trays:
                                    _dm2k_a_batys = [
                                        b for b in _dm2k_a_grp_trays
                                        if isinstance(b, int) and 1 <= b <= MAX_BATTERY_NUMBER
                                    ]
                                else:
                                    _dm2k_a_batys = (
                                        _dm2k_a_auto_trays[_dm2k_a_g_idx]
                                        if _dm2k_a_g_idx < len(_dm2k_a_auto_trays)
                                        else _dm2k_a_all_batys
                                    )
                                if not _dm2k_a_batys:
                                    continue
                                _dm2k_a_tav = _get_tav_for_batteries(_dm2k_a_resolved, _dm2k_a_batys)
                                _dm2k_a_perf = _compute_perf_values(_dm2k_a_ep_str, _dm2k_a_tav, _dm2k_a_batys)
                                _dm2k_a_perf["hfsj_unit"] = "hour"
                                _dm2k_a_grp_chuyen = _dm2k_a_eff_grp.get("chuyen") or ""
                                _dm2k_a_grp_loai = _dm2k_a_eff_grp.get("loai") or ""
                                _dm2k_a_sheet_key = (
                                    entry.model.strip()
                                    if _dm2k_model_upper_m in _dm2k_no_chuyen_m
                                    else f"{entry.model.strip()} {_dm2k_a_grp_chuyen.strip()}"
                                )
                                _dm2k_a_fdfs_label = _dm2k_a_fdfs or _dm2k_a_grp_loai
                                groups.setdefault(_dm2k_a_sheet_key, {}).setdefault(
                                    (_dm2k_a_row_label, _dm2k_a_grp_loai), {}
                                )[_dm2k_a_fdfs_label] = _dm2k_a_perf
                        continue  # all archives × all groups processed; skip DMP path
                    _dm2k_ai_gi_pairs = _pair_dm2000_archives_to_groups(_arch_rows, entry.groups)
                    _dm2k_ai_to_gi = dict(_dm2k_ai_gi_pairs)
                    for _dm2k_ai, _dm2k_a in enumerate(_arch_rows):
                        _dm2k_gi = _dm2k_ai_to_gi.get(_dm2k_ai)
                        if _dm2k_gi is None:
                            continue
                        _dm2k_grp = entry.groups[_dm2k_gi]
                        _dm2k_a_resolved = (
                            str(_dm2000_get_value(_dm2k_a, "cdid", "archname") or _dm2k_arch).strip()
                            or _dm2k_arch
                        )
                        _dm2k_a_fdfs_raw = str(_dm2000_get_value(_dm2k_a, "fdfs") or "").strip()
                        _dm2k_a_load_res = str(_dm2000_get_value(
                            _dm2k_a, "load_resistance", "fzdz", "fzlkdz", "dw", "fddl", "fdz", "resistance", "load_r", "r_ohm",
                        ) or "").strip()
                        _dm2k_a_ep_raw = _dm2000_get_value(
                            _dm2k_a,
                            "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
                            "endpoint_v", "vcut", "cutoffv", "cutoff_v",
                            "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
                        )
                        _dm2k_a_ep_str = str(_dm2k_a_ep_raw or "").strip()
                        _dm2k_a_fdfs = _build_dm2000_condition_label(
                            _dm2k_a_fdfs_raw, _dm2k_a_load_res, _dm2k_a_ep_str, _dm2k_arch,
                        )
                        _dm2k_a_startdate = _dm2000_get_value(_dm2k_a, "startdate", "fdrq")
                        _dm2k_a_fdrq = _to_date(_dm2k_a_startdate) if _dm2k_a_startdate else ""
                        if entry.special_type in _SPECIAL_TYPE_LABEL:
                            _dm2k_a_row_label = _SPECIAL_TYPE_LABEL[entry.special_type]
                        elif _dm2k_a_fdrq:
                            _dm2k_a_row_label = _dm2k_a_fdrq
                        elif entry.report_date:
                            _dm2k_a_row_label = _to_date(entry.report_date) or entry.report_date
                        else:
                            _dm2k_a_row_label = _dm2k_arch
                        _dm2k_a_all_batys = _get_batys_for_archive(_dm2k_a_resolved)
                        _dm2k_a_batys = (
                            [b for b in _dm2k_grp.trays if isinstance(b, int) and 1 <= b <= MAX_BATTERY_NUMBER]
                            if _dm2k_grp.trays else _dm2k_a_all_batys
                        )
                        if not _dm2k_a_batys:
                            continue
                        _dm2k_a_tav = _get_tav_for_batteries(_dm2k_a_resolved, _dm2k_a_batys)
                        _dm2k_a_perf = _compute_perf_values(_dm2k_a_ep_str, _dm2k_a_tav, _dm2k_a_batys)
                        _dm2k_a_perf["hfsj_unit"] = "hour"
                        _dm2k_a_sheet_key = (
                            entry.model.strip()
                            if _dm2k_model_upper_m in _dm2k_no_chuyen_m
                            else f"{entry.model.strip()} {_dm2k_grp.chuyen.strip()}"
                        )
                        _dm2k_a_fdfs_label = _dm2k_a_fdfs or _dm2k_grp.loai
                        groups.setdefault(_dm2k_a_sheet_key, {}).setdefault(
                            (_dm2k_a_row_label, _dm2k_grp.loai), {}
                        )[_dm2k_a_fdfs_label] = _dm2k_a_perf
                    continue  # all archives processed; skip DMP path

                _arch_meta = _arch_rows[0]
                # Resolve the actual cdid so battery/TAV queries use the correct key
                _dm2k_resolved = (
                    str(_dm2000_get_value(_arch_meta, "cdid", "archname") or _dm2k_arch).strip()
                    or _dm2k_arch
                )

            _dm2k_fdfs_raw = str(_dm2000_get_value(_arch_meta, "fdfs") or "").strip()
            _dm2k_load_res = str(_dm2000_get_value(
                _arch_meta, "load_resistance", "fzdz", "fzlkdz", "dw", "fddl", "fdz", "resistance", "load_r", "r_ohm",
            ) or "").strip()
            _dm2k_ep_raw = _dm2000_get_value(
                _arch_meta,
                "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
                "endpoint_v", "vcut", "cutoffv", "cutoff_v",
                "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
            )
            _dm2k_ep_str = str(_dm2k_ep_raw or "").strip()
            _dm2k_fdfs = _build_dm2000_condition_label(
                _dm2k_fdfs_raw, _dm2k_load_res, _dm2k_ep_str, _dm2k_arch,
            )

            # Determine row label — prefer the archive's own start date (fdrq) over
            # entry.report_date, which is typically set to today when no date prefix
            # was entered in the remark.  The DM2000 archive's fdrq is the actual
            # test start date and should be used as the row label.
            _dm2k_startdate_raw = _dm2000_get_value(_arch_meta, "startdate", "fdrq")
            _dm2k_fdrq = _to_date(_dm2k_startdate_raw) if _dm2k_startdate_raw else ""
            if entry.special_type in _SPECIAL_TYPE_LABEL:
                _dm2k_row_label = _SPECIAL_TYPE_LABEL[entry.special_type]
            elif _dm2k_fdrq:
                _dm2k_row_label = _dm2k_fdrq
            elif entry.report_date:
                _dm2k_row_label = _to_date(entry.report_date) or entry.report_date
            else:
                _dm2k_row_label = _dm2k_arch

            # Get all active batteries for auto-assignment fallback
            _dm2k_all_batys = _get_batys_for_archive(_dm2k_resolved)

            # Auto-detect groups from the archive's bz/remarks when the entry's groups
            # have no explicit tray assignments.  A bz like "LR6 UDP501 HP503" implies
            # two groups with positional tray assignment ([1-4] for group-0, [6-9] for
            # group-1).  Using all batteries for a single group would mix channels from
            # different production lines and produce an incorrect average.
            _dm2k_bz_raw = str(_dm2000_get_value(
                _arch_meta, "remarks", "remark", "bz", "note", "memo", "bzh",
            ) or "").strip()
            _dm2k_bz_groups = _parse_bz_groups(_dm2k_bz_raw)
            _dm2k_eff_groups = [
                {"loai": grp.loai, "chuyen": grp.chuyen, "trays": list(grp.trays or [])}
                for grp in entry.groups
            ]
            _dm2k_n = len(_dm2k_eff_groups)
            if len(_dm2k_bz_groups) > len(_dm2k_eff_groups) and not any(g["trays"] for g in _dm2k_eff_groups):
                # More groups in bz than entry specifies; use bz groups for full coverage
                _dm2k_eff_groups = [
                    {"loai": g["loai"], "chuyen": g["chuyen"], "trays": []}
                    for g in _dm2k_bz_groups
                ]
                _dm2k_n = len(_dm2k_eff_groups)
            # Correct loai from the archive's bz field (matched by chuyen) so that
            # entries whose groups_json was saved with an incorrect loai still
            # display the right battery grade (e.g. bz "LR6 UDP501 HP503" →
            # chuyen 501 → "UD+", chuyen 503 → "HP").
            _dm2k_bz_loai_by_chuyen: dict[str, str] = {
                str(g["chuyen"]): g["loai"]
                for g in _dm2k_bz_groups
                if g.get("chuyen")
            }
            if _dm2k_bz_loai_by_chuyen:
                _dm2k_eff_groups = [
                    {**eg, "loai": _dm2k_bz_loai_by_chuyen.get(
                        str(eg.get("chuyen") or "").strip(), eg["loai"]
                    )}
                    for eg in _dm2k_eff_groups
                ]
            _dm2k_auto_trays: list[list[int]] = _DMP_TRAY_ASSIGNMENT.get(
                _dm2k_n, [_dm2k_all_batys]
            )

            model_upper = entry.model.strip().upper()
            no_chuyen_models = {"LR61", "9V", "6LR61"}

            for _dm2k_g_idx, _dm2k_eff_grp in enumerate(_dm2k_eff_groups):
                _dm2k_grp_trays = _dm2k_eff_grp.get("trays") or []
                if _dm2k_grp_trays:
                    _dm2k_batys = [b for b in _dm2k_grp_trays if isinstance(b, int) and 1 <= b <= MAX_BATTERY_NUMBER]
                else:
                    _dm2k_batys = (
                        _dm2k_auto_trays[_dm2k_g_idx]
                        if _dm2k_g_idx < len(_dm2k_auto_trays)
                        else _dm2k_all_batys
                    )
                if not _dm2k_batys:
                    continue

                _dm2k_tav_map = _get_tav_for_batteries(_dm2k_resolved, _dm2k_batys)
                _dm2k_perf = _compute_perf_values(_dm2k_ep_str, _dm2k_tav_map, _dm2k_batys)
                _dm2k_perf["hfsj_unit"] = "hour"

                _dm2k_grp_chuyen = _dm2k_eff_grp.get("chuyen") or ""
                _dm2k_grp_loai = _dm2k_eff_grp.get("loai") or ""
                if model_upper in no_chuyen_models:
                    _dm2k_sheet_key = entry.model.strip()
                else:
                    _dm2k_sheet_key = f"{entry.model.strip()} {_dm2k_grp_chuyen.strip()}"

                _dm2k_fdfs_label = _dm2k_fdfs or _dm2k_grp_loai
                _dm2k_row_key = (_dm2k_row_label, _dm2k_grp_loai)
                groups.setdefault(_dm2k_sheet_key, {}).setdefault(_dm2k_row_key, {})[_dm2k_fdfs_label] = _dm2k_perf

            continue  # Skip DMP lookup path for this entry

        # Fetch batch metadata from DMPDATA.mdb.
        # batch_id may be a 6-digit DDMMYY date (the agreed convention for column-A
        # queries) rather than a literal para_pub.id.  When it matches that pattern,
        # convert it to a proper date and query by fdrq first; fall back to an id
        # lookup so that actual para_pub.id values still work.
        batch_rows = []
        _bid = entry.batch_id.strip()
        _ddmmyy_match = re.fullmatch(r"\d{6}", _bid)
        if _ddmmyy_match:
            _s = _bid
            try:
                _yy = int(_s[4:6])
                # Sliding-window century: treat YY within 50 years of today as 2000s,
                # anything older as 1900s (covers the practical 2000-2099 range of DMP data).
                _cur_yy = date.today().year % 100
                _century = 2000 if (_yy - _cur_yy) % 100 <= 50 else 1900
                _qdate = date(_century + _yy, int(_s[2:4]), int(_s[0:2]))
                try:
                    batch_rows = _read_dmpdata(
                        "SELECT * FROM para_pub WHERE fdrq = ?", (_qdate,)
                    )
                except pyodbc.Error:
                    try:
                        batch_rows = _read_dmpdata(
                            "SELECT id, dcxh, fdrq, fdfs, hfsj, zzdy, bz FROM para_pub WHERE fdrq = ?",
                            (_qdate,),
                        )
                    except pyodbc.Error as exc:
                        logger.warning(
                            "_compute_dmp_perf_groups: fdrq date query failed %s: %s",
                            entry.batch_id, exc,
                        )
            except (ValueError, TypeError):
                pass  # invalid DDMMYY — fall through to id lookup below

        if not batch_rows:
            # Direct id lookup (also handles non-DDMMYY batch_id values)
            try:
                batch_rows = _read_dmpdata(
                    "SELECT * FROM para_pub WHERE id = ?", (entry.batch_id,)
                )
            except pyodbc.Error:
                try:
                    batch_rows = _read_dmpdata(
                        "SELECT id, dcxh, fdrq, fdfs, hfsj, zzdy, bz FROM para_pub WHERE id = ?",
                        (entry.batch_id,),
                    )
                except pyodbc.Error as exc:
                    logger.warning("_compute_dmp_perf_groups: batch read failed %s: %s", entry.batch_id, exc)

        # Fallback: search para_pub.bz (remark field) using the raw_remark text.
        # This handles the case where the user entered a remark without a date
        # prefix (e.g. "LR6 UD501 UD502") so batch_id defaults to today's date
        # and the date-based lookup yields no results, but the corresponding
        # para_pub row can still be identified by its bz value.
        if not batch_rows and entry.raw_remark and entry.raw_remark.strip():
            _remark_search = entry.raw_remark.strip()
            # Strip leading 6-digit DDMMYY date prefix so "160226 LR6 UD501"
            # becomes "LR6 UD501" and matches bz = "160226 LR6 UD501 UD502".
            _remark_tokens = _remark_search.split()
            if _remark_tokens and re.fullmatch(r"\d{6}", _remark_tokens[0]):
                _remark_search = " ".join(_remark_tokens[1:]).strip()
            if _remark_search:
                # Leading wildcard is intentional: bz values often contain a
                # DDMMYY date prefix before the remark text (e.g. "160226 LR6 UD501").
                _bz_pattern = f"%{_remark_search}%"
                _bz_sql = (
                    "SELECT id, dcxh, fdrq, fdfs, hfsj, zzdy, bz FROM para_pub"
                    " WHERE bz LIKE ? ORDER BY fdrq DESC"
                )
                try:
                    batch_rows = _read_dmpdata(_bz_sql, (_bz_pattern,))
                except pyodbc.Error as exc:
                    logger.warning(
                        "_compute_dmp_perf_groups: bz LIKE lookup failed '%s': %s",
                        _remark_search, exc,
                    )

        if batch_rows:
            batch = batch_rows[0]
            # Resolve the actual para_pub.id so that _dmp_compute_group_perf can look
            # up the matching para_singl rows (which use sid = para_pub.id, not DDMMYY).
            actual_batch_id = str(_dm2000_get_value(batch, "id") or entry.batch_id)

            fdrq = _to_date(_dm2000_get_value(batch, "fdrq"))
            fdfs = str(_dm2000_get_value(batch, "fdfs") or "").strip()
            # When para_pub.fdfs is empty (DMP often leaves it blank), fall back to
            # para_pub.jstj (discharge test condition, e.g. "(1500mW2s,650mW28s)10T/h,24h/d").
            # This allows _perf_fdfs_matches_header to find the correct column in the
            # Excel template (e.g. "1500mW,(1500mW2s,650mW28s)10T/h,24h/d") instead of
            # falling through to the position-based fallback that writes to the first
            # RESULT column regardless of the actual test condition.
            if not fdfs:
                fdfs = str(_dm2000_get_value(batch, "jstj") or "").strip()
            # Normalise the unit from para_pub.hfsj ("minute"/"hour"/"times")
            _hfsj_raw = str(_dm2000_get_value(batch, "hfsj") or "").strip().lower()
            _HFSJ_TIMES = {"times", "lần", "t", "lan", "count"}
            _HFSJ_MINUTE = {"minute", "minutes", "phút", "m", "min", "phu"}
            if _hfsj_raw in _HFSJ_TIMES:
                hfsj_unit = "times"
            elif _hfsj_raw in _HFSJ_MINUTE:
                hfsj_unit = "minute"
            else:
                hfsj_unit = "hour"
            zzdy_raw = str(_dm2000_get_value(batch, "zzdy") or "").strip()

            # Parse endpoint voltage
            ep_v: Optional[float] = None
            if zzdy_raw:
                tok = re.sub(r"[^0-9.\-]+$", "", zzdy_raw.split()[0])
                try:
                    ep_v = float(tok)
                except (TypeError, ValueError):
                    ep_v = None
        else:
            # DMP batch not found. Try DM2000 path using raw_remark as the bz key.
            # This allows a single remark value (e.g. "LR6 UDP501 HP503") to match
            # either a DMP batch or a DM2000 archive transparently.
            if entry.raw_remark and entry.raw_remark.strip():
                _fb_remark = entry.raw_remark.strip()
                _fb_arch_rows: list[dict] = []
                try:
                    _fb_arch_rows = _read_dm2000_ls(
                        "SELECT * FROM ls_jb_cs WHERE bz = ?", (_fb_remark,)
                    )
                except pyodbc.Error:
                    pass
                if not _fb_arch_rows:
                    _fb_escaped = _fb_remark.replace("%", "[%]").replace("_", "[_]")
                    try:
                        _fb_arch_rows = _read_dm2000_ls(
                            "SELECT * FROM ls_jb_cs WHERE bz LIKE ?",
                            (f"%{_fb_escaped}%",),
                        )
                    except pyodbc.Error:
                        pass
                if _fb_arch_rows:
                    # When multiple archives are found and at least one can be paired
                    # with a group by production-line (chuyen), process every archive
                    # for its matching group so that:
                    #   • each group reads its own archive (correct grade/manufacturer),
                    #   • all discharge conditions across archives are collected.
                    # _pair_dm2000_archives_to_groups handles the ambiguous case where
                    # archives share the same manufacturer (e.g. both "501-502") so
                    # chuyen matching alone cannot distinguish them — it falls back to
                    # positional sort by min battery number.
                    _fb_multi = len(_fb_arch_rows) > 1 and len(entry.groups) > 0 and any(
                        _dm2000_archive_matches_chuyen(a, g.chuyen)
                        for a in _fb_arch_rows for g in entry.groups
                    )
                    if _fb_multi:
                        _fb_model_upper_m = entry.model.strip().upper()
                        _fb_no_chuyen_m = {"LR61", "9V", "6LR61"}
                        # When every archive matches every group the archives share the
                        # same remark and represent different test dates — each archive
                        # covers ALL production lines.  Process each archive with all
                        # groups via _parse_bz_groups + _DMP_TRAY_ASSIGNMENT (same
                        # logic the single-archive path applies correctly).
                        _fb_all_match_all = _dm2000_all_archives_match_all_groups(
                            _fb_arch_rows, entry.groups
                        )
                        if _fb_all_match_all:
                            for _fb_a in _fb_arch_rows:
                                _fb_a_resolved = (
                                    str(_dm2000_get_value(_fb_a, "cdid", "archname") or _fb_remark).strip()
                                    or _fb_remark
                                )
                                _fb_a_fdfs_raw = str(_dm2000_get_value(_fb_a, "fdfs") or "").strip()
                                _fb_a_load_res = str(_dm2000_get_value(
                                    _fb_a, "load_resistance", "fzdz", "fzlkdz", "dw", "fddl", "fdz", "resistance", "load_r", "r_ohm",
                                ) or "").strip()
                                _fb_a_ep_raw = _dm2000_get_value(
                                    _fb_a,
                                    "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
                                    "endpoint_v", "vcut", "cutoffv", "cutoff_v",
                                    "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
                                )
                                _fb_a_ep_str = str(_fb_a_ep_raw or "").strip()
                                _fb_a_fdfs = _build_dm2000_condition_label(
                                    _fb_a_fdfs_raw, _fb_a_load_res, _fb_a_ep_str, _fb_remark,
                                )
                                _fb_a_startdate = _dm2000_get_value(_fb_a, "startdate", "fdrq")
                                _fb_a_fdrq = _to_date(_fb_a_startdate) if _fb_a_startdate else ""
                                if entry.special_type in _SPECIAL_TYPE_LABEL:
                                    _fb_a_row_label = _SPECIAL_TYPE_LABEL[entry.special_type]
                                elif _fb_a_fdrq:
                                    _fb_a_row_label = _fb_a_fdrq
                                elif entry.report_date:
                                    _fb_a_row_label = _to_date(entry.report_date) or entry.report_date
                                else:
                                    _fb_a_row_label = _fb_remark
                                _fb_a_all_batys = _get_batys_for_archive(_fb_a_resolved)
                                _fb_a_bz_raw = str(_dm2000_get_value(
                                    _fb_a, "remarks", "remark", "bz", "note", "memo", "bzh",
                                ) or "").strip()
                                _fb_a_bz_groups = _parse_bz_groups(_fb_a_bz_raw)
                                _fb_a_eff_groups = [
                                    {"loai": grp.loai, "chuyen": grp.chuyen, "trays": list(grp.trays or [])}
                                    for grp in entry.groups
                                ]
                                _fb_a_n = len(_fb_a_eff_groups)
                                if len(_fb_a_bz_groups) > len(_fb_a_eff_groups) and not any(
                                    g["trays"] for g in _fb_a_eff_groups
                                ):
                                    _fb_a_eff_groups = [
                                        {"loai": g["loai"], "chuyen": g["chuyen"], "trays": []}
                                        for g in _fb_a_bz_groups
                                    ]
                                    _fb_a_n = len(_fb_a_eff_groups)
                                _fb_a_auto_trays = _DMP_TRAY_ASSIGNMENT.get(
                                    _fb_a_n, [_fb_a_all_batys]
                                )
                                for _fb_a_g_idx, _fb_a_eff_grp in enumerate(_fb_a_eff_groups):
                                    _fb_a_grp_trays = _fb_a_eff_grp.get("trays") or []
                                    if _fb_a_grp_trays:
                                        _fb_a_batys = [
                                            b for b in _fb_a_grp_trays
                                            if isinstance(b, int) and 1 <= b <= MAX_BATTERY_NUMBER
                                        ]
                                    else:
                                        _fb_a_batys = (
                                            _fb_a_auto_trays[_fb_a_g_idx]
                                            if _fb_a_g_idx < len(_fb_a_auto_trays)
                                            else _fb_a_all_batys
                                        )
                                    if not _fb_a_batys:
                                        continue
                                    _fb_a_tav = _get_tav_for_batteries(_fb_a_resolved, _fb_a_batys)
                                    _fb_a_perf = _compute_perf_values(_fb_a_ep_str, _fb_a_tav, _fb_a_batys)
                                    _fb_a_perf["hfsj_unit"] = "hour"
                                    _fb_a_grp_chuyen = _fb_a_eff_grp.get("chuyen") or ""
                                    _fb_a_grp_loai = _fb_a_eff_grp.get("loai") or ""
                                    _fb_a_sheet = (
                                        entry.model.strip()
                                        if _fb_model_upper_m in _fb_no_chuyen_m
                                        else f"{entry.model.strip()} {_fb_a_grp_chuyen.strip()}"
                                    )
                                    _fb_a_fdfs_label = _fb_a_fdfs or _fb_a_grp_loai
                                    groups.setdefault(_fb_a_sheet, {}).setdefault(
                                        (_fb_a_row_label, _fb_a_grp_loai), {}
                                    )[_fb_a_fdfs_label] = _fb_a_perf
                            continue  # all archives × all groups processed; skip DMP path
                        _fb_ai_gi_pairs = _pair_dm2000_archives_to_groups(_fb_arch_rows, entry.groups)
                        _fb_ai_to_gi = dict(_fb_ai_gi_pairs)
                        for _fb_ai, _fb_a in enumerate(_fb_arch_rows):
                            _fb_gi = _fb_ai_to_gi.get(_fb_ai)
                            if _fb_gi is None:
                                continue
                            _fb_grp = entry.groups[_fb_gi]
                            _fb_a_resolved = (
                                str(_dm2000_get_value(_fb_a, "cdid", "archname") or _fb_remark).strip()
                                or _fb_remark
                            )
                            _fb_a_fdfs_raw = str(_dm2000_get_value(_fb_a, "fdfs") or "").strip()
                            _fb_a_load_res = str(_dm2000_get_value(
                                _fb_a, "load_resistance", "fzdz", "fzlkdz", "dw", "fddl", "fdz", "resistance", "load_r", "r_ohm",
                            ) or "").strip()
                            _fb_a_ep_raw = _dm2000_get_value(
                                _fb_a,
                                "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
                                "endpoint_v", "vcut", "cutoffv", "cutoff_v",
                                "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
                            )
                            _fb_a_ep_str = str(_fb_a_ep_raw or "").strip()
                            _fb_a_fdfs = _build_dm2000_condition_label(
                                _fb_a_fdfs_raw, _fb_a_load_res, _fb_a_ep_str, _fb_remark,
                            )
                            _fb_a_startdate = _dm2000_get_value(_fb_a, "startdate", "fdrq")
                            _fb_a_fdrq = _to_date(_fb_a_startdate) if _fb_a_startdate else ""
                            if entry.special_type in _SPECIAL_TYPE_LABEL:
                                _fb_a_row_label = _SPECIAL_TYPE_LABEL[entry.special_type]
                            elif _fb_a_fdrq:
                                _fb_a_row_label = _fb_a_fdrq
                            elif entry.report_date:
                                _fb_a_row_label = _to_date(entry.report_date) or entry.report_date
                            else:
                                _fb_a_row_label = _fb_remark
                            _fb_a_all_batys = _get_batys_for_archive(_fb_a_resolved)
                            _fb_a_batys = (
                                [b for b in _fb_grp.trays if isinstance(b, int) and 1 <= b <= MAX_BATTERY_NUMBER]
                                if _fb_grp.trays else _fb_a_all_batys
                            )
                            if not _fb_a_batys:
                                continue
                            _fb_a_tav = _get_tav_for_batteries(_fb_a_resolved, _fb_a_batys)
                            _fb_a_perf = _compute_perf_values(_fb_a_ep_str, _fb_a_tav, _fb_a_batys)
                            _fb_a_perf["hfsj_unit"] = "hour"
                            _fb_a_sheet = (
                                entry.model.strip()
                                if _fb_model_upper_m in _fb_no_chuyen_m
                                else f"{entry.model.strip()} {_fb_grp.chuyen.strip()}"
                            )
                            _fb_a_fdfs_label = _fb_a_fdfs or _fb_grp.loai
                            groups.setdefault(_fb_a_sheet, {}).setdefault(
                                (_fb_a_row_label, _fb_grp.loai), {}
                            )[_fb_a_fdfs_label] = _fb_a_perf
                        continue  # DM2000 path handled; skip rest of DMP path

                    _fb_meta = _fb_arch_rows[0]
                    _fb_resolved = (
                        str(_dm2000_get_value(_fb_meta, "cdid", "archname") or _fb_remark).strip()
                        or _fb_remark
                    )
                    _fb_fdfs_raw = str(_dm2000_get_value(_fb_meta, "fdfs") or "").strip()
                    _fb_load_res = str(_dm2000_get_value(
                        _fb_meta, "load_resistance", "fzdz", "fzlkdz", "dw", "fddl", "fdz", "resistance", "load_r", "r_ohm",
                    ) or "").strip()
                    _fb_ep_raw = _dm2000_get_value(
                        _fb_meta,
                        "endpoint_voltage", "jzdy", "jzdianyi", "jzdv", "jz",
                        "endpoint_v", "vcut", "cutoffv", "cutoff_v",
                        "jzdian", "evy", "minv", "cutv", "jz_dy", "zzdy",
                    )
                    _fb_ep_str = str(_fb_ep_raw or "").strip()
                    _fb_fdfs = _build_dm2000_condition_label(
                        _fb_fdfs_raw, _fb_load_res, _fb_ep_str, _fb_remark,
                    )
                    # Determine row label — prefer the archive's own start date (fdrq)
                    # over entry.report_date, which is typically set to today when no
                    # date prefix was entered in the remark.
                    _fb_startdate_raw = _dm2000_get_value(_fb_meta, "startdate", "fdrq")
                    _fb_fdrq = _to_date(_fb_startdate_raw) if _fb_startdate_raw else ""
                    if entry.special_type in _SPECIAL_TYPE_LABEL:
                        _fb_row_label = _SPECIAL_TYPE_LABEL[entry.special_type]
                    elif _fb_fdrq:
                        _fb_row_label = _fb_fdrq
                    elif entry.report_date:
                        _fb_row_label = _to_date(entry.report_date) or entry.report_date
                    else:
                        _fb_row_label = _fb_remark
                    _fb_all_batys = _get_batys_for_archive(_fb_resolved)
                    # Auto-detect groups from the archive's bz/remarks when the entry's
                    # groups have no explicit tray assignments.  A bz like "LR6 UDP501
                    # HP503" implies two groups with positional tray assignment ([1-4]
                    # for group-0, [6-9] for group-1).  Using all batteries for a single
                    # group would mix channels from different production lines.
                    _fb_bz_raw = str(_dm2000_get_value(
                        _fb_meta, "remarks", "remark", "bz", "note", "memo", "bzh",
                    ) or "").strip()
                    _fb_bz_groups = _parse_bz_groups(_fb_bz_raw)
                    _fb_eff_groups = [
                        {"loai": grp.loai, "chuyen": grp.chuyen, "trays": list(grp.trays or [])}
                        for grp in entry.groups
                    ]
                    _fb_n = len(_fb_eff_groups)
                    if len(_fb_bz_groups) > len(_fb_eff_groups) and not any(g["trays"] for g in _fb_eff_groups):
                        _fb_eff_groups = [
                            {"loai": g["loai"], "chuyen": g["chuyen"], "trays": []}
                            for g in _fb_bz_groups
                        ]
                        _fb_n = len(_fb_eff_groups)
                    _fb_auto_trays = _DMP_TRAY_ASSIGNMENT.get(_fb_n, [_fb_all_batys])
                    _fb_model_upper = entry.model.strip().upper()
                    _fb_no_chuyen = {"LR61", "9V", "6LR61"}
                    for _fb_g_idx, _fb_eff_grp in enumerate(_fb_eff_groups):
                        _fb_grp_trays = _fb_eff_grp.get("trays") or []
                        if _fb_grp_trays:
                            _fb_batys = [b for b in _fb_grp_trays if isinstance(b, int) and 1 <= b <= MAX_BATTERY_NUMBER]
                        else:
                            _fb_batys = (
                                _fb_auto_trays[_fb_g_idx]
                                if _fb_g_idx < len(_fb_auto_trays)
                                else _fb_all_batys
                            )
                        if not _fb_batys:
                            continue
                        _fb_tav = _get_tav_for_batteries(_fb_resolved, _fb_batys)
                        _fb_perf = _compute_perf_values(_fb_ep_str, _fb_tav, _fb_batys)
                        _fb_perf["hfsj_unit"] = "hour"
                        _fb_grp_chuyen = _fb_eff_grp.get("chuyen") or ""
                        _fb_grp_loai = _fb_eff_grp.get("loai") or ""
                        _fb_sheet = (
                            entry.model.strip()
                            if _fb_model_upper in _fb_no_chuyen
                            else f"{entry.model.strip()} {_fb_grp_chuyen.strip()}"
                        )
                        _fb_fdfs_label = _fb_fdfs or _fb_grp_loai
                        groups.setdefault(_fb_sheet, {}).setdefault(
                            (_fb_row_label, _fb_grp_loai), {}
                        )[_fb_fdfs_label] = _fb_perf
                    continue  # DM2000 path handled; skip rest of DMP path

            # Both DMP and DM2000 lookups failed.
            logger.warning("_compute_dmp_perf_groups: batch not found: %s", entry.batch_id)
            # For the web JSON preview we skip entries with no data entirely
            # so that spurious rows (empty date) and spurious columns (grp.loai
            # used as fdfs_label instead of the real condition string) do not
            # appear in the table.  The Excel report generator keeps the old
            # behaviour (empty cells) so users can still review missing batches.
            if skip_not_found:
                continue
            actual_batch_id = entry.batch_id
            fdfs = ""
            hfsj_unit = "hour"
            ep_v = None

            # Derive a proper YYYY-MM-DD row label from the available fields.
            # Priority: report_date (from SQLite) > DDMMYY conversion > raw batch_id.
            if entry.report_date:
                fdrq = _to_date(entry.report_date) or entry.batch_id
            elif _ddmmyy_match:
                try:
                    _yy2 = int(_bid[4:6])
                    _cur_yy2 = date.today().year % 100
                    _century2 = 2000 if (_yy2 - _cur_yy2) % 100 <= 50 else 1900
                    fdrq = date(_century2 + _yy2, int(_bid[2:4]), int(_bid[0:2])).strftime("%Y-%m-%d")
                except (ValueError, TypeError):
                    fdrq = entry.batch_id
            else:
                fdrq = entry.batch_id

        # Determine row label (date or special)
        if entry.special_type in _SPECIAL_TYPE_LABEL:
            row_label = _SPECIAL_TYPE_LABEL[entry.special_type]
        else:
            row_label = fdrq or entry.batch_id

        # Auto-assign trays if not specified
        n_groups = len(entry.groups)
        auto_trays = _DMP_TRAY_ASSIGNMENT.get(n_groups, [list(range(1, 10))])

        # Re-derive authoritative loai by chuyen from raw_remark so that stored
        # entries with an incorrect loai in groups_json still display the correct
        # battery grade.  Example: raw_remark "LR6 UDP501 HP503" yields the
        # mapping {"501": "UD+", "503": "HP"}.  If raw_remark is absent or a
        # chuyen has no match, the stored grp.loai is used as fallback.
        _remark_loai_by_chuyen: dict[str, str] = {}
        if entry.raw_remark:
            for _rg in _parse_bz_groups(entry.raw_remark):
                if _rg.get("chuyen"):
                    _remark_loai_by_chuyen[str(_rg["chuyen"])] = _rg["loai"]

        for g_idx, grp in enumerate(entry.groups):
            trays = grp.trays if grp.trays else (auto_trays[g_idx] if g_idx < len(auto_trays) else [])
            if not trays:
                continue

            # Compute performance values for this group's trays using the resolved
            # para_pub.id (actual_batch_id) so that the para_singl lookup succeeds.
            perf = _dmp_compute_group_perf(actual_batch_id, trays, ep_v)
            perf["hfsj_unit"] = hfsj_unit

            # Determine sheet name
            model_upper = entry.model.strip().upper()
            # Models without production-line requirement use only the model name
            no_chuyen_models = {"LR61", "9V", "6LR61"}
            if model_upper in no_chuyen_models:
                sheet_key = entry.model.strip()
            else:
                sheet_key = f"{entry.model.strip()} {grp.chuyen.strip()}"

            # Use loai from raw_remark (re-parsed) when available so that entries
            # whose groups_json was saved with an incorrect loai are displayed with
            # the correct battery grade derived from the remark token (UDP→UD+,
            # HP→HP, UD→UD) matched by chuyen.
            effective_loai = _remark_loai_by_chuyen.get(grp.chuyen.strip(), grp.loai)

            # fdfs label for column matching
            fdfs_label = fdfs if fdfs else effective_loai

            row_key = (row_label, effective_loai)
            groups.setdefault(sheet_key, {}).setdefault(row_key, {})[fdfs_label] = perf

    return groups


@app.post("/dmp-perf-report/generate")
def generate_dmp_perf_report(payload: DmpPerfReportRequest):
    """Generate a DMP battery performance report (Bảng theo dõi hiệu suất pin).

    Each entry maps a DMP batch (batch_id) to one or more production-line groups.
    The result is written into the appropriate sheet/row/column of the Excel template.
    """
    if not payload.entries:
        raise HTTPException(status_code=400, detail="entries must not be empty")

    groups = _compute_dmp_perf_groups(payload)

    if not groups:
        raise HTTPException(
            status_code=422, detail="No data could be extracted for any entry"
        )

    if not payload.template_name:
        raise HTTPException(
            status_code=400,
            detail="template_name is required for DMP perf report generation"
        )

    template_path = _resolve_dmp_perf_template_path(payload.template_name)
    workbook_bytes = _render_perf_template(template_path, groups)
    filename = payload.template_name
    return StreamingResponse(
        BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/dmp-perf-data")
def get_dmp_perf_data(payload: DmpPerfReportRequest):
    """Return DMP performance data as JSON for web visualization.

    Same computation as ``/dmp-perf-report/generate`` but returns JSON instead
    of filling an Excel template, enabling the web UI to render an interactive
    preview table.

    Response shape::

        {
          "sheets": {
            "LR6 501": {
              "rows": [
                {
                  "date": "2026-01-01",
                  "loai": "UD",
                  "conditions": {
                    "<fdfs_label>": {
                      "avg_hours": 25.3,
                      "avg_minutes": 1518.0,
                      "avg_count": 12,
                      "uniform_rate": 96.5
                    }
                  }
                }
              ],
              "conditions": ["<fdfs_label1>", ...]
            }
          }
        }
    """
    if not payload.entries:
        raise HTTPException(status_code=400, detail="entries must not be empty")

    # skip_not_found=True: entries with no matching DMP data are omitted from the
    # web preview so the table does not show spurious empty rows or columns whose
    # header is the battery grade (e.g. "UD+") rather than the real test condition.
    groups = _compute_dmp_perf_groups(payload, skip_not_found=True)

    sheets: dict = {}
    for sheet_key, date_type_map in groups.items():
        # Collect all unique condition labels for this sheet (preserving insertion order)
        all_conditions: list[str] = []
        seen: set[str] = set()
        for row_data in date_type_map.values():
            for lbl in row_data:
                if lbl not in seen:
                    seen.add(lbl)
                    all_conditions.append(lbl)

        # Build a per-sheet units dict: fdfs_label → "hour"|"minute"|"times".
        # Built before sorting so unit type can be used as a primary sort key.
        # Take the first non-null value seen across all rows for each condition.
        units: dict[str, str] = {}
        for row_data in date_type_map.values():
            for lbl, perf in row_data.items():
                if lbl not in units and perf.get("hfsj_unit"):
                    units[lbl] = perf["hfsj_unit"]

        # Sort conditions: use IEC template order for the detected battery family as
        # the primary sort key.  Conditions not found in the template fall back to the
        # legacy unit-type ordering (times → minute → hour) so that reports for
        # non-standard battery types are still reasonable.
        # Battery type is the first whitespace-separated token of the sheet key
        # (e.g. "LR6" from "LR6 501"; "|" candidates are stripped first).
        _sheet_key_clean = sheet_key.split("|")[0].strip()
        _battery_type = _sheet_key_clean.split()[0] if _sheet_key_clean else ""
        _template = _TEMPLATE_CONDITION_ORDER.get(_battery_type.upper(), [])
        _cond_first: dict[str, str] = {}
        for (row_lbl, _), rd in date_type_map.items():
            for lbl in rd:
                if lbl not in _cond_first or row_lbl < _cond_first[lbl]:
                    _cond_first[lbl] = row_lbl
        _UNIT_ORDER = {"times": 0, "minute": 1, "hour": 2}
        _tmpl_len = len(_template)

        def _sort_key(c: str) -> tuple:
            tmpl_pos = _template_condition_sort_key(c, _battery_type)[0]
            if tmpl_pos < _tmpl_len:
                # Found in template: primary key is template position, tie-break by label.
                # The unit-order slot is kept for tuple-length parity with the fallback branch.
                return (0, tmpl_pos, _UNIT_ORDER.get(units.get(c, "hour"), 2), c)
            # Not in template: sort after template conditions using unit-type then first-date.
            # The constant 0 in slot 2 is a parity placeholder (same tuple length as above).
            return (
                1,
                _UNIT_ORDER.get(units.get(c, "hour"), 2),
                0,  # parity placeholder — slot unused when group=1
                _cond_first.get(c, "9999-99-99"),
            )

        all_conditions.sort(key=_sort_key)

        rows = []
        for (row_label, loai), conditions in sorted(date_type_map.items(), key=lambda x: x[0]):
            rows.append({
                "date": row_label,
                "loai": loai,
                "conditions": {
                    fdfs_label: {
                        "avg_hours": perf.get("avg_hours"),
                        "avg_minutes": perf.get("avg_minutes"),
                        "avg_count": perf.get("avg_count"),
                        "uniform_rate": perf.get("uniform_rate"),
                    }
                    for fdfs_label, perf in conditions.items()
                },
            })

        sheets[sheet_key] = {"rows": rows, "conditions": all_conditions, "units": units}

    return {"sheets": sheets}
