import logging
import math
import os
import re
import shutil
import socket
import statistics
import tempfile
import threading
import time
from contextlib import asynccontextmanager, contextmanager
from copy import copy
from io import BytesIO
from pathlib import Path

import pyodbc
import requests
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("dmp_service")

STATION_NAME: str = os.environ.get("DMP_STATION_NAME", "").strip()
VONIKO_SERVER_URL: str = os.environ.get("VONIKO_SERVER_URL", "").rstrip("/")
STATION_PORT: int = int(os.environ.get("DMP_STATION_PORT", "8766"))
DMP_DATA_DIR: str = os.environ.get("DMP_DATA_DIR", r"C:\DMP\Data")
DMP_TEMPLATES_DIR: str = os.environ.get("DMP_TEMPLATES_DIR", "./dmp_templates")


def _get_local_ip() -> str:
    try:
        host = VONIKO_SERVER_URL.split("://")[-1].split(":")[0].split("/")[0]
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect((host, 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def _registration_loop() -> None:
    while True:
        try:
            ip = _get_local_ip()
            url = f"http://{ip}:{STATION_PORT}"
            requests.post(
                f"{VONIKO_SERVER_URL}/api/dmp/register",
                json={"name": STATION_NAME, "url": url},
                timeout=5,
            )
        except Exception as exc:
            logger.error("DMP registration heartbeat failed: %s", exc)
        time.sleep(30)


@asynccontextmanager
async def _lifespan(_app: FastAPI):
    if VONIKO_SERVER_URL and STATION_NAME:
        t = threading.Thread(target=_registration_loop, daemon=True)
        t.start()
        logger.info("DMP registration thread started for station '%s'", STATION_NAME)
    yield


app = FastAPI(title="DMP Battery Data Bridge", version="1.0.0", lifespan=_lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class ReportRequest(BaseModel):
    batch_id: str
    cdmc: str
    channel: int
    template_name: str


@contextmanager
def shadow_copy(mdb_path: str):
    """Copy live .mdb to temp before querying to avoid file-lock errors."""
    suffix = Path(mdb_path).suffix
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    os.close(tmp_fd)
    try:
        shutil.copy2(mdb_path, tmp_path)
        yield tmp_path
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def query_mdb(mdb_path: str, sql: str, params: tuple = ()) -> list[dict]:
    conn_str = (
        r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
        f"DBQ={mdb_path};"
    )
    with pyodbc.connect(conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description] if cursor.description else []
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _is_inside(base: Path, target: Path) -> bool:
    try:
        target.resolve().relative_to(base.resolve())
        return True
    except ValueError:
        return False


def _resolve_data_mdb(cdmc: str) -> Path:
    raw = (cdmc or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="cdmc is required")

    file_name = Path(raw).name
    if file_name != raw:
        raise HTTPException(status_code=400, detail="Invalid cdmc path")

    stem = Path(file_name).stem
    suffix = Path(file_name).suffix.lower()
    if suffix not in ("", ".mdb"):
        raise HTTPException(status_code=400, detail="cdmc must be .mdb")
    if not re.fullmatch(r"[A-Za-z0-9_-]+", stem):
        raise HTTPException(status_code=400, detail="Invalid cdmc filename")

    normalized = f"{stem}.mdb"
    data_dir = Path(DMP_DATA_DIR).resolve()
    mdb_path = (data_dir / normalized).resolve()
    if not _is_inside(data_dir, mdb_path):
        raise HTTPException(status_code=400, detail="Invalid cdmc path")
    if not mdb_path.exists() or not mdb_path.is_file():
        raise HTTPException(status_code=404, detail="Data file not found")
    return mdb_path


def _resolve_template_path(template_name: str) -> Path:
    raw = (template_name or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="template_name is required")

    file_name = Path(raw).name
    if file_name != raw or Path(file_name).suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Invalid template name")

    templates_dir = Path(DMP_TEMPLATES_DIR).resolve()
    template_path = (templates_dir / file_name).resolve()
    if not _is_inside(templates_dir, template_path):
        raise HTTPException(status_code=400, detail="Invalid template path")
    if not template_path.exists() or not template_path.is_file():
        raise HTTPException(status_code=404, detail="Template not found")
    return template_path


def _to_number(value):
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    return num if math.isfinite(num) else None


def _calc_stats(rows: list[dict]) -> dict:
    volt = [_to_number(row.get("VOLT")) for row in rows]
    curr = [_to_number(row.get("Im")) for row in rows]
    volt_values = [v for v in volt if v is not None]
    curr_values = [v for v in curr if v is not None]

    def _agg(values: list[float]):
        if not values:
            return None, None, None
        return max(values), min(values), statistics.mean(values)

    vmax, vmin, vavg = _agg(volt_values)
    imax, imin, iavg = _agg(curr_values)

    return {
        "VOLT_MAX": vmax,
        "VOLT_MIN": vmin,
        "VOLT_AVG": vavg,
        "IM_MAX": imax,
        "IM_MIN": imin,
        "IM_AVG": iavg,
    }


def _interpolate_text(text: str, context: dict) -> str:
    if not isinstance(text, str):
        return text

    def _replace(match: re.Match) -> str:
        key = match.group(1)
        value = context.get(key, "")
        return "" if value is None else str(value)

    return re.sub(r"\{\{\s*([A-Za-z0-9_]+)\s*\}\}", _replace, text)


def _copy_row(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.value = src.value
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def _process_sheet(ws, ctx: dict) -> None:
    open_tag = re.compile(r"\{\{\s*#([A-Za-z0-9_]+)\s*\}\}")
    close_tag = re.compile(r"\{\{\s*/([A-Za-z0-9_]+)\s*\}\}")
    strip_array_tags = re.compile(r"\{\{\s*[#/][A-Za-z0-9_]+\s*\}\}")

    row_idx = 1
    while row_idx <= ws.max_row:
        row_cells = [ws.cell(row=row_idx, column=c) for c in range(1, ws.max_column + 1)]
        values = [cell.value for cell in row_cells if isinstance(cell.value, str)]

        found_open = None
        found_close = None
        for value in values:
            m_open = open_tag.search(value)
            if m_open:
                found_open = m_open.group(1)
            m_close = close_tag.search(value)
            if m_close:
                found_close = m_close.group(1)

        if found_open and found_open == found_close:
            items = ctx.get(found_open, [])
            if not isinstance(items, list):
                items = []

            if not items:
                ws.delete_rows(row_idx, 1)
                continue

            for item_idx, item in enumerate(items):
                target_row = row_idx + item_idx
                if item_idx > 0:
                    ws.insert_rows(target_row, 1)
                    _copy_row(ws, row_idx, target_row)

                local_ctx = dict(ctx)
                if isinstance(item, dict):
                    local_ctx.update(item)

                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=target_row, column=col)
                    if isinstance(cell.value, str):
                        without_tags = strip_array_tags.sub("", cell.value)
                        cell.value = _interpolate_text(without_tags, local_ctx)

            row_idx += len(items)
            continue

        row_idx += 1

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = _interpolate_text(cell.value, ctx)


def render_excel_template(template_path: str, context: dict) -> bytes:
    wb = load_workbook(template_path)
    for ws in wb.worksheets:
        _process_sheet(ws, context)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _query_data_mdb(cdmc: str, channel: int) -> list[dict]:
    mdb_path = _resolve_data_mdb(cdmc)
    sql = "SELECT baty, TIM, VOLT, Im FROM vidata WHERE baty = ? ORDER BY TIM ASC"
    with shadow_copy(str(mdb_path)) as copied:
        return query_mdb(copied, sql, (channel,))


@app.get("/")
def health_root():
    return {"service": "DMP Battery Data Bridge", "version": "1.0.0", "status": "ok"}


@app.get("/batches")
def get_batches():
    try:
        dmpdata = (Path(DMP_DATA_DIR).resolve() / "DMPDATA.mdb").resolve()
        if not dmpdata.exists():
            raise HTTPException(status_code=404, detail="DMPDATA.mdb not found")
        with shadow_copy(str(dmpdata)) as copied:
            rows = query_mdb(copied, "SELECT id, dcxh, fdrq, fdfs FROM para_pub ORDER BY fdrq DESC")
        return {"batches": rows}
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch batches")
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/batches/{batch_id}/channels")
def get_channels(batch_id: str):
    try:
        dmpdata = (Path(DMP_DATA_DIR).resolve() / "DMPDATA.mdb").resolve()
        if not dmpdata.exists():
            raise HTTPException(status_code=404, detail="DMPDATA.mdb not found")
        with shadow_copy(str(dmpdata)) as copied:
            rows = query_mdb(copied, "SELECT baty, cdmc FROM para_singl WHERE id = ?", (batch_id,))
        return {"channels": rows}
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch channels")
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/telemetry")
def get_telemetry(cdmc: str = Query(...), channel: int = Query(...)):
    try:
        telemetry = _query_data_mdb(cdmc, channel)
        return {"telemetry": telemetry}
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch telemetry")
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/stats")
def get_stats(cdmc: str = Query(...), channel: int = Query(...)):
    try:
        telemetry = _query_data_mdb(cdmc, channel)
        return _calc_stats(telemetry)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to compute stats")
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/templates")
def get_templates():
    try:
        templates_dir = Path(DMP_TEMPLATES_DIR).resolve()
        if not templates_dir.exists():
            return {"templates": []}
        templates = sorted([
            p.name
            for p in templates_dir.iterdir()
            if p.is_file() and p.suffix.lower() == ".xlsx"
        ])
        return {"templates": templates}
    except Exception as exc:
        logger.exception("Failed to list templates")
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/report")
def generate_report(request: ReportRequest):
    try:
        dmpdata = (Path(DMP_DATA_DIR).resolve() / "DMPDATA.mdb").resolve()
        if not dmpdata.exists():
            raise HTTPException(status_code=404, detail="DMPDATA.mdb not found")

        with shadow_copy(str(dmpdata)) as copied:
            batch_rows = query_mdb(
                copied,
                "SELECT id, dcxh, fdrq, fdfs FROM para_pub WHERE id = ?",
                (request.batch_id,),
            )

        if not batch_rows:
            raise HTTPException(status_code=404, detail="Batch not found")

        telemetry = _query_data_mdb(request.cdmc, request.channel)
        stats = _calc_stats(telemetry)
        batch = batch_rows[0]

        context = {
            "BATCH_ID": batch.get("id"),
            "MODEL": batch.get("dcxh"),
            "DATE": batch.get("fdrq"),
            "DISCHARGE_PATTERN": batch.get("fdfs"),
            "CHANNEL": request.channel,
            **stats,
            "HISTORY_DATA": [
                {
                    "TIM": row.get("TIM"),
                    "VOLT": row.get("VOLT"),
                    "Im": row.get("Im"),
                    "BATY": row.get("baty"),
                }
                for row in telemetry
            ],
        }

        template_path = _resolve_template_path(request.template_name)
        report_bytes = render_excel_template(str(template_path), context)

        safe_batch = re.sub(r"[^A-Za-z0-9._-]", "_", str(request.batch_id))
        safe_channel = re.sub(r"[^A-Za-z0-9._-]", "_", str(request.channel))
        filename = f"dmp_report_{safe_batch}_{safe_channel}.xlsx"

        return StreamingResponse(
            BytesIO(report_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to generate report")
        raise HTTPException(status_code=500, detail=str(exc)) from exc
