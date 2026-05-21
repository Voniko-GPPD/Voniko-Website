import io
import os
import uuid
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse

import openpyxl
from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import desc, extract, func, inspect, or_, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Query as SAQuery
from sqlalchemy.orm import Session

from .database import Base, SessionLocal, engine, get_db
from .default_dictionaries import DEFAULT_DICTIONARIES, DEFAULT_DICTIONARY_SEED_VERSION
from .models import (
    DefectType,
    GradeMapping,
    LineMapping,
    MonthMapping,
    ProductionOutput,
    QualityRecord,
    SuffixMapping,
    SystemSetting,
    YearMapping,
)
from .schemas import (
    DefectTypeCreate,
    DefectTypeOut,
    DefectTypeUpdate,
    GradeMappingCreate,
    GradeMappingOut,
    GradeMappingUpdate,
    LineMappingCreate,
    LineMappingOut,
    LineMappingUpdate,
    MonthlyPpmRow,
    MonthlySummaryRow,
    MonthMappingCreate,
    MonthMappingOut,
    MonthMappingUpdate,
    OCRIngestRequest,
    OCRIngestResponse,
    ParseRequest,
    ParseResponse,
    ProductionOutputCreate,
    ProductionOutputOut,
    ProductionOutputUpdate,
    QualityRecordCreate,
    QualityRecordOut,
    SuffixMappingCreate,
    SuffixMappingOut,
    SuffixMappingUpdate,
    YearMappingCreate,
    YearMappingOut,
    YearMappingUpdate,
    YearlySummaryRow,
)
from .services.parser import ParseCodeError, parse_battery_codes

app = FastAPI(title="BQMS API", version="1.0.0")

_BASE = Path(__file__).resolve().parent.parent.parent
_DATA_DIR = _BASE / "backend" / "data"
_DATA_DIR.mkdir(parents=True, exist_ok=True)
_QC_MEDIA_BASE_URL = os.getenv("QC_MEDIA_BASE_URL", "/uploads/qc").rstrip("/")
_PPM_DECIMALS = 4

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3001",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = _DATA_DIR / "qc-uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/uploads/qc", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads_qc")


@app.on_event("startup")
def startup() -> None:
    Base.metadata.create_all(bind=engine)
    _apply_runtime_migrations()
    _seed_default_dictionaries_once()


def _ensure_column(conn, table_name: str, column_name: str, ddl: str) -> None:
    cols = {c["name"] for c in inspect(conn).get_columns(table_name)}
    if column_name not in cols:
        conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {ddl}"))


def _round_ppm(value: float) -> float:
    return float(f"{value:.{_PPM_DECIMALS}f}")


def _apply_runtime_migrations() -> None:
    # Lightweight SQLite-safe migration path for local upgrades.
    with engine.begin() as conn:
        insp = inspect(conn)
        if insp.has_table("line_mapping"):
            _ensure_column(conn, "line_mapping", "battery_model", "battery_model VARCHAR(64)")

        if insp.has_table("quality_records"):
            _ensure_column(conn, "quality_records", "parsed_line_code", "parsed_line_code VARCHAR(16)")
            _ensure_column(conn, "quality_records", "parsed_line_desc", "parsed_line_desc VARCHAR(64)")
            _ensure_column(conn, "quality_records", "parsed_battery_model", "parsed_battery_model VARCHAR(64)")
            _ensure_column(conn, "quality_records", "parsed_station_no", "parsed_station_no VARCHAR(2)")
            _ensure_column(conn, "quality_records", "photo_url", "photo_url VARCHAR(255)")
            _ensure_column(conn, "quality_records", "defect_description", "defect_description VARCHAR(255)")

        # Safe index creation
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_quality_records_line_code ON quality_records(parsed_line_code)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_quality_records_model ON quality_records(parsed_battery_model)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_quality_records_station ON quality_records(parsed_station_no)"))


def _seed_default_dictionaries_once() -> None:
    seed_key = "default_dictionary_seed_version"
    db = SessionLocal()
    try:
        existing_marker = db.get(SystemSetting, seed_key)
        if existing_marker and existing_marker.value == DEFAULT_DICTIONARY_SEED_VERSION:
            return

        registry = [
            ("line_mappings", LineMapping, "code"),
            ("defect_types", DefectType, "name"),
            ("year_mappings", YearMapping, "code"),
            ("month_mappings", MonthMapping, "code"),
            ("grade_mappings", GradeMapping, "code"),
            ("suffix_mappings", SuffixMapping, "suffix_code"),
        ]

        for seed_name, model, unique_field in registry:
            for payload in DEFAULT_DICTIONARIES[seed_name]:
                unique_value = payload[unique_field]
                item = db.query(model).filter(getattr(model, unique_field) == unique_value).first()
                if item:
                    for key, value in payload.items():
                        setattr(item, key, value)
                else:
                    db.add(model(**payload))

        if existing_marker:
            existing_marker.value = DEFAULT_DICTIONARY_SEED_VERSION
        else:
            db.add(SystemSetting(key=seed_key, value=DEFAULT_DICTIONARY_SEED_VERSION))
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def commit_or_400(db: Session) -> None:
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail="数据唯一性冲突，请检查编码或名称是否重复") from exc


def normalize_code(code: str) -> str:
    return code.strip().upper()


def get_model_or_404(db: Session, model: Any, item_id: int, model_name: str):
    obj = db.get(model, item_id)
    if not obj:
        raise HTTPException(status_code=404, detail=f"{model_name}不存在")
    return obj


def _csv_strings(value: str | None) -> list[str] | None:
    if not value:
        return None
    items = [x.strip() for x in value.split(",") if x.strip()]
    return items or None


def _csv_ints(value: str | None) -> list[int] | None:
    if not value:
        return None
    result: list[int] = []
    for token in value.split(","):
        token = token.strip()
        if not token:
            continue
        try:
            result.append(int(token))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"无效的数字参数: {token}") from exc
    return result or None


def _save_photo(photo: UploadFile | None) -> str | None:
    if not photo:
        return None
    filename = photo.filename or ""
    ext = Path(filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".bmp", ".webp"}:
        raise HTTPException(status_code=400, detail="照片仅支持 jpg/jpeg/png/bmp/webp 格式")
    new_name = f"{uuid.uuid4().hex}{ext}"
    dst = UPLOAD_DIR / new_name
    content = photo.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传照片为空")
    with dst.open("wb") as f:
        f.write(content)
    return f"{_QC_MEDIA_BASE_URL}/{new_name}"


def _remove_photo(photo_url: str | None) -> None:
    if not photo_url:
        return
    parsed_path = urlparse(photo_url).path if "://" in photo_url else photo_url
    filename = Path(parsed_path).name
    if not filename:
        return
    path = UPLOAD_DIR / filename
    if path.exists():
        path.unlink()


def _get_line_info(db: Session, line_code: str) -> tuple[str | None, str | None]:
    line = db.query(LineMapping).filter(LineMapping.code == line_code).first()
    if not line:
        return None, None
    return line.line_desc, line.battery_model


def build_record_out(record: QualityRecord) -> QualityRecordOut:
    line_code = record.parsed_line_code or ""
    line_desc = record.parsed_line_desc or record.parsed_line or ""
    station_no = record.parsed_station_no or ""
    return QualityRecordOut(
        id=record.id,
        record_time=record.record_time,
        detected_date=record.detected_date,
        upper_code=record.upper_code,
        lower_code=record.lower_code,
        parsed_line=record.parsed_line,
        parsed_line_code=line_code,
        parsed_line_desc=line_desc,
        parsed_battery_model=record.parsed_battery_model,
        parsed_station_no=station_no,
        parsed_production_time=record.parsed_production_time,
        parsed_grade=record.parsed_grade,
        parsed_special_status=record.parsed_special_status,
        photo_url=record.photo_url,
        defect_description=record.defect_description,
        defect_type_id=record.defect_type_id,
        defect_type_name=record.defect_type.name,
        operator_name=record.operator_name,
    )


def _create_quality_record_entity(
    db: Session,
    detected_date: date,
    upper_code: str,
    lower_code: str,
    defect_type_id: int,
    operator_name: str,
    defect_description: str | None = None,
    photo_url: str | None = None,
) -> QualityRecord:
    defect_type = get_model_or_404(db, DefectType, defect_type_id, "不良类型")
    if not defect_type.status:
        raise HTTPException(status_code=400, detail="所选不良类型已禁用，请联系管理员")

    try:
        parsed = parse_battery_codes(db, upper_code, lower_code)
    except ParseCodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    record = QualityRecord(
        detected_date=detected_date,
        upper_code=upper_code.strip().upper(),
        lower_code=lower_code.strip(),
        parsed_line=parsed["production_line"],
        parsed_line_code=parsed["line_code"],
        parsed_line_desc=parsed["line_desc"],
        parsed_battery_model=parsed.get("battery_model"),
        parsed_station_no=parsed["station_no"],
        parsed_production_time=parsed["production_time"],
        parsed_grade=parsed["grade"],
        parsed_special_status=parsed["special_status"],
        photo_url=photo_url,
        defect_description=defect_description.strip() if defect_description else None,
        defect_type_id=defect_type_id,
        operator_name=operator_name.strip(),
    )
    db.add(record)
    commit_or_400(db)
    db.refresh(record)
    db.refresh(defect_type)
    return record


def _apply_record_filters(
    query: SAQuery,
    *,
    start_date: date | None,
    end_date: date | None,
    defect_type_ids: list[int] | None,
    line_codes: list[str] | None,
    battery_models: list[str] | None,
    station_nos: list[str] | None,
    keyword: str | None,
) -> SAQuery:
    if start_date:
        query = query.filter(QualityRecord.detected_date >= start_date)
    if end_date:
        query = query.filter(QualityRecord.detected_date <= end_date)
    if defect_type_ids:
        query = query.filter(QualityRecord.defect_type_id.in_(defect_type_ids))
    if line_codes:
        query = query.filter(QualityRecord.parsed_line_code.in_(line_codes))
    if battery_models:
        query = query.filter(QualityRecord.parsed_battery_model.in_(battery_models))
    if station_nos:
        query = query.filter(QualityRecord.parsed_station_no.in_(station_nos))
    if keyword:
        kw = f"%{keyword.strip()}%"
        query = query.filter(
            or_(
                QualityRecord.upper_code.ilike(kw),
                QualityRecord.lower_code.ilike(kw),
                QualityRecord.parsed_line.ilike(kw),
                QualityRecord.parsed_grade.ilike(kw),
                QualityRecord.parsed_special_status.ilike(kw),
                QualityRecord.operator_name.ilike(kw),
                DefectType.name.ilike(kw),
            )
        )
    return query


@app.get("/api/v1/health")
def health() -> dict:
    return {"status": "ok"}


# ── Dict registry (export / import shared logic) ──────────────────────────────

DICT_REGISTRY: dict[str, dict] = {
    "line-mappings": {
        "model": LineMapping,
        "label": "生产线映射",
        "unique_field": "code",
        "columns": [
            {"field": "code", "header": "代码字母", "type": "str", "required": True},
            {"field": "line_desc", "header": "生产线描述", "type": "str", "required": True},
            {"field": "battery_model", "header": "电池型号", "type": "str", "required": False},
            {"field": "status", "header": "状态(1/0)", "type": "bool", "required": False},
        ],
    },
    "defect-types": {
        "model": DefectType,
        "label": "不良类型",
        "unique_field": "name",
        "columns": [
            {"field": "name", "header": "名称", "type": "str", "required": True},
            {"field": "status", "header": "状态(1/0)", "type": "bool", "required": False},
        ],
    },
    "year-mappings": {
        "model": YearMapping,
        "label": "年份映射",
        "unique_field": "code",
        "columns": [
            {"field": "code", "header": "代码字母", "type": "str", "required": True},
            {"field": "year_value", "header": "年份", "type": "int", "required": True},
            {"field": "status", "header": "状态(1/0)", "type": "bool", "required": False},
        ],
    },
    "month-mappings": {
        "model": MonthMapping,
        "label": "月份映射",
        "unique_field": "code",
        "columns": [
            {"field": "code", "header": "代码字母", "type": "str", "required": True},
            {"field": "month_value", "header": "月份(1-12)", "type": "int", "required": True},
            {"field": "status", "header": "状态(1/0)", "type": "bool", "required": False},
        ],
    },
    "grade-mappings": {
        "model": GradeMapping,
        "label": "等级映射",
        "unique_field": "code",
        "columns": [
            {"field": "code", "header": "代码字母", "type": "str", "required": True},
            {"field": "grade_desc", "header": "等级描述", "type": "str", "required": True},
            {"field": "status", "header": "状态(1/0)", "type": "bool", "required": False},
        ],
    },
    "suffix-mappings": {
        "model": SuffixMapping,
        "label": "后缀映射",
        "unique_field": "suffix_code",
        "columns": [
            {"field": "suffix_code", "header": "后缀代码", "type": "str", "required": True},
            {"field": "status_desc", "header": "状态描述", "type": "str", "required": True},
            {"field": "status", "header": "状态(1/0)", "type": "bool", "required": False},
        ],
    },
}

HEADER_FILL = PatternFill("solid", fgColor="2D5FA3")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _build_excel(reg: dict, rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = reg["label"]
    columns = reg["columns"]

    ws.append([c["header"] for c in columns])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = 22

    for row in rows:
        data = []
        for col in columns:
            val = getattr(row, col["field"], None)
            if col["type"] == "bool":
                val = 1 if val else 0
            elif val is None:
                val = ""
            data.append(val)
        ws.append(data)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _parse_excel_rows(file_bytes: bytes, columns: list) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header_row = next(rows_iter, None)
    if header_row is None:
        return [], ["文件为空"]

    expected_headers = [c["header"] for c in columns]
    actual_headers = [str(h).strip() if h is not None else "" for h in header_row]
    if actual_headers[: len(expected_headers)] != expected_headers:
        return [], [f"表头不匹配。期望: {expected_headers}，实际: {actual_headers}"]

    parsed: list[dict] = []
    errors: list[str] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if not any(v not in (None, "") for v in row):
            continue
        record: dict[str, Any] = {}
        for col_idx, col in enumerate(columns):
            raw = row[col_idx] if col_idx < len(row) else None
            if col["required"] and raw in (None, ""):
                errors.append(f"第{row_num}行 [{col['header']}] 不能为空")
                break
            if raw in (None, ""):
                record[col["field"]] = None
            elif col["type"] == "int":
                try:
                    record[col["field"]] = int(raw)
                except (ValueError, TypeError):
                    errors.append(f"第{row_num}行 [{col['header']}] 须为整数，实际值: {raw}")
                    break
            elif col["type"] == "bool":
                text = str(raw).strip().lower()
                if text in {"1", "true", "yes", "y"}:
                    record[col["field"]] = True
                elif text in {"0", "false", "no", "n"}:
                    record[col["field"]] = False
                else:
                    errors.append(f"第{row_num}行 [{col['header']}] 须为 1 或 0，实际值: {raw}")
                    break
            else:
                record[col["field"]] = str(raw).strip()
        else:
            parsed.append(record)
    return parsed, errors


@app.get("/api/v1/dictionaries/{dict_key}/export")
def export_dict(dict_key: str, db: Session = Depends(get_db)):
    if dict_key not in DICT_REGISTRY:
        raise HTTPException(status_code=404, detail=f"未知字典类型: {dict_key}")
    reg = DICT_REGISTRY[dict_key]
    rows = db.query(reg["model"]).order_by(reg["model"].id).all()
    file_bytes = _build_excel(reg, rows)
    filename = f"{reg['label']}_{date.today()}.xlsx"
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return Response(
        content=file_bytes,
        media_type=mime,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.post("/api/v1/dictionaries/{dict_key}/import")
async def import_dict(dict_key: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if dict_key not in DICT_REGISTRY:
        raise HTTPException(status_code=404, detail=f"未知字典类型: {dict_key}")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式文件")

    reg = DICT_REGISTRY[dict_key]
    file_bytes = await file.read()
    parsed_rows, errors = _parse_excel_rows(file_bytes, reg["columns"])
    if errors:
        raise HTTPException(status_code=422, detail={"errors": errors})

    model = reg["model"]
    unique_field = reg["unique_field"]
    created_count = 0
    updated_count = 0

    for record in parsed_rows:
        unique_val = record.get(unique_field)
        if unique_val is None:
            continue
        existing = db.query(model).filter(getattr(model, unique_field) == unique_val).first()
        if existing:
            for key, val in record.items():
                if val is not None:
                    setattr(existing, key, val)
            updated_count += 1
        else:
            obj = model(**{k: v for k, v in record.items() if v is not None})
            if hasattr(obj, "status") and obj.status is None:
                obj.status = True
            db.add(obj)
            created_count += 1

    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail="导入时发生唯一性冲突，请检查数据") from exc

    return {"imported": created_count, "updated": updated_count, "total": len(parsed_rows)}


# ── Line Mapping ───────────────────────────────────────────────────────────────

@app.get("/api/v1/dictionaries/line-mappings", response_model=list[LineMappingOut])
def list_line_mappings(enabled_only: bool = Query(False), db: Session = Depends(get_db)):
    query = db.query(LineMapping)
    if enabled_only:
        query = query.filter(LineMapping.status.is_(True))
    return query.order_by(LineMapping.id.desc()).all()


@app.get("/api/v1/dictionaries/line-mappings/{item_id}", response_model=LineMappingOut)
def get_line_mapping(item_id: int, db: Session = Depends(get_db)):
    return get_model_or_404(db, LineMapping, item_id, "生产线映射")


@app.post("/api/v1/dictionaries/line-mappings", response_model=LineMappingOut)
def create_line_mapping(payload: LineMappingCreate, db: Session = Depends(get_db)):
    item = LineMapping(
        code=normalize_code(payload.code),
        line_desc=payload.line_desc.strip(),
        battery_model=payload.battery_model.strip() if payload.battery_model else None,
        status=payload.status,
    )
    db.add(item)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.put("/api/v1/dictionaries/line-mappings/{item_id}", response_model=LineMappingOut)
def update_line_mapping(item_id: int, payload: LineMappingUpdate, db: Session = Depends(get_db)):
    item = get_model_or_404(db, LineMapping, item_id, "生产线映射")
    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"] is not None:
        data["code"] = normalize_code(data["code"])
    if "line_desc" in data and data["line_desc"] is not None:
        data["line_desc"] = data["line_desc"].strip()
    if "battery_model" in data and data["battery_model"] is not None:
        data["battery_model"] = data["battery_model"].strip()
    for key, value in data.items():
        setattr(item, key, value)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.delete("/api/v1/dictionaries/line-mappings/{item_id}", status_code=204)
def delete_line_mapping(item_id: int, db: Session = Depends(get_db)):
    item = get_model_or_404(db, LineMapping, item_id, "生产线映射")
    db.delete(item)
    commit_or_400(db)
    return Response(status_code=204)


# ── Defect Types ───────────────────────────────────────────────────────────────

@app.get("/api/v1/dictionaries/defect-types", response_model=list[DefectTypeOut])
def list_defect_types(enabled_only: bool = Query(False), db: Session = Depends(get_db)):
    query = db.query(DefectType)
    if enabled_only:
        query = query.filter(DefectType.status.is_(True))
    return query.order_by(DefectType.id.desc()).all()


@app.get("/api/v1/dictionaries/defect-types/{item_id}", response_model=DefectTypeOut)
def get_defect_type(item_id: int, db: Session = Depends(get_db)):
    return get_model_or_404(db, DefectType, item_id, "不良类型")


@app.post("/api/v1/dictionaries/defect-types", response_model=DefectTypeOut)
def create_defect_type(payload: DefectTypeCreate, db: Session = Depends(get_db)):
    item = DefectType(name=payload.name.strip(), status=payload.status)
    db.add(item)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.put("/api/v1/dictionaries/defect-types/{item_id}", response_model=DefectTypeOut)
def update_defect_type(item_id: int, payload: DefectTypeUpdate, db: Session = Depends(get_db)):
    item = get_model_or_404(db, DefectType, item_id, "不良类型")
    data = payload.model_dump(exclude_unset=True)
    if "name" in data and data["name"] is not None:
        data["name"] = data["name"].strip()
    for key, value in data.items():
        setattr(item, key, value)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.delete("/api/v1/dictionaries/defect-types/{item_id}", status_code=204)
def delete_defect_type(item_id: int, db: Session = Depends(get_db)):
    item = get_model_or_404(db, DefectType, item_id, "不良类型")
    db.delete(item)
    commit_or_400(db)
    return Response(status_code=204)


# ── Year Mapping ───────────────────────────────────────────────────────────────

@app.get("/api/v1/dictionaries/year-mappings", response_model=list[YearMappingOut])
def list_year_mappings(enabled_only: bool = Query(False), db: Session = Depends(get_db)):
    query = db.query(YearMapping)
    if enabled_only:
        query = query.filter(YearMapping.status.is_(True))
    return query.order_by(YearMapping.id.desc()).all()


@app.get("/api/v1/dictionaries/year-mappings/{item_id}", response_model=YearMappingOut)
def get_year_mapping(item_id: int, db: Session = Depends(get_db)):
    return get_model_or_404(db, YearMapping, item_id, "年份映射")


@app.post("/api/v1/dictionaries/year-mappings", response_model=YearMappingOut)
def create_year_mapping(payload: YearMappingCreate, db: Session = Depends(get_db)):
    item = YearMapping(code=normalize_code(payload.code), year_value=payload.year_value, status=payload.status)
    db.add(item)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.put("/api/v1/dictionaries/year-mappings/{item_id}", response_model=YearMappingOut)
def update_year_mapping(item_id: int, payload: YearMappingUpdate, db: Session = Depends(get_db)):
    item = get_model_or_404(db, YearMapping, item_id, "年份映射")
    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"] is not None:
        data["code"] = normalize_code(data["code"])
    for key, value in data.items():
        setattr(item, key, value)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.delete("/api/v1/dictionaries/year-mappings/{item_id}", status_code=204)
def delete_year_mapping(item_id: int, db: Session = Depends(get_db)):
    item = get_model_or_404(db, YearMapping, item_id, "年份映射")
    db.delete(item)
    commit_or_400(db)
    return Response(status_code=204)


# ── Month Mapping ──────────────────────────────────────────────────────────────

@app.get("/api/v1/dictionaries/month-mappings", response_model=list[MonthMappingOut])
def list_month_mappings(enabled_only: bool = Query(False), db: Session = Depends(get_db)):
    query = db.query(MonthMapping)
    if enabled_only:
        query = query.filter(MonthMapping.status.is_(True))
    return query.order_by(MonthMapping.id.desc()).all()


@app.get("/api/v1/dictionaries/month-mappings/{item_id}", response_model=MonthMappingOut)
def get_month_mapping(item_id: int, db: Session = Depends(get_db)):
    return get_model_or_404(db, MonthMapping, item_id, "月份映射")


@app.post("/api/v1/dictionaries/month-mappings", response_model=MonthMappingOut)
def create_month_mapping(payload: MonthMappingCreate, db: Session = Depends(get_db)):
    item = MonthMapping(code=normalize_code(payload.code), month_value=payload.month_value, status=payload.status)
    db.add(item)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.put("/api/v1/dictionaries/month-mappings/{item_id}", response_model=MonthMappingOut)
def update_month_mapping(item_id: int, payload: MonthMappingUpdate, db: Session = Depends(get_db)):
    item = get_model_or_404(db, MonthMapping, item_id, "月份映射")
    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"] is not None:
        data["code"] = normalize_code(data["code"])
    for key, value in data.items():
        setattr(item, key, value)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.delete("/api/v1/dictionaries/month-mappings/{item_id}", status_code=204)
def delete_month_mapping(item_id: int, db: Session = Depends(get_db)):
    item = get_model_or_404(db, MonthMapping, item_id, "月份映射")
    db.delete(item)
    commit_or_400(db)
    return Response(status_code=204)


# ── Grade Mapping ──────────────────────────────────────────────────────────────

@app.get("/api/v1/dictionaries/grade-mappings", response_model=list[GradeMappingOut])
def list_grade_mappings(enabled_only: bool = Query(False), db: Session = Depends(get_db)):
    query = db.query(GradeMapping)
    if enabled_only:
        query = query.filter(GradeMapping.status.is_(True))
    return query.order_by(GradeMapping.id.desc()).all()


@app.get("/api/v1/dictionaries/grade-mappings/{item_id}", response_model=GradeMappingOut)
def get_grade_mapping(item_id: int, db: Session = Depends(get_db)):
    return get_model_or_404(db, GradeMapping, item_id, "等级映射")


@app.post("/api/v1/dictionaries/grade-mappings", response_model=GradeMappingOut)
def create_grade_mapping(payload: GradeMappingCreate, db: Session = Depends(get_db)):
    item = GradeMapping(code=normalize_code(payload.code), grade_desc=payload.grade_desc.strip(), status=payload.status)
    db.add(item)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.put("/api/v1/dictionaries/grade-mappings/{item_id}", response_model=GradeMappingOut)
def update_grade_mapping(item_id: int, payload: GradeMappingUpdate, db: Session = Depends(get_db)):
    item = get_model_or_404(db, GradeMapping, item_id, "等级映射")
    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"] is not None:
        data["code"] = normalize_code(data["code"])
    if "grade_desc" in data and data["grade_desc"] is not None:
        data["grade_desc"] = data["grade_desc"].strip()
    for key, value in data.items():
        setattr(item, key, value)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.delete("/api/v1/dictionaries/grade-mappings/{item_id}", status_code=204)
def delete_grade_mapping(item_id: int, db: Session = Depends(get_db)):
    item = get_model_or_404(db, GradeMapping, item_id, "等级映射")
    db.delete(item)
    commit_or_400(db)
    return Response(status_code=204)


# ── Suffix Mapping ─────────────────────────────────────────────────────────────

@app.get("/api/v1/dictionaries/suffix-mappings", response_model=list[SuffixMappingOut])
def list_suffix_mappings(enabled_only: bool = Query(False), db: Session = Depends(get_db)):
    query = db.query(SuffixMapping)
    if enabled_only:
        query = query.filter(SuffixMapping.status.is_(True))
    return query.order_by(SuffixMapping.id.desc()).all()


@app.get("/api/v1/dictionaries/suffix-mappings/{item_id}", response_model=SuffixMappingOut)
def get_suffix_mapping(item_id: int, db: Session = Depends(get_db)):
    return get_model_or_404(db, SuffixMapping, item_id, "后缀映射")


@app.post("/api/v1/dictionaries/suffix-mappings", response_model=SuffixMappingOut)
def create_suffix_mapping(payload: SuffixMappingCreate, db: Session = Depends(get_db)):
    item = SuffixMapping(
        suffix_code=payload.suffix_code.strip().upper(),
        status_desc=payload.status_desc.strip(),
        status=payload.status,
    )
    db.add(item)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.put("/api/v1/dictionaries/suffix-mappings/{item_id}", response_model=SuffixMappingOut)
def update_suffix_mapping(item_id: int, payload: SuffixMappingUpdate, db: Session = Depends(get_db)):
    item = get_model_or_404(db, SuffixMapping, item_id, "后缀映射")
    data = payload.model_dump(exclude_unset=True)
    if "suffix_code" in data and data["suffix_code"] is not None:
        data["suffix_code"] = data["suffix_code"].strip().upper()
    if "status_desc" in data and data["status_desc"] is not None:
        data["status_desc"] = data["status_desc"].strip()
    for key, value in data.items():
        setattr(item, key, value)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.delete("/api/v1/dictionaries/suffix-mappings/{item_id}", status_code=204)
def delete_suffix_mapping(item_id: int, db: Session = Depends(get_db)):
    item = get_model_or_404(db, SuffixMapping, item_id, "后缀映射")
    db.delete(item)
    commit_or_400(db)
    return Response(status_code=204)


# ── Parse / Records / OCR ──────────────────────────────────────────────────────

@app.post("/api/v1/parse", response_model=ParseResponse)
def parse_codes(payload: ParseRequest, db: Session = Depends(get_db)):
    try:
        parsed = parse_battery_codes(db, payload.upper_code, payload.lower_code)
    except ParseCodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return ParseResponse(**parsed)


@app.post("/api/v1/quality-records", response_model=QualityRecordOut)
def create_quality_record(payload: QualityRecordCreate, db: Session = Depends(get_db)):
    record = _create_quality_record_entity(
        db=db,
        detected_date=payload.detected_date,
        upper_code=payload.upper_code,
        lower_code=payload.lower_code,
        defect_type_id=payload.defect_type_id,
        defect_description=payload.defect_description,
        operator_name=payload.operator_name,
    )
    return build_record_out(record)


@app.post("/api/v1/quality-records/upload", response_model=QualityRecordOut)
def create_quality_record_with_photo(
    detected_date: date = Form(...),
    upper_code: str = Form(...),
    lower_code: str = Form(...),
    defect_type_id: int = Form(...),
    defect_description: str | None = Form(None),
    operator_name: str = Form(...),
    photo: UploadFile | None = File(None),
    db: Session = Depends(get_db),
):
    photo_url = _save_photo(photo)
    try:
        record = _create_quality_record_entity(
            db=db,
            detected_date=detected_date,
            upper_code=upper_code,
            lower_code=lower_code,
            defect_type_id=defect_type_id,
            defect_description=defect_description,
            operator_name=operator_name,
            photo_url=photo_url,
        )
    except Exception:
        _remove_photo(photo_url)
        raise
    return build_record_out(record)


@app.get("/api/v1/quality-records", response_model=list[QualityRecordOut])
def list_quality_records(
    start_date: date | None = None,
    end_date: date | None = None,
    defect_type_ids: str | None = Query(default=None, description="逗号分隔，例如 1,2"),
    line_codes: str | None = Query(default=None, description="逗号分隔，例如 V1,V2"),
    battery_models: str | None = Query(default=None, description="逗号分隔"),
    station_nos: str | None = Query(default=None, description="逗号分隔，例如 01,02"),
    keyword: str | None = Query(default=None, description="模糊搜索关键词"),
    sort_by: str = Query(default="id"),
    sort_order: str = Query(default="desc"),
    limit: int = Query(default=1000, ge=1, le=10000),
    db: Session = Depends(get_db),
):
    defect_ids = _csv_ints(defect_type_ids)
    lines = _csv_strings(line_codes)
    models = _csv_strings(battery_models)
    stations = _csv_strings(station_nos)

    query = db.query(QualityRecord).join(DefectType, QualityRecord.defect_type_id == DefectType.id)
    query = _apply_record_filters(
        query,
        start_date=start_date,
        end_date=end_date,
        defect_type_ids=defect_ids,
        line_codes=lines,
        battery_models=models,
        station_nos=stations,
        keyword=keyword,
    )

    sortable_map = {
        "id": QualityRecord.id,
        "record_time": QualityRecord.record_time,
        "detected_date": QualityRecord.detected_date,
        "parsed_production_time": QualityRecord.parsed_production_time,
        "line_code": QualityRecord.parsed_line_code,
        "battery_model": QualityRecord.parsed_battery_model,
        "station_no": QualityRecord.parsed_station_no,
    }
    sort_col = sortable_map.get(sort_by, QualityRecord.id)
    query = query.order_by(sort_col if sort_order == "asc" else desc(sort_col))
    rows = query.limit(limit).all()
    return [build_record_out(r) for r in rows]


@app.delete("/api/v1/quality-records/{record_id}", status_code=204)
def delete_quality_record(record_id: int, db: Session = Depends(get_db)):
    record = get_model_or_404(db, QualityRecord, record_id, "质量记录")
    _remove_photo(record.photo_url)
    db.delete(record)
    commit_or_400(db)
    return Response(status_code=204)


@app.get("/api/v1/quality-records/filter-options")
def quality_record_filter_options(db: Session = Depends(get_db)):
    defect_types = db.query(DefectType).order_by(DefectType.name).all()
    line_mappings = db.query(LineMapping).order_by(LineMapping.code).all()
    station_rows = (
        db.query(QualityRecord.parsed_station_no)
        .filter(QualityRecord.parsed_station_no.isnot(None))
        .group_by(QualityRecord.parsed_station_no)
        .order_by(QualityRecord.parsed_station_no)
        .all()
    )
    return {
        "defect_types": [{"id": d.id, "name": d.name} for d in defect_types],
        "line_codes": [{"code": l.code, "label": l.line_desc, "battery_model": l.battery_model} for l in line_mappings],
        "battery_models": sorted({l.battery_model for l in line_mappings if l.battery_model}),
        "station_nos": [r[0] for r in station_rows],
    }


@app.post("/api/v1/ocr/ingest", response_model=OCRIngestResponse)
def ingest_ocr(payload: OCRIngestRequest, db: Session = Depends(get_db)):
    try:
        parsed = parse_battery_codes(db, payload.upper_code, payload.lower_code)
    except ParseCodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    saved_record_id = None
    if payload.auto_save:
        if not payload.defect_type_id or not payload.operator_name:
            raise HTTPException(status_code=400, detail="auto_save=true 时必须传 defect_type_id 和 operator_name")
        record = _create_quality_record_entity(
            db=db,
            detected_date=payload.detected_date or date.today(),
            upper_code=payload.upper_code,
            lower_code=payload.lower_code,
            defect_type_id=payload.defect_type_id,
            operator_name=payload.operator_name,
        )
        saved_record_id = record.id

    return OCRIngestResponse(parsed=ParseResponse(**parsed), saved_record_id=saved_record_id)


# ── Production output records ──────────────────────────────────────────────────

@app.get("/api/v1/production-outputs", response_model=list[ProductionOutputOut])
def list_production_outputs(
    year: int | None = Query(default=None, ge=2000, le=2100),
    month: int | None = Query(default=None, ge=1, le=12),
    line_codes: str | None = Query(default=None),
    battery_models: str | None = Query(default=None),
    keyword: str | None = Query(default=None),
    sort_by: str = Query(default="year"),
    sort_order: str = Query(default="desc"),
    limit: int = Query(default=500, ge=1, le=5000),
    db: Session = Depends(get_db),
):
    lines = _csv_strings(line_codes)
    models = _csv_strings(battery_models)
    query = db.query(ProductionOutput)
    if year:
        query = query.filter(ProductionOutput.year == year)
    if month:
        query = query.filter(ProductionOutput.month == month)
    if lines:
        query = query.filter(ProductionOutput.line_code.in_(lines))
    if models:
        query = query.filter(ProductionOutput.battery_model.in_(models))
    if keyword:
        kw = f"%{keyword.strip()}%"
        query = query.filter(
            or_(
                ProductionOutput.line_code.ilike(kw),
                ProductionOutput.line_desc.ilike(kw),
                ProductionOutput.battery_model.ilike(kw),
                ProductionOutput.note.ilike(kw),
            )
        )

    sortable_map = {
        "id": ProductionOutput.id,
        "year": ProductionOutput.year,
        "month": ProductionOutput.month,
        "line_code": ProductionOutput.line_code,
        "output_qty": ProductionOutput.output_qty,
        "updated_at": ProductionOutput.updated_at,
    }
    sort_col = sortable_map.get(sort_by, ProductionOutput.year)
    query = query.order_by(sort_col if sort_order == "asc" else desc(sort_col), desc(ProductionOutput.id))
    return query.limit(limit).all()


@app.post("/api/v1/production-outputs", response_model=ProductionOutputOut)
def create_production_output(payload: ProductionOutputCreate, db: Session = Depends(get_db)):
    line_code = normalize_code(payload.line_code)
    line_desc = payload.line_desc.strip() if payload.line_desc else None
    battery_model = payload.battery_model.strip() if payload.battery_model else None
    auto_line_desc, auto_model = _get_line_info(db, line_code)
    if not line_desc:
        line_desc = auto_line_desc
    if not battery_model:
        battery_model = auto_model

    item = ProductionOutput(
        year=payload.year,
        month=payload.month,
        line_code=line_code,
        line_desc=line_desc,
        battery_model=battery_model,
        output_qty=payload.output_qty,
        note=payload.note.strip() if payload.note else None,
    )
    db.add(item)
    commit_or_400(db)
    db.refresh(item)
    return item


@app.put("/api/v1/production-outputs/{item_id}", response_model=ProductionOutputOut)
def update_production_output(item_id: int, payload: ProductionOutputUpdate, db: Session = Depends(get_db)):
    item = get_model_or_404(db, ProductionOutput, item_id, "产量记录")
    data = payload.model_dump(exclude_unset=True)
    if "line_code" in data and data["line_code"] is not None:
        data["line_code"] = normalize_code(data["line_code"])
    if "line_desc" in data and data["line_desc"] is not None:
        data["line_desc"] = data["line_desc"].strip()
    if "battery_model" in data and data["battery_model"] is not None:
        data["battery_model"] = data["battery_model"].strip()
    if "note" in data and data["note"] is not None:
        data["note"] = data["note"].strip()

    for key, value in data.items():
        setattr(item, key, value)

    if item.line_code:
        auto_line_desc, auto_model = _get_line_info(db, item.line_code)
        if not item.line_desc:
            item.line_desc = auto_line_desc
        if not item.battery_model:
            item.battery_model = auto_model

    commit_or_400(db)
    db.refresh(item)
    return item


@app.delete("/api/v1/production-outputs/{item_id}", status_code=204)
def delete_production_output(item_id: int, db: Session = Depends(get_db)):
    item = get_model_or_404(db, ProductionOutput, item_id, "产量记录")
    db.delete(item)
    commit_or_400(db)
    return Response(status_code=204)


# ── Dashboard ──────────────────────────────────────────────────────────────────

@app.get("/api/v1/dashboard/monthly-summary", response_model=list[MonthlySummaryRow])
def monthly_summary(
    year: int = Query(..., ge=2000, le=2100),
    line_codes: str | None = Query(default=None),
    battery_models: str | None = Query(default=None),
    defect_type_ids: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    lines = _csv_strings(line_codes)
    models = _csv_strings(battery_models)
    defect_ids = _csv_ints(defect_type_ids)

    query = (
        db.query(
            extract("month", QualityRecord.parsed_production_time).label("month"),
            QualityRecord.defect_type_id.label("defect_type_id"),
            DefectType.name.label("defect_name"),
            func.count(QualityRecord.id).label("count"),
        )
        .join(DefectType, QualityRecord.defect_type_id == DefectType.id)
        .filter(extract("year", QualityRecord.parsed_production_time) == year)
    )
    if lines:
        query = query.filter(QualityRecord.parsed_line_code.in_(lines))
    if models:
        query = query.filter(QualityRecord.parsed_battery_model.in_(models))
    if defect_ids:
        query = query.filter(QualityRecord.defect_type_id.in_(defect_ids))

    rows = (
        query.group_by(
            extract("month", QualityRecord.parsed_production_time),
            QualityRecord.defect_type_id,
            DefectType.name,
        )
        .order_by(extract("month", QualityRecord.parsed_production_time))
        .all()
    )
    return [
        MonthlySummaryRow(
            month=int(row.month),
            defect_type_id=int(row.defect_type_id),
            defect_name=row.defect_name,
            count=int(row.count),
        )
        for row in rows
    ]


@app.get("/api/v1/dashboard/yearly-summary", response_model=list[YearlySummaryRow])
def yearly_summary(
    start_year: int = Query(..., ge=2000, le=2100),
    end_year: int = Query(..., ge=2000, le=2100),
    line_codes: str | None = Query(default=None),
    battery_models: str | None = Query(default=None),
    defect_type_ids: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    if end_year < start_year:
        raise HTTPException(status_code=400, detail="end_year 不能小于 start_year")

    lines = _csv_strings(line_codes)
    models = _csv_strings(battery_models)
    defect_ids = _csv_ints(defect_type_ids)

    query = (
        db.query(
            extract("year", QualityRecord.parsed_production_time).label("year"),
            QualityRecord.defect_type_id.label("defect_type_id"),
            DefectType.name.label("defect_name"),
            func.count(QualityRecord.id).label("count"),
        )
        .join(DefectType, QualityRecord.defect_type_id == DefectType.id)
        .filter(extract("year", QualityRecord.parsed_production_time) >= start_year)
        .filter(extract("year", QualityRecord.parsed_production_time) <= end_year)
    )
    if lines:
        query = query.filter(QualityRecord.parsed_line_code.in_(lines))
    if models:
        query = query.filter(QualityRecord.parsed_battery_model.in_(models))
    if defect_ids:
        query = query.filter(QualityRecord.defect_type_id.in_(defect_ids))

    rows = (
        query.group_by(
            extract("year", QualityRecord.parsed_production_time),
            QualityRecord.defect_type_id,
            DefectType.name,
        )
        .order_by(extract("year", QualityRecord.parsed_production_time))
        .all()
    )
    return [
        YearlySummaryRow(
            year=int(row.year),
            defect_type_id=int(row.defect_type_id),
            defect_name=row.defect_name,
            count=int(row.count),
        )
        for row in rows
    ]


@app.get("/api/v1/dashboard/monthly-ppm", response_model=list[MonthlyPpmRow])
def monthly_ppm(
    year: int = Query(..., ge=2000, le=2100),
    line_codes: str | None = Query(default=None),
    battery_models: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    lines = _csv_strings(line_codes)
    models = _csv_strings(battery_models)

    output_query = db.query(ProductionOutput).filter(ProductionOutput.year == year)
    if lines:
        output_query = output_query.filter(ProductionOutput.line_code.in_(lines))
    if models:
        output_query = output_query.filter(ProductionOutput.battery_model.in_(models))
    outputs = output_query.all()

    defect_query = (
        db.query(
            extract("month", QualityRecord.parsed_production_time).label("month"),
            QualityRecord.parsed_line_code.label("line_code"),
            QualityRecord.parsed_battery_model.label("battery_model"),
            func.count(QualityRecord.id).label("defect_count"),
        )
        .filter(extract("year", QualityRecord.parsed_production_time) == year)
    )
    if lines:
        defect_query = defect_query.filter(QualityRecord.parsed_line_code.in_(lines))
    if models:
        defect_query = defect_query.filter(QualityRecord.parsed_battery_model.in_(models))
    defect_rows = (
        defect_query.group_by(
            extract("month", QualityRecord.parsed_production_time),
            QualityRecord.parsed_line_code,
            QualityRecord.parsed_battery_model,
        )
        .all()
    )

    defect_map: dict[tuple[int, str, str | None], int] = {}
    for row in defect_rows:
        key = (int(row.month), row.line_code, row.battery_model)
        defect_map[key] = int(row.defect_count)

    result: list[MonthlyPpmRow] = []
    for item in outputs:
        key = (item.month, item.line_code, item.battery_model)
        defect_count = defect_map.get(key, 0)
        ppm = 0.0 if item.output_qty <= 0 else defect_count * 1_000_000 / item.output_qty
        result.append(
            MonthlyPpmRow(
                year=item.year,
                month=item.month,
                line_code=item.line_code,
                line_desc=item.line_desc,
                battery_model=item.battery_model,
                defect_count=defect_count,
                output_qty=item.output_qty,
                ppm=_round_ppm(ppm),
            )
        )

    result.sort(key=lambda x: (x.month, x.line_code))
    return result


@app.get("/api/v1/dashboard/range-summary")
def range_summary(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    line_codes: str | None = Query(default=None),
    battery_models: str | None = Query(default=None),
    defect_type_ids: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """Aggregate defects by detected_date (year-month) within a date range."""
    lines = _csv_strings(line_codes)
    models_list = _csv_strings(battery_models)
    defect_ids = _csv_ints(defect_type_ids)

    query = (
        db.query(
            QualityRecord.detected_date.label("detected_date"),
            QualityRecord.defect_type_id.label("defect_type_id"),
            DefectType.name.label("defect_name"),
            func.count(QualityRecord.id).label("count"),
        )
        .join(DefectType, QualityRecord.defect_type_id == DefectType.id)
        .filter(QualityRecord.detected_date >= start_date)
        .filter(QualityRecord.detected_date <= end_date)
    )
    if lines:
        query = query.filter(QualityRecord.parsed_line_code.in_(lines))
    if models_list:
        query = query.filter(QualityRecord.parsed_battery_model.in_(models_list))
    if defect_ids:
        query = query.filter(QualityRecord.defect_type_id.in_(defect_ids))

    rows = (
        query.group_by(
            QualityRecord.detected_date,
            QualityRecord.defect_type_id,
            DefectType.name,
        )
        .order_by(QualityRecord.detected_date)
        .all()
    )

    agg: dict[tuple, int] = {}
    for row in rows:
        ym = str(row.detected_date)[:7]
        key = (ym, int(row.defect_type_id), row.defect_name)
        agg[key] = agg.get(key, 0) + int(row.count)

    return [
        {"year_month": k[0], "defect_type_id": k[1], "defect_name": k[2], "count": v}
        for k, v in sorted(agg.items())
    ]


@app.get("/api/v1/dashboard/range-ppm")
def range_ppm(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    line_codes: str | None = Query(default=None),
    battery_models: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """Compute PPM per line per month within a date range using detected_date."""
    lines = _csv_strings(line_codes)
    models_list = _csv_strings(battery_models)

    start_ym = start_date[:7]
    end_ym = end_date[:7]
    start_year = int(start_date[:4])
    end_year = int(end_date[:4])

    output_query = db.query(ProductionOutput).filter(
        ProductionOutput.year >= start_year,
        ProductionOutput.year <= end_year,
    )
    if lines:
        output_query = output_query.filter(ProductionOutput.line_code.in_(lines))
    if models_list:
        output_query = output_query.filter(ProductionOutput.battery_model.in_(models_list))
    outputs = output_query.all()

    defect_query = (
        db.query(
            QualityRecord.detected_date.label("detected_date"),
            QualityRecord.parsed_line_code.label("line_code"),
            QualityRecord.parsed_battery_model.label("battery_model"),
            func.count(QualityRecord.id).label("defect_count"),
        )
        .filter(QualityRecord.detected_date >= start_date)
        .filter(QualityRecord.detected_date <= end_date)
    )
    if lines:
        defect_query = defect_query.filter(QualityRecord.parsed_line_code.in_(lines))
    if models_list:
        defect_query = defect_query.filter(QualityRecord.parsed_battery_model.in_(models_list))

    defect_rows = (
        defect_query.group_by(
            QualityRecord.detected_date,
            QualityRecord.parsed_line_code,
            QualityRecord.parsed_battery_model,
        ).all()
    )

    defect_map: dict[tuple, int] = {}
    for row in defect_rows:
        ym = str(row.detected_date)[:7]
        key = (ym, row.line_code or "", row.battery_model or "")
        defect_map[key] = defect_map.get(key, 0) + int(row.defect_count)

    result = []
    for item in outputs:
        ym = f"{item.year:04d}-{item.month:02d}"
        if start_ym <= ym <= end_ym:
            key = (ym, item.line_code, item.battery_model or "")
            defect_count = defect_map.get(key, 0)
            ppm = 0.0 if item.output_qty <= 0 else defect_count * 1_000_000 / item.output_qty
            result.append({
                "year_month": ym,
                "line_code": item.line_code,
                "line_desc": item.line_desc,
                "battery_model": item.battery_model,
                "defect_count": defect_count,
                "output_qty": item.output_qty,
                "ppm": _round_ppm(ppm),
            })

    return sorted(result, key=lambda x: (x["year_month"], x["line_code"]))


# ---------------------------------------------------------------------------
# SPA Frontend Static Serving (Production Mode)
# Serves the built Vue3 app from frontend/dist/
# API routes (/api/..., /uploads/...) take priority; all others → index.html
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("QC_PORT", "8002"))
    uvicorn.run("app.main:app", host="127.0.0.1", port=port, reload=False)
